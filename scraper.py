"""
NEMWeb scraper - concurrent fetches, Origin assets, fuel mix via OpenNEM.

AEMO CSV format:
  C rows = comments
  I rows = headers: [I, TABLE, SUBTABLE, VERSION, col1, col2, ...]
  D rows = data:    [D, TABLE, SUBTABLE, VERSION, val1, val2, ...]
  Trailing comma means last field is always empty - we handle by slicing from index 4.
"""

import re
import csv
import io
import logging
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from typing import Optional
from zoneinfo import ZoneInfo

import requests

logger = logging.getLogger(__name__)

AEST = ZoneInfo("Australia/Brisbane")  # UTC+10 fixed - AEMO never uses daylight saving
NEMWEB_BASE       = "https://www.nemweb.com.au"
DISPATCH_IS_URL   = f"{NEMWEB_BASE}/REPORTS/CURRENT/DispatchIS_Reports/"
PREDISPATCH_URL   = f"{NEMWEB_BASE}/REPORTS/CURRENT/PredispatchIS_Reports/"
SCADA_URL         = f"{NEMWEB_BASE}/REPORTS/CURRENT/Dispatch_SCADA/"
TRADING_IS_URL    = f"{NEMWEB_BASE}/REPORTS/CURRENT/TradingIS_Reports/"
TRADING_CURRENT   = f"{NEMWEB_BASE}/REPORTS/CURRENT/TradingIS_Reports/"
TRADING_ARCHIVE   = f"{NEMWEB_BASE}/REPORTS/ARCHIVE/TradingIS_Reports/"
MTPASA_DUID_URL   = f"{NEMWEB_BASE}/REPORTS/CURRENT/MTPASA_DUIDAvailability/"
PDPASA_DUID_URL   = f"{NEMWEB_BASE}/REPORTS/CURRENT/PDPASA_DUIDAvailability/"
STPASA_DUID_URL   = f"{NEMWEB_BASE}/REPORTS/CURRENT/STPASA_DUIDAvailability/"
ST_PASA_URL       = f"{NEMWEB_BASE}/REPORTS/CURRENT/Short_Term_PASA_Reports/"
OPENNEM_API       = "https://api.opennem.org.au"
AEMO_REG_LIST_URL = "https://www.aemo.com.au/-/media/Files/Electricity/NEM/Participant_Information/Current-Participants/NEM-Registration-and-Exemption-List.xls"
# Additional registration list URL variants to try
AEMO_REG_LIST_URLS = [
    "https://aemo.com.au/-/media/files/electricity/nem/participant_information/current-participants/nem-registration-and-exemption-list.xls",
    "https://www.aemo.com.au/-/media/Files/Electricity/NEM/Participant_Information/Current-Participants/NEM-Registration-and-Exemption-List.xls",
    "https://aemo.com.au/-/media/Files/Electricity/NEM/Participant_Information/Current-Participants/Generators-and-Scheduled-Loads.xls",
]

NEM_REGIONS = ["QLD1", "NSW1", "VIC1", "SA1", "TAS1"]

# ---------------------------------------------------------------------------
# Static unit registry - loaded once from nem_units.json at startup
# ---------------------------------------------------------------------------
import json as _json
from pathlib import Path as _Path

_AEMO_REG_URL = "https://www.aemo.com.au/-/media/files/electricity/nem/participant_information/nem-registration-and-exemption-list.xlsx"

# Fuel type mapping from AEMO's "Fuel Source - Primary" and "Technology Type" columns
_FUEL_MAP = {
    # Coal
    "black coal": "Black Coal",
    "coal": "Black Coal",
    "brown coal": "Brown Coal",
    "lignite": "Brown Coal",
    # Gas
    "natural gas": "Gas",
    "gas": "Gas",
    "coal seam methane": "Gas",
    "coal seam gas": "Gas",
    "landfill methane gas": "Gas",
    "biogas": "Gas",
    "waste coal mine gas": "Gas",
    # Hydro
    "water": "Hydro",
    "hydro": "Hydro",
    "pumped hydro": "Hydro",
    # Wind
    "wind": "Wind",
    # Solar
    "solar": "Solar",
    "solar thermal": "Solar",
    # Battery
    "battery storage": "Battery",
    "battery": "Battery",
    # Liquid
    "liquid fuel": "Liquid",
    "diesel": "Liquid",
    "fuel oil": "Liquid",
    # Other
    "biomass": "Other",
    "waste": "Other",
    "kerosene": "Liquid",
}

# DUID prefix/pattern → fuel - used when registry has no entry for a DUID
_DUID_FUEL_PATTERNS = [
    # ── Black Coal ────────────────────────────────────────────────────────────
    (re.compile(r"^(KSG|KOGAN)", re.I),            "Black Coal"),  # Kogan Creek
    (re.compile(r"^SW[1-6]$", re.I),               "Black Coal"),  # Stanwell
    (re.compile(r"^(STAN[-_]?\d)", re.I),           "Black Coal"),  # Stanwell alt
    (re.compile(r"^TARONG\d", re.I),                "Black Coal"),  # Tarong
    (re.compile(r"^TNPS\d", re.I),                  "Black Coal"),  # Tarong North
    (re.compile(r"^MILLMERRAN\d", re.I),             "Black Coal"),  # Millmerran
    (re.compile(r"^GLAD\d", re.I),                   "Black Coal"),  # Gladstone
    (re.compile(r"^CALL", re.I),                     "Black Coal"),  # Callide B/C
    (re.compile(r"^(ERARING|ER0\d)", re.I),          "Black Coal"),  # Eraring
    (re.compile(r"^BW0\d$", re.I),                   "Black Coal"),  # Bayswater
    (re.compile(r"^(VP[56]|VALES)", re.I),           "Black Coal"),  # Vales Point
    (re.compile(r"^(MTPIPER|MT.?PIPER|MP\d$)", re.I),"Black Coal"), # Mt Piper
    (re.compile(r"^(LD0\d|LIDDELL)", re.I),          "Black Coal"),  # Liddell
    (re.compile(r"^(LYA\d|LYB\d)", re.I),            "Brown Coal"),  # Loy Yang A/B
    (re.compile(r"^(LOYYB|LOYANG)", re.I),           "Brown Coal"),  # Loy Yang alt
    (re.compile(r"^(HPS\d|HAZEL)", re.I),            "Brown Coal"),  # Hazelwood (retired)
    (re.compile(r"^(YWPS\d|YALLOURN)", re.I),        "Brown Coal"),  # Yallourn
    (re.compile(r"^(ANGLESEA)", re.I),               "Brown Coal"),  # Anglesea (retired)
    (re.compile(r"^(ENERGY.?BRIX|EBRIX)", re.I),     "Brown Coal"),  # Energy Brix (ret.)

    # ── Gas / CCGT / OCGT ────────────────────────────────────────────────────
    (re.compile(r"^DDPS\d", re.I),                   "Gas"),   # Darling Downs
    (re.compile(r"^BRAEMAR\d", re.I),                "Gas"),   # Braemar
    (re.compile(r"^SWAN_E", re.I),                   "Gas"),   # Swanbank E
    (re.compile(r"^CONDAMINE\d", re.I),              "Gas"),   # Condamine
    (re.compile(r"^OAKEY\d", re.I),                  "Gas"),   # Oakey
    (re.compile(r"^MSTUART\d", re.I),                "Gas"),   # Mt Stuart
    (re.compile(r"^YABULU\d", re.I),                 "Gas"),   # Yabulu
    (re.compile(r"^LOYNB\d", re.I),                  "Gas"),   # Loynton
    (re.compile(r"^COLONGRA\d", re.I),               "Gas"),   # Colongra
    (re.compile(r"^TALLAWARRA", re.I),               "Gas"),   # Tallawarra
    (re.compile(r"^URANQ\d", re.I),                  "Gas"),   # Uranquinty
    (re.compile(r"^HVGTS", re.I),                    "Gas"),   # Hunter Valley GT
    (re.compile(r"^(TORRA|TORRB)\d", re.I),          "Gas"),   # Torrens Island
    (re.compile(r"^PPCCGT", re.I),                   "Gas"),   # Pelican Point
    (re.compile(r"^OSBORNE\d", re.I),                "Gas"),   # Osborne
    (re.compile(r"^AGLSOM", re.I),                   "Gas"),   # Somerton SA
    (re.compile(r"^MINTARO\d", re.I),                "Gas"),   # Mintaro
    (re.compile(r"^DRYCGT\d", re.I),                 "Gas"),   # Dry Creek
    (re.compile(r"^LADBROKE\d", re.I),               "Gas"),   # Ladbroke Grove
    (re.compile(r"^JEERALANG\d", re.I),              "Gas"),   # Jeeralang A
    (re.compile(r"^JLBGT\d", re.I),                  "Gas"),   # Jeeralang B
    (re.compile(r"^(MORTLK|MORTLAKE)\d", re.I),      "Gas"),   # Mortlake
    (re.compile(r"^LAVNORTH", re.I),                 "Gas"),   # Laverton North
    (re.compile(r"^NEWPORT$", re.I),                 "Gas"),   # Newport
    (re.compile(r"^SOMERTON$", re.I),                "Gas"),   # Somerton VIC
    (re.compile(r"^VPGS\d", re.I),                   "Gas"),   # Valley Power
    (re.compile(r"^QPS\d", re.I),                    "Gas"),   # QPS Peakers
    (re.compile(r"^ROMA_\d", re.I),                  "Gas"),   # Roma
    (re.compile(r"^TOWN.?VILLE", re.I),              "Gas"),   # Townsville
    (re.compile(r"^MACKAYGT\d", re.I),               "Gas"),   # Mackay GT
    # Generic patterns for unrecognised gas units
    (re.compile(r"(OCGT|CCGT|_GT\d|GT_\d|GAS)", re.I), "Gas"),

    # ── Hydro ─────────────────────────────────────────────────────────────────
    (re.compile(r"^(KAREEYA|W.?HOE_|WIVENHOE)", re.I),  "Hydro"),
    (re.compile(r"(HYDRO|TUMUT|HUME|MURRAY|DARTM|EILDON|GORDON|POATINA|"
                r"TREVALLYN|JOHN.?BUTTERS|SNOWY|SHOALHAVEN|JINGELLIC|"
                r"BLOWERING|BENDEELA|BATESMAN|DARTM|CETHANA|MACKINTOSH|"
                r"BASTYAN|CATAGUNYA|DEVILS|FISHER|LEMONTHYME|DEVILS_G|"
                r"REPULSE|ROWALLAN|TARRALEAH|TUNGATINAH|LIAPOOTAH)", re.I), "Hydro"),
    # Pumped hydro load DUIDs (negative MW = pumping mode)
    (re.compile(r"(PUMP|_PUMP\d?$|PUMP\d?$)", re.I), "Hydro"),

    # ── Wind ──────────────────────────────────────────────────────────────────
    (re.compile(r"(WF\d|_WF\d?$|WIND|SNOWYWIND)", re.I),  "Wind"),

    # ── Solar ─────────────────────────────────────────────────────────────────
    (re.compile(r"(SOLAR|_SF\d|SF\d$|_PV|PV\d|BUNGALA|TAILEM|DARLINGTON.?PT"
                r"|DARLPNT|FINNSF|BOMEN)", re.I), "Solar"),

    # ── Battery ───────────────────────────────────────────────────────────────
    (re.compile(r"(BATT|BATTERY|_BAT\d?|BAT_|HPR\d|HORNSDALE.?P|BYP|"
                r"BESS|_BESS|BARCABAT|GANNAWARRA|LAKELANDS|WANDOAN)", re.I), "Battery"),

    # ── Liquid (diesel/oil) ───────────────────────────────────────────────────
    (re.compile(r"(DIESEL|DISTILLATE|LIQUID|FUEL.?OIL)", re.I), "Liquid"),
]

def _infer_region_from_duid(duid: str) -> str:
    """Last-resort region inference from DUID name prefix."""
    d = duid.upper()
    # Common patterns: Q=QLD, N=NSW, V=VIC, S=SA, T=TAS prefix in many DUIDs
    _PREFIX = {
        "QLD": "QLD1", "NSW": "NSW1", "VIC": "VIC1", "SA1": "SA1",
        "TAS": "TAS1", "SNO": "NSW1",  # Snowy
    }
    for prefix, region in _PREFIX.items():
        if d.startswith(prefix):
            return region
    # Single-letter prefix used in some DUIDs
    _SINGLE = {"Q": "QLD1", "N": "NSW1", "V": "VIC1", "S": "SA1", "T": "TAS1"}
    if d[0] in _SINGLE:
        return _SINGLE[d[0]]
    return ""


def _infer_fuel_from_duid(duid: str) -> str:
    """Last-resort fuel type inference from DUID name patterns."""
    for pattern, fuel in _DUID_FUEL_PATTERNS:
        if pattern.search(duid):
            return fuel
    return "Other"

def _fetch_aemo_registration() -> dict:
    """
    Download and parse the AEMO NEM Registration and Exemption List XLSX.
    Returns { DUID: { station, fuel, region, capacity } }
    """
    import io
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available for registration fetch")
        return {}

    try:
        logger.info("Fetching AEMO registration list...")
        resp = requests.get(_AEMO_REG_URL, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

        # Find the "Generators and Scheduled Loads" sheet
        sheet = None
        for name in wb.sheetnames:
            if "generator" in name.lower() or "scheduled" in name.lower():
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active

        # Read header row to find column indices
        rows = list(sheet.iter_rows(values_only=True))
        # Find header row (contains "DUID")
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            if row and any(str(c).strip().upper() == "DUID" for c in row if c):
                header_row = [str(c).strip() if c else "" for c in row]
                header_idx = i
                break

        if not header_row:
            logger.warning("Could not find header row in AEMO registration list")
            return {}

        # Map column names to indices (case-insensitive)
        col = {h.lower(): i for i, h in enumerate(header_row)}
        logger.info(f"Registration sheet columns: {list(col.keys())[:20]}")

        # Find key columns
        def find_col(*names):
            for n in names:
                if n in col: return col[n]
            # Partial match
            for k, v in col.items():
                for n in names:
                    if n in k: return v
            return None

        duid_col     = find_col("duid")
        station_col  = find_col("station name", "station", "generating unit name", "name")
        region_col   = find_col("region", "regionid")
        # AEMO XLSX actual column: "Fuel Source - Descriptor" (not "primary")
        fuel_col     = find_col("fuel source - descriptor", "fuel source - primary", "fuel source", "fuel type", "primary fuel", "fuel")
        tech_col     = find_col("technology type - primary", "technology type - descriptor", "technology type", "tech type", "technology")
        capacity_col = find_col("reg cap (mw)", "registered capacity", "capacity", "max cap", "reg cap")

        if duid_col is None:
            logger.warning("Could not find DUID column in registration list")
            return {}

        logger.info(f"Registration columns found - duid:{duid_col} station:{station_col} "
                    f"region:{region_col} fuel:{fuel_col} tech:{tech_col} cap:{capacity_col}")

        result = {}
        for row in rows[header_idx + 1:]:
            if not row or not row[duid_col]:
                continue
            duid = str(row[duid_col]).strip().upper()
            if not duid or duid == "NONE":
                continue

            station  = str(row[station_col]).strip()  if station_col  is not None and row[station_col]  else duid
            region   = str(row[region_col]).strip()   if region_col   is not None and row[region_col]   else ""
            fuel_raw = str(row[fuel_col]).strip().lower() if fuel_col is not None and row[fuel_col] else ""
            tech_raw = str(row[tech_col]).strip().lower() if tech_col is not None and row[tech_col] else ""

            # Normalise region
            if region and not region.endswith("1"):
                region = region + "1"
            if region not in ("QLD1", "NSW1", "VIC1", "SA1", "TAS1"):
                continue  # skip non-NEM regions

            # Normalise fuel
            fuel = None
            for key, val in _FUEL_MAP.items():
                if key in fuel_raw or key in tech_raw:
                    fuel = val
                    break
            # Battery override via tech
            if "battery" in tech_raw or "storage" in tech_raw:
                fuel = "Battery"
            if fuel is None or fuel == "Other":
                # Try to infer from DUID and station name before giving up
                fuel = _infer_fuel_from_duid(duid) or _infer_fuel_from_duid(station.upper().replace(" ", "_"))
                if not fuel:
                    fuel = "Other"

            # Capacity
            cap = None
            if capacity_col is not None and row[capacity_col]:
                try:
                    cap = int(round(float(str(row[capacity_col]).replace(",", ""))))
                except (ValueError, TypeError):
                    pass

            result[duid] = {"station": station, "fuel": fuel, "region": region, "capacity": cap}

        logger.info(f"AEMO registration list: {len(result)} DUIDs parsed")
        return result

    except Exception as e:
        logger.warning(f"AEMO registration fetch failed: {e}")
        return {}


def _load_nem_units() -> dict:
    """Load DUID registry from static JSON only (fast, no network).
    The live AEMO registration list is fetched later during scrape_gen() runs."""
    p = _Path(__file__).parent / "nem_units.json"
    try:
        return _json.loads(p.read_text())
    except Exception as e:
        logger.warning(f"nem_units.json load failed: {e}")
        return {}


NEM_UNITS: dict = _load_nem_units()
logger.info(f"NEM_UNITS loaded: {len(NEM_UNITS)} DUIDs")

# Pump-load DUIDs: registered as scheduled LOAD in AEMO - SCADA reports positive MW
# when consuming (pumping). We negate so they display as negative generation.
PUMP_LOAD_DUIDS: set = {
    "SHPUMP",       # Shoalhaven pump (NSW)
    "WIVENPUMP1",   # Wivenhoe pump 1 (QLD)
    "WIVENPUMP2",   # Wivenhoe pump 2 (QLD)
    "MURRAY_PUMP",  # Murray pump (NSW)
    "BENDEELA_P",   # Bendeela pump (NSW)
    "KANGVALLEY",   # Kangaroo Valley pump (NSW)
}

FUEL_COLORS = {
    "Black Coal":        "#6b7280",
    "Brown Coal":        "#8B4513",
    "Gas":               "#ff9f40",
    "Hydro":             "#36a2eb",
    "Wind":              "#4bc0c0",
    "Solar":             "#ffd700",
    "Rooftop Solar":     "#ffe066",
    "Battery":           "#9b59b6",
    "Battery (charging)":"#a855f7",
    "Pump Hydro":        "#06b6d4",
    "Liquid":            "#e74c3c",
    "Other":             "#95a5a6",
}
ALL_FUELS = list(FUEL_COLORS.keys())

# ---------------------------------------------------------------------------
# Origin Energy assets - DUID -> display info
# These are Origin's registered generating units in the NEM
# ---------------------------------------------------------------------------
# ── Origin Energy asset registry ──────────────────────────────────────────────
# ORIGIN_DUIDS: the set of DUIDs that belong to Origin's portfolio.
# All metadata (station, fuel, region, capacity) comes from nem_units.json / live reg list.
# ORIGIN_DISPLAY_NAMES: overrides the AEMO station name with an Origin-branded name.
# Only add overrides where the AEMO name differs from what we want to show.
ORIGIN_DUIDS: set = {
    # QLD
    "DDPS1", "MSTUART1", "MSTUART2", "MSTUART3", "ROMA_7",
    "CLARESF1", "DDSF1", "DAYDSF1",
    "SNB01",
    # NSW
    "ER01", "ER02", "ER03", "ER04",
    "URANQ1", "URANQ2", "URANQ3", "URANQ4",
    "URANQ11", "URANQ12", "URANQ13", "URANQ14",  # aliases used in some AEMO files
    "SHGEN", "SHPUMP",
    "MOREESF1", "GUNNING1",
    "ERGT01",
    "ERB01",
    # VIC
    "MORTLK11", "MORTLK12",
    "STOCKYD1",
    # SA
    "OSB-AG",
    "QPS1", "QPS2", "QPS3", "QPS4", "QPS5",
    "LADBROK1", "LADBROK2",
    "BNGSF1", "BNGSF2",
    "SNOWNTH1", "SNOWSTH1",
}

# Display name overrides - maps DUID → display station name shown on Origin page.
# Only needed where AEMO's registered name differs from what Origin calls it.
ORIGIN_DISPLAY_NAMES: dict = {
    "SNB01":    "Supernode Battery",
    "OSB-AG":   "Osborne Cogen",
    "ROMA_7":   "Roma Gas",
    "BNGSF1":   "Bungala Solar 1",
    "BNGSF2":   "Bungala Solar 2",
    "GUNNING1": "Gunning Wind Farm",
    "ERB01":    "Eraring Battery",
    "ERGT01":   "Eraring Liquid",
    # Eraring coal
    "ER01":     "Eraring",
    "ER02":     "Eraring",
    "ER03":     "Eraring",
    "ER04":     "Eraring",
    # Shoalhaven
    "SHGEN":    "Shoalhaven",
    "SHPUMP":   "Shoalhaven Pump",
    # Darling Downs gas + solar
    "DDPS1":    "Darling Downs",
    "DDSF1":    "Darling Downs Solar",
    # Daydream
    "DAYDSF1":  "Daydream Solar",
    # Mt Stuart
    "MSTUART1": "Mt Stuart",
    "MSTUART2": "Mt Stuart",
    "MSTUART3": "Mt Stuart",
    # Mortlake
    "MORTLK11": "Mortlake",
    "MORTLK12": "Mortlake",
    # Stockyard Hill
    "STOCKYD1": "Stockyard Hill",
    # Ladbroke Grove
    "LADBROK1": "Ladbroke Grove",
    "LADBROK2": "Ladbroke Grove",
    # Quarantine
    "QPS1":     "Quarantine",
    "QPS2":     "Quarantine",
    "QPS3":     "Quarantine",
    "QPS4":     "Quarantine",
    "QPS5":     "Quarantine",
    # Snowtown
    "SNOWNTH1": "Snowtown 1",
    "SNOWSTH1": "Snowtown 2",
    # Uranquinty — two DUID families, same display name
    "URANQ1":   "Uranquinty",
    "URANQ2":   "Uranquinty",
    "URANQ3":   "Uranquinty",
    "URANQ4":   "Uranquinty",
    "URANQ11":  "Uranquinty",
    "URANQ12":  "Uranquinty",
    "URANQ13":  "Uranquinty",
    "URANQ14":  "Uranquinty",
}


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "NEM-Dashboard/1.0"})
# Cap connection pool to avoid overwhelming NEMWeb with concurrent requests
_adapter = requests.adapters.HTTPAdapter(pool_connections=4, pool_maxsize=8)
SESSION.mount("https://", _adapter)
SESSION.mount("http://", _adapter)

def _get(url: str, timeout: int = 15, retries: int = 2) -> Optional[requests.Response]:
    import time as _time
    for attempt in range(retries + 1):
        try:
            r = SESSION.get(url, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt < retries:
                _time.sleep(0.5 * (attempt + 1))  # 0.5s, 1.0s backoff
                logger.debug(f"GET retry {attempt+1} {url}: {e}")
            else:
                logger.warning(f"GET failed {url}: {e}")
    return None


def _list_hrefs(url: str) -> list[str]:
    r = _get(url, timeout=10)
    if not r:
        return []
    found = []
    for m in re.finditer(r'href="([^"]+\.zip)"', r.text, re.IGNORECASE):
        href = m.group(1)
        if href.startswith("http"):
            # Normalise any absolute URL to use www. and uppercase /REPORTS/
            href = re.sub(r'https?://(?:www\.)?nemweb\.com\.au', NEMWEB_BASE, href, flags=re.I)
            href = re.sub(r'/[Rr]eports/[Cc]urrent/', '/REPORTS/CURRENT/', href)
            href = re.sub(r'/[Rr]eports/[Aa]rchive/', '/REPORTS/ARCHIVE/', href)
            found.append(href)
        elif href.startswith("/"):
            # Normalise path casing
            href = re.sub(r'/[Rr]eports/[Cc]urrent/', '/REPORTS/CURRENT/', href)
            href = re.sub(r'/[Rr]eports/[Aa]rchive/', '/REPORTS/ARCHIVE/', href)
            found.append(f"{NEMWEB_BASE}{href}")
        else:
            found.append(url.rstrip("/") + "/" + href)
    return sorted(set(found))


def _read_zip(url: str) -> str:
    r = _get(url, timeout=30)
    if not r:
        return ""
    try:
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            csvs = [n for n in z.namelist() if n.lower().endswith(".csv")]
            if not csvs:
                return ""
            with z.open(csvs[0]) as f:
                return f.read().decode("utf-8", errors="replace")
    except Exception as e:
        logger.warning(f"ZIP read failed {url}: {e}")
        return ""


def _read_zip_all(url: str) -> str:
    """Read ALL CSVs from a ZIP and concatenate — for multi-file archive ZIPs."""
    r = _get(url, timeout=60)
    if not r:
        return ""
    try:
        parts = []
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            csvs = sorted(n for n in z.namelist() if n.lower().endswith(".csv"))
            if not csvs:
                return ""
            for name in csvs:
                with z.open(name) as f:
                    parts.append(f.read().decode("utf-8", errors="replace"))
        return "\n".join(parts)
    except Exception as e:
        logger.warning(f"ZIP read all failed {url}: {e}")
        return ""


def _read_zip_of_zips(url: str) -> str:
    """
    Read a ZIP that contains inner ZIPs (each containing one CSV).
    Used for TradingIS weekly archive files.
    """
    r = _get(url, timeout=60)
    if not r:
        return ""
    try:
        parts = []
        with zipfile.ZipFile(io.BytesIO(r.content)) as outer:
            inner_zips = sorted(n for n in outer.namelist() if n.lower().endswith(".zip"))
            for name in inner_zips:
                try:
                    inner_bytes = outer.read(name)
                    with zipfile.ZipFile(io.BytesIO(inner_bytes)) as inner:
                        csvs = [n for n in inner.namelist() if n.lower().endswith(".csv")]
                        if csvs:
                            with inner.open(csvs[0]) as f:
                                parts.append(f.read().decode("utf-8", errors="replace"))
                except Exception:
                    continue
        return "\n".join(parts)
    except Exception as e:
        logger.warning(f"ZIP of ZIPs read failed {url}: {e}")
        return ""


# ---------------------------------------------------------------------------
# AEMO CSV parser
# ---------------------------------------------------------------------------

def _parse_aemo(text: str, table_key: str) -> list[dict]:
    """Return data rows for table matching table_key (checked against TABLE_SUBTABLE)."""
    results = []
    headers: list[str] = []
    in_table = False
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if not row:
            continue
        ind = row[0].strip().upper()
        if ind == "I" and len(row) >= 5:
            key = f"{row[1].strip()}_{row[2].strip()}".upper()
            if table_key.upper() in key:
                headers = [c.strip().upper() for c in row[4:] if c.strip()]
                in_table = True
            else:
                in_table = False
        elif ind == "D" and in_table and headers:
            vals = [c.strip() for c in row[4:]]
            vals = vals[:len(headers)]
            while len(vals) < len(headers):
                vals.append("")
            results.append(dict(zip(headers, vals)))
    return results


def get_latest_file_url(directory_url: str, prefix: str = "") -> Optional[str]:
    urls = _list_hrefs(directory_url)
    if prefix:
        urls = [u for u in urls if prefix.lower() in u.lower()]
    return urls[-1] if urls else None


def get_all_file_urls(directory_url: str, prefix: str = "") -> list[str]:
    urls = _list_hrefs(directory_url)
    if prefix:
        urls = [u for u in urls if prefix.lower() in u.lower()]
    return urls


# ---------------------------------------------------------------------------
# Dispatch IS - prices, demand, generation, interconnectors (one file fetch)
# ---------------------------------------------------------------------------

def _fetch_dispatch_is() -> str:
    url = get_latest_file_url(DISPATCH_IS_URL, "PUBLIC_DISPATCHIS")
    return _read_zip(url) if url else ""


def scrape_region_summary(text: str) -> dict:
    summary: dict[str, dict] = {}
    for row in _parse_aemo(text, "DISPATCH_PRICE"):
        region = row.get("REGIONID", "")
        if region not in NEM_REGIONS:
            continue
        if row.get("INTERVENTION", "0") not in ("0", ""):
            continue  # skip intervention runs
        e = summary.setdefault(region, {})
        for f in ["RRP"]:
            v = row.get(f, "")
            if v:
                try: e[f] = round(float(v), 2)
                except ValueError: pass
    for row in _parse_aemo(text, "DISPATCH_REGIONSUM"):
        region = row.get("REGIONID", "")
        if region not in NEM_REGIONS:
            continue
        if row.get("INTERVENTION", "0") not in ("0", ""):
            continue  # skip intervention runs
        e = summary.setdefault(region, {})
        for f in ["TOTALDEMAND","DEMANDFORECAST","INITIALSUPPLY",
                  "DISPATCHABLEGENERATION","DISPATCHABLELOAD","SEMISCHEDULE_CLEAREDMW","NETINTERCHANGE",
                  "DEMAND_AND_NONSCHEDGEN","TOTALINTERMITTENTGENERATION",
                  "BDU_ENERGY_STORAGE","BDU_MAX_AVAIL","BDU_MIN_AVAIL",
                  "BDU_CLEAREDMW_GEN","BDU_CLEAREDMW_LOAD","BDU_INITIAL_ENERGY_STORAGE"]:
            v = row.get(f, "")
            if v:
                try: e[f] = round(float(v), 1)
                except ValueError: pass
    return summary


def scrape_interconnectors(text: str) -> dict:
    flows = {}
    for row in _parse_aemo(text, "DISPATCH_INTERCONNECTORRES"):
        ic = row.get("INTERCONNECTORID", "")
        if not ic:
            continue
        try:
            flows[ic] = {
                "flow":   round(float(row.get("MWFLOW", 0) or 0), 1),
                "losses": round(float(row.get("MWLOSSES", 0) or 0), 1),
            }
        except ValueError:
            pass
    return flows


# ---------------------------------------------------------------------------
# SCADA - per-DUID actual MW output (for Origin assets)
# ---------------------------------------------------------------------------

def scrape_scada_duids(duids: set) -> dict:
    """
    Fetch DISPATCH_UNIT_SCADA and return { duid: mw } for requested DUIDs.
    """
    url = get_latest_file_url(SCADA_URL, "PUBLIC_DISPATCHSCADA")
    if not url:
        # fallback: try DispatchIS UNIT_SOLUTION
        return {}
    text = _read_zip(url)
    result = {}
    for row in _parse_aemo(text, "DISPATCH_UNIT_SCADA"):
        duid = row.get("DUID", "").strip().upper()
        if duid in duids:
            v = row.get("SCADAVALUE", "")
            try:
                result[duid] = round(float(v), 1)
            except (ValueError, TypeError):
                pass
    # Also try UNIT_SOLUTION in DispatchIS if SCADA didn't find them
    return result


def scrape_unit_solution(text: str, duids: set) -> dict:
    """Extract INITIALMW from DISPATCH_UNIT_SOLUTION for given DUIDs."""
    result = {}
    for row in _parse_aemo(text, "DISPATCH_UNIT_SOLUTION"):
        duid = row.get("DUID", "").strip().upper()
        if duid in duids:
            v = row.get("INITIALMW", row.get("TOTALCLEARED", ""))
            try:
                result[duid] = round(float(v), 1)
            except (ValueError, TypeError):
                pass
    return result


# ---------------------------------------------------------------------------
# Historical prices + demand from TradingIS CURRENT
# Each file is ~1.5KB and covers exactly ONE 30-min interval.
# Files are named PUBLIC_TRADINGIS_YYYYMMDDHHMM_<id>.zip
# Contains TRADING_PRICE (actual settled RRP) and TRADING_REGIONSUM (actual demand).
# SETTLEMENTDATE is end-of-interval - subtract 30min to get interval start.
# ---------------------------------------------------------------------------

def scrape_trading_history() -> dict:
    """
    Fetch today's TradingIS CURRENT files for price history.
    Each file covers ONE 30-min interval (~1.5KB).
    TRADING_PRICE uses INVALIDFLAG (not INTERVENTION).
    SETTLEMENTDATE is end-of-interval - subtract 30min.
    Returns { "prices": {region: [{interval, rrp}]} }
    """
    now_aest = datetime.now(AEST)
    today_str = now_aest.strftime("%Y%m%d")

    all_zips = _list_hrefs(TRADING_CURRENT)
    today_zips = sorted([u for u in all_zips if today_str in u])
    if not today_zips:
        today_zips = sorted(all_zips)[-48:]

    # TradingIS has ~5 files per 30-min interval (one per dispatch run).
    # Keep only the LAST file per HHMM timestamp to get the final settled price.
    # Filename format: PUBLIC_TRADINGIS_YYYYMMDDHHММ_<seq>.zip
    seen_hhmm = {}
    for url in today_zips:
        fname = url.split('/')[-1]
        parts = fname.split('_')
        # parts[2] is the datetime stamp e.g. 202603121335
        if len(parts) >= 3 and len(parts[2]) >= 12:
            hhmm = parts[2][8:12]  # extract HHMM
            seen_hhmm[hhmm] = url  # last one wins (sorted ascending)
    today_zips = sorted(seen_hhmm.values())

    prices: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    fetch_ok = 0
    fetch_fail = 0
    fetch_empty = 0
    now_aest   = datetime.now(AEST)
    today_date = now_aest.date()
    now_label  = now_aest.strftime("%H:%M")

    def fetch_one(url):
        import time as _time, random
        _time.sleep(random.uniform(0, 0.1))  # small jitter to avoid burst
        try:
            text = _read_zip(url)
            if not text:
                return [], "empty"
            pts = []
            for row in _parse_aemo(text, "TRADING_PRICE"):
                region = row.get("REGIONID", "")
                if region not in NEM_REGIONS:
                    continue
                # TradingIS uses INVALIDFLAG, not INTERVENTION
                if row.get("INVALIDFLAG", "0") not in ("0", ""):
                    continue
                dt_str = row.get("SETTLEMENTDATE", "")
                rrp_str = row.get("RRP", "")
                if not dt_str or not rrp_str:
                    continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                    # Only keep today's intervals, capped at now
                    if dt.date() != today_date:
                        continue
                    label = dt.strftime("%H:%M")
                    if label > now_label:
                        continue
                    pts.append((region, label, round(float(rrp_str), 2)))
                except (ValueError, TypeError):
                    pass
            return pts, "ok" if pts else "empty"
        except Exception as e:
            logger.warning(f"fetch_one failed {url}: {e}")
            return [], "fail"

    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(fetch_one, u): u for u in today_zips}
        for fut in as_completed(futures):
            pts, status = fut.result()
            if status == "ok":
                fetch_ok += 1
            elif status == "fail":
                fetch_fail += 1
            else:
                fetch_empty += 1
            for region, label, val in pts:
                prices[region][label] = val

    price_result = {r: [{"interval": k, "rrp": v} for k, v in sorted(s.items())]
                    for r, s in prices.items() if s}

    logger.info(f"TradingIS prices: {sum(len(v) for v in price_result.values())} pts "
                f"from {len(today_zips)} files (ok={fetch_ok} empty={fetch_empty} fail={fetch_fail})")
    return {"prices": price_result, "fetch_stats": {"ok": fetch_ok, "empty": fetch_empty, "fail": fetch_fail}}


# ---------------------------------------------------------------------------
# Historical demand from DispatchIS CURRENT (5-min intervals)
# ---------------------------------------------------------------------------

def scrape_dispatch_history() -> dict:
    """
    Fetch today's DispatchIS files for 5-min demand AND price history in one pass.
    Returns {
        "demand": { region: [{interval, demand}] },
        "prices": { region: [{interval, rrp}] }   ← 5-min dispatch prices
    }
    These 5-min dispatch prices fill the ~1hr gap that TradingIS (30-min) lags behind.
    Cap at 300 files.
    """
    now_aest = datetime.now(AEST)
    today_str = now_aest.strftime("%Y%m%d")

    all_zips = _list_hrefs(DISPATCH_IS_URL)
    today_zips = [u for u in all_zips if today_str in u and "PUBLIC_DISPATCHIS" in u.upper()]
    if not today_zips:
        today_zips = all_zips[-288:]
    today_zips = sorted(today_zips)[-300:]

    demand: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    op_demand: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    gen_demand: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    prices: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    ss_solar: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    ss_wind: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    ic_flows: dict[str, dict] = {}   # { ic_id: { label: flow } }
    rooftop: dict[str, dict] = {r: {} for r in NEM_REGIONS}  # TOTALINTERMITTENTGENERATION
    fetch_ok = fetch_fail = fetch_empty = 0
    now_aest   = datetime.now(AEST)
    today_date = now_aest.date()
    now_label  = now_aest.strftime("%H:%M")

    def fetch_one(url):
        import time as _time, random
        _time.sleep(random.uniform(0, 0.05))
        try:
            text = _read_zip(url)
            if not text:
                return [], "empty"
            pts = []
            # Extract demand from DISPATCH_REGIONSUM
            for row in _parse_aemo(text, "DISPATCH_REGIONSUM"):
                region = row.get("REGIONID", "")
                if region not in NEM_REGIONS:
                    continue
                if row.get("INTERVENTION", "0") not in ("0", ""):
                    continue
                dt_str = row.get("SETTLEMENTDATE", "")
                demand_str = row.get("TOTALDEMAND", "")
                op_demand_str = row.get("DEMAND_AND_NONSCHEDGEN", "")
                semisched_str = row.get("SEMISCHEDULE_CLEAREDMW", "")
                if not dt_str or not demand_str:
                    continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=5)
                    if dt.date() != today_date or dt.strftime("%H:%M") > now_label:
                        continue
                    label = dt.strftime("%H:%M")
                    pts.append(("demand", region, label, round(float(demand_str), 1)))
                    if op_demand_str:
                        pts.append(("op_demand", region, label, round(float(op_demand_str), 1)))
                    # gen_demand = TOTALDEMAND + SEMISCHEDULE_CLEAREDMW (sits at top of full supply stack)
                    if semisched_str:
                        try:
                            gen_dem = round(float(demand_str), 1) + round(float(semisched_str), 1)
                            pts.append(("gen_demand", region, label, gen_dem))
                        except (ValueError, TypeError):
                            pass
                    rooftop_str = row.get("SS_SOLAR_CLEAREDMW", "")
                    wind_str    = row.get("SS_WIND_CLEAREDMW", "")
                    if rooftop_str:
                        try:
                            rooftop = round(float(rooftop_str), 1)
                            if rooftop > 0:
                                pts.append(("rooftop", region, label, rooftop))
                            pts.append(("solar", region, label, rooftop))
                        except (ValueError, TypeError):
                            pass
                    if wind_str:
                        try:
                            pts.append(("wind", region, label, round(float(wind_str), 1)))
                        except (ValueError, TypeError):
                            pass
                    # BDU (battery) fields
                    bdu_gen  = row.get("BDU_CLEAREDMW_GEN", "")
                    bdu_load = row.get("BDU_CLEAREDMW_LOAD", "")
                    bdu_soc  = row.get("BDU_ENERGY_STORAGE", "")
                    bdu_cap  = row.get("BDU_MAX_AVAIL", "")
                    disp_load_str = row.get("DISPATCHABLELOAD", "")
                    if bdu_gen or bdu_load:
                        try:
                            g = round(float(bdu_gen or 0), 1)
                            l = round(float(bdu_load or 0), 1)
                            s = round(float(bdu_soc), 1) if bdu_soc else None
                            c = round(float(bdu_cap), 1) if bdu_cap else None
                            dl = round(float(disp_load_str), 1) if disp_load_str else l
                            pump = max(round(dl - l, 1), 0)
                            pts.append(("bdu", region, label, {"net_mw": round(g-l,1), "gen": g, "load": l, "pump_load": pump, "storage": s, "max_avail": c}))
                        except (ValueError, TypeError):
                            pass
                except (ValueError, TypeError):
                    pass
            # Extract price from DISPATCH_PRICE
            for row in _parse_aemo(text, "DISPATCH_PRICE"):
                region = row.get("REGIONID", "")
                if region not in NEM_REGIONS:
                    continue
                if row.get("INTERVENTION", "0") not in ("0", ""):
                    continue
                dt_str = row.get("SETTLEMENTDATE", "")
                rrp_str = row.get("RRP", "")
                if not dt_str or not rrp_str:
                    continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=5)
                    if dt.date() != today_date or dt.strftime("%H:%M") > now_label:
                        continue
                    pts.append(("price", region, dt.strftime("%H:%M"), round(float(rrp_str), 2)))
                except (ValueError, TypeError):
                    pass
            # Extract IC flows from DISPATCH_INTERCONNECTORRES
            for row in _parse_aemo(text, "DISPATCH_INTERCONNECTORRES"):
                ic_id = row.get("INTERCONNECTORID", "").strip()
                if not ic_id:
                    continue
                if row.get("INTERVENTION", "0") not in ("0", ""):
                    continue
                dt_str = row.get("SETTLEMENTDATE", "")
                flow_str = row.get("MWFLOW", "")
                if not dt_str or not flow_str:
                    continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=5)
                    if dt.date() != today_date or dt.strftime("%H:%M") > now_label:
                        continue
                    pts.append(("ic", ic_id, dt.strftime("%H:%M"), round(float(flow_str), 1)))
                except (ValueError, TypeError):
                    pass
            return pts, "ok" if pts else "empty"
        except Exception as e:
            logger.warning(f"dispatch_history fetch_one failed {url}: {e}")
            return [], "fail"

    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(fetch_one, u): u for u in today_zips}
        for fut in as_completed(futures):
            pts, status = fut.result()
            if status == "ok":
                fetch_ok += 1
            elif status == "fail":
                fetch_fail += 1
            else:
                fetch_empty += 1
            for item in pts:
                kind = item[0]
                if kind == "demand":
                    _, region, label, val = item
                    demand[region][label] = val
                elif kind == "op_demand":
                    _, region, label, val = item
                    op_demand[region][label] = val
                elif kind == "gen_demand":
                    _, region, label, val = item
                    gen_demand[region][label] = val
                elif kind == "solar":
                    _, region, label, val = item
                    ss_solar[region][label] = val
                elif kind == "wind":
                    _, region, label, val = item
                    ss_wind[region][label] = val
                elif kind == "ic":
                    _, ic_id, label, val = item
                    if ic_id not in ic_flows:
                        ic_flows[ic_id] = {}
                    ic_flows[ic_id][label] = val
                elif kind == "rooftop":
                    _, region, label, val = item
                    rooftop[region][label] = val
                elif kind == "bdu":
                    _, region, label, val = item
                    _bdu_history[region][label] = val
                else:
                    _, region, label, val = item
                    prices[region][label] = val

    demand_result = {r: [{"interval": k, "demand": v} for k, v in sorted(s.items())]
                     for r, s in demand.items() if s}
    op_demand_result = {r: [{"interval": k, "demand": v} for k, v in sorted(s.items())]
                        for r, s in op_demand.items() if s}
    price_result  = {r: [{"interval": k, "rrp": v}    for k, v in sorted(s.items())]
                     for r, s in prices.items() if s}

    # Backfill _ic_history so IC chart shows full day on startup
    for ic_id, series in ic_flows.items():
        if ic_id not in _ic_history:
            _ic_history[ic_id] = {}
        _ic_history[ic_id].update(series)

    # Store rooftop data globally so scrape_scada_history can merge it in after backfill
    global _rooftop_history
    for region in NEM_REGIONS:
        if region not in _rooftop_history:
            _rooftop_history[region] = {}
        _rooftop_history[region].update(rooftop[region])

    logger.info(f"DispatchIS history: demand={sum(len(v) for v in demand_result.values())} pts, "
                f"prices={sum(len(v) for v in price_result.values())} pts, "
                f"ic_flows={sum(len(v) for v in ic_flows.values())} pts "
                f"from {len(today_zips)} files (ok={fetch_ok} empty={fetch_empty} fail={fetch_fail})")
    solar_result    = {r: [{"interval": k, "mw": v} for k, v in sorted(s.items())]
                        for r, s in ss_solar.items() if s}
    wind_result     = {r: [{"interval": k, "mw": v} for k, v in sorted(s.items())]
                        for r, s in ss_wind.items() if s}
    gen_demand_result = {r: [{"interval": k, "demand": v} for k, v in sorted(s.items())]
                         for r, s in gen_demand.items() if s}
    return {"demand": demand_result, "op_demand": op_demand_result,
            "prices": price_result, "solar": solar_result, "wind": wind_result,
            "gen_demand": gen_demand_result}


# Keep old name as alias for compatibility
def scrape_dispatch_demand_history() -> dict:
    return scrape_dispatch_history()["demand"]


# ---------------------------------------------------------------------------
# Predispatch
# ---------------------------------------------------------------------------

def scrape_historical_dispatch_prices(date_str: str) -> dict:
    """
    Fetch 5-min dispatch prices for a given date (YYYYMMDD).
    Returns { region: [ {interval: "HH:MM", rrp: float} ] }

    Sources (in priority order):
    1. DispatchIS CURRENT — last ~3 days, table PRICE, filter INTERVENTION=0
    2. TradingIS CURRENT  — last ~15 days, table PRICE, uses PERIODID for time
    """
    from datetime import datetime as _dt, timedelta as _td
    now_aest = datetime.now(AEST)
    today    = now_aest.date()

    try:
        req_date = _dt.strptime(date_str, "%Y%m%d").date()
    except ValueError:
        return {}

    prices: dict = {r: {} for r in NEM_REGIONS}

    # ── Source 1: DispatchIS CURRENT (last ~3 days, 5-min, INTERVENTION filter) ──
    try:
        dispatch_files = _list_hrefs(DISPATCH_IS_URL)
        date_files = sorted([f for f in dispatch_files if date_str in f and "PUBLIC_DISPATCHIS" in f.upper()])
        logger.info(f"scrape_historical_dispatch_prices: {len(date_files)} DispatchIS files for {date_str}")
        for url in date_files:
            try:
                text = _read_zip(url)
                if not text: continue
                for row in _parse_aemo(text, "PRICE"):
                    if row.get("INTERVENTION", "0").strip() != "0":
                        continue
                    region = row.get("REGIONID", "").strip()
                    if region not in NEM_REGIONS: continue
                    dt_str2 = row.get("SETTLEMENTDATE", "")
                    rrp_str = row.get("RRP", "")
                    if not dt_str2 or not rrp_str: continue
                    try:
                        dt = datetime.fromisoformat(dt_str2.replace("/", "-")) - _td(minutes=5)
                        if dt.date() != req_date: continue
                        label = dt.strftime("%H:%M")
                        prices[region][label] = round(float(rrp_str), 2)
                    except (ValueError, TypeError):
                        continue
            except Exception as e:
                logger.debug(f"scrape_historical_dispatch_prices: DispatchIS file error {url}: {e}")
    except Exception as e:
        logger.warning(f"scrape_historical_dispatch_prices: DispatchIS listing failed: {e}")

    # If we got data from DispatchIS, return it
    if any(prices.values()):
        return {r: [{"interval": k, "rrp": v} for k, v in sorted(pts.items())]
                for r, pts in prices.items() if pts}

    # ── Source 2: TradingIS CURRENT (last ~15 days, 5-min via PERIODID) ──
    try:
        trading_files = _list_hrefs(TRADING_IS_URL)
        date_files = sorted([f for f in trading_files if date_str in f and "PUBLIC_TRADINGIS" in f.upper()])
        logger.info(f"scrape_historical_dispatch_prices: {len(date_files)} TradingIS files for {date_str}")
        for url in date_files:
            try:
                text = _read_zip(url)
                if not text: continue
                for row in _parse_aemo(text, "PRICE"):
                    region = row.get("REGIONID", "").strip()
                    if region not in NEM_REGIONS: continue
                    period_str = row.get("PERIODID", "")
                    rrp_str    = row.get("RRP", "")
                    if not period_str or not rrp_str: continue
                    try:
                        # PERIODID 1-288, each = 5 min from midnight
                        period = int(float(period_str))
                        total_min = (period - 1) * 5
                        hh, mm = divmod(total_min, 60)
                        label = f"{hh:02d}:{mm:02d}"
                        prices[region][label] = round(float(rrp_str), 2)
                    except (ValueError, TypeError):
                        continue
            except Exception as e:
                logger.debug(f"scrape_historical_dispatch_prices: TradingIS file error {url}: {e}")
    except Exception as e:
        logger.warning(f"scrape_historical_dispatch_prices: TradingIS listing failed: {e}")

    if not any(prices.values()):
        logger.warning(f"scrape_historical_dispatch_prices: no data found for {date_str}")

    return {r: [{"interval": k, "rrp": v} for k, v in sorted(pts.items())]
            for r, pts in prices.items() if pts}


def _fetch_predispatch() -> str:
    url = get_latest_file_url(PREDISPATCH_URL, "PUBLIC_PREDISPATCHIS")
    return _read_zip(url) if url else ""


def scrape_predispatch_prices(text: str) -> dict:
    now_aest = datetime.now(AEST).replace(tzinfo=None)
    today    = now_aest.date()
    region_series: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    for tk in ["PREDISPATCH_REGION_PRICES", "PREDISPATCH_PRICE", "PREDISPATCH_REGIONPRICE"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            region = row.get("REGIONID", "")
            if region not in NEM_REGIONS:
                continue
            if row.get("INTERVENTION", "0") not in ("0", ""):
                continue
            dt_str = row.get("DATETIME", row.get("SETTLEMENTDATE", ""))
            rrp_str = row.get("RRP", "")
            if not dt_str or not rrp_str:
                continue
            try:
                rrp = round(float(rrp_str), 2)
                # DATETIME is end-of-interval; shift back 30min for display
                dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                # Only keep today's future intervals (AEST) - keep anything within last 30min too
                if dt.date() == today and dt >= now_aest - timedelta(minutes=30):
                    region_series[region][dt.strftime("%H:%M")] = rrp
            except (ValueError, TypeError):
                pass
        break
    result = {}
    for region, series in region_series.items():
        if series:
            result[region] = [{"interval": k, "rrp": v} for k, v in sorted(series.items())]
    logger.info(f"Predispatch prices: {sum(len(v) for v in result.values())} pts")
    return result


def scrape_predispatch_demand(text: str) -> dict:
    now_aest = datetime.now(AEST).replace(tzinfo=None)
    today    = now_aest.date()
    region_series: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    for tk in ["PREDISPATCH_REGION_SOLUTION", "PREDISPATCH_REGIONSOLUTION"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            region = row.get("REGIONID", "")
            if region not in NEM_REGIONS:
                continue
            if row.get("INTERVENTION", "0") not in ("0", ""):
                continue
            dt_str = row.get("DATETIME", row.get("SETTLEMENTDATE", ""))
            demand_str = row.get("DEMAND_AND_NONSCHEDGEN", row.get("TOTALDEMAND", row.get("DEMAND", "")))
            if not dt_str or not demand_str:
                continue
            try:
                demand = round(float(demand_str), 1)
                # DATETIME is end-of-interval; shift back 30min for display
                dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                if dt.date() == today and dt > now_aest:
                    region_series[region][dt.strftime("%H:%M")] = demand
            except (ValueError, TypeError):
                pass
        break
    result = {}
    for region, series in region_series.items():
        if series:
            result[region] = [{"interval": k, "demand": v} for k, v in sorted(series.items())]
    return result


def scrape_predispatch_generation(text: str) -> dict:
    """
    Extract today's future generation forecast from PREDISPATCH files.
    Returns { region: [{interval, Scheduled, SemiScheduled}] } for today's future intervals.
    """
    now_aest = datetime.now(AEST).replace(tzinfo=None)
    today    = now_aest.date()
    region_series: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    for tk in ["PREDISPATCH_REGION_SOLUTION", "PREDISPATCH_REGIONSOLUTION"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            region = row.get("REGIONID", "")
            if region not in NEM_REGIONS:
                continue
            if row.get("INTERVENTION", "0") not in ("0", ""):
                continue
            dt_str = row.get("DATETIME", row.get("SETTLEMENTDATE", ""))
            if not dt_str:
                continue
            try:
                dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                if dt.date() == today and dt >= now_aest:
                    label = dt.strftime("%H:%M")
                    sched = float(row.get("DISPATCHABLEGENERATION", 0) or 0)
                    semi  = float(row.get("SEMISCHEDULE_CLEAREDMW", 0) or 0)
                    sol   = float(row.get("SS_SOLAR_UIGF", row.get("SS_SOLAR_CLEAREDMW", 0)) or 0)
                    win   = float(row.get("SS_WIND_UIGF",  row.get("SS_WIND_CLEAREDMW",  0)) or 0)
                    region_series[region][label] = {
                        "Scheduled":     round(sched, 1),
                        "SemiScheduled": round(semi, 1),
                        "solar":         round(sol, 1),
                        "wind":          round(win, 1),
                    }
            except (ValueError, TypeError):
                pass
        break
    result = {}
    for region, series in region_series.items():
        if series:
            result[region] = [{"interval": k, **v} for k, v in sorted(series.items())]
    return result


def scrape_p5min_unit_solution(text: str) -> dict:
    """
    Extract DUID-level MW forecasts from P5MIN_UNIT_SOLUTION.
    Returns { duid: [{interval: "HH:MM", mw: float}] } for future intervals today.
    P5MIN covers ~1hr ahead in 5-min intervals.
    INTERVAL_DATETIME is end-of-period - subtract 5min for display label.
    """
    if not text:
        return {}

    now_aest = datetime.now(AEST).replace(tzinfo=None)
    today    = now_aest.date()
    now_hhmm = now_aest.strftime("%H:%M")

    result: dict[str, dict] = {}

    for tk in ["P5MIN_UNIT_SOLUTION", "P5MIN_UNITSOLUTION"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            if row.get("INTERVENTION", "0") not in ("0", ""):
                continue
            duid = row.get("DUID", "").strip()
            if not duid:
                continue
            dt_str = row.get("INTERVAL_DATETIME", "")
            if not dt_str:
                continue
            try:
                dt = datetime.strptime(dt_str[:16], "%Y/%m/%d %H:%M")
                dt -= timedelta(minutes=5)
            except Exception:
                try:
                    dt = datetime.strptime(dt_str[:16], "%Y-%m-%d %H:%M")
                    dt -= timedelta(minutes=5)
                except Exception:
                    continue
            if dt.date() != today:
                continue
            hhmm = dt.strftime("%H:%M")
            if hhmm <= now_hhmm:
                continue
            mw_str = row.get("TOTALCLEARED", row.get("SEMIDISPATCHCAP", ""))
            try:
                mw = float(mw_str)
            except Exception:
                continue
            result.setdefault(duid, {})[hhmm] = mw
        if result:
            break

    return {duid: [{"interval": k, "mw": v} for k, v in sorted(pts.items())]
            for duid, pts in result.items() if pts}


# Keep old name as alias so nothing else breaks
scrape_predispatch_unit_solution = scrape_p5min_unit_solution




def scrape_tomorrow_prices(text: str) -> dict:
    """
    Extract tomorrow's predispatch price forecast from PREDISPATCH files.
    Returns { region: [{interval: "HH:MM", rrp: float}] } for tomorrow only.
    """
    now_aest = datetime.now(AEST).replace(tzinfo=None)
    tomorrow = (now_aest + timedelta(days=1)).date()
    region_series: dict[str, dict] = {r: {} for r in NEM_REGIONS}
    for tk in ["PREDISPATCH_REGION_PRICES", "PREDISPATCH_PRICE", "PREDISPATCH_REGIONPRICE"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            region = row.get("REGIONID", "")
            if region not in NEM_REGIONS:
                continue
            if row.get("INTERVENTION", "0") not in ("0", ""):
                continue
            dt_str = row.get("DATETIME", row.get("SETTLEMENTDATE", ""))
            rrp_str = row.get("RRP", "")
            if not dt_str or not rrp_str:
                continue
            try:
                rrp = round(float(rrp_str), 2)
                dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                if dt.date() == tomorrow:
                    region_series[region][dt.strftime("%H:%M")] = rrp
            except (ValueError, TypeError):
                pass
        break
    result = {}
    for region, series in region_series.items():
        if series:
            result[region] = [{"interval": k, "rrp": v} for k, v in sorted(series.items())]
    logger.info(f"Tomorrow prices: {sum(len(v) for v in result.values())} pts")
    return result


def scrape_tomorrow_demand(text: str, stpasa: dict) -> dict:
    """
    Tomorrow's demand forecast - combines predispatch (finer resolution near midnight)
    and STPASA (broader 30-min intervals for the full day).
    Returns { region: [{interval: "HH:MM", demand: float}] }
    """
    now_aest = datetime.now(AEST).replace(tzinfo=None)
    tomorrow = (now_aest + timedelta(days=1)).date()
    region_series: dict[str, dict] = {r: {} for r in NEM_REGIONS}

    # First: predispatch region solution (high-res near midnight)
    for tk in ["PREDISPATCH_REGION_SOLUTION", "PREDISPATCH_REGIONSOLUTION"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            region = row.get("REGIONID", "")
            if region not in NEM_REGIONS:
                continue
            if row.get("INTERVENTION", "0") not in ("0", ""):
                continue
            dt_str = row.get("DATETIME", row.get("SETTLEMENTDATE", ""))
            demand_str = row.get("DEMAND_AND_NONSCHEDGEN", row.get("TOTALDEMAND", row.get("DEMAND", "")))
            if not dt_str or not demand_str:
                continue
            try:
                demand = round(float(demand_str), 1)
                dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                if dt.date() == tomorrow:
                    region_series[region][dt.strftime("%H:%M")] = demand
            except (ValueError, TypeError):
                pass
        break

    # Second: fill remainder from STPASA (30-min intervals)
    for region, pts in stpasa.items():
        if region not in NEM_REGIONS:
            continue
        for pt in pts:
            label_full = pt.get("interval", "")  # "YYYY-MM-DD HH:MM"
            d50 = pt.get("demand_50")
            if not label_full or d50 is None:
                continue
            try:
                dt = datetime.strptime(label_full, "%Y-%m-%d %H:%M")
                if dt.date() == tomorrow:
                    hhmm = dt.strftime("%H:%M")
                    if hhmm not in region_series[region]:  # don't overwrite predispatch
                        region_series[region][hhmm] = d50
            except ValueError:
                pass

    result = {}
    for region, series in region_series.items():
        if series:
            result[region] = [{"interval": k, "demand": v} for k, v in sorted(series.items())]
    logger.info(f"Tomorrow demand: {sum(len(v) for v in result.values())} pts")
    return result


# ---------------------------------------------------------------------------
# Fuel mix - OpenNEM (single NEM-wide call, not 5 calls)
# ---------------------------------------------------------------------------

def _normalise_fuel(fuel: str) -> str:
    f = fuel.upper()
    if "COAL_BLACK" in f or "BLACK" in f:          return "Black Coal"
    if "COAL_BROWN" in f or "BROWN" in f:          return "Brown Coal"
    if "GAS" in f or "OCGT" in f or "CCGT" in f:  return "Gas"
    if "HYDRO" in f or "WATER" in f:               return "Hydro"
    if "WIND" in f:                                return "Wind"
    if "SOLAR" in f or "ROOFTOP" in f:             return "Solar"
    if "BATTERY" in f or "STORAGE" in f:           return "Battery"
    if "LIQUID" in f or "DISTILLATE" in f:         return "Liquid"
    return "Other"


def scrape_fuel_mix_live() -> dict:
    """
    Compute current fuel mix from SCADA output + registration list.
    Returns { region: { fuel: mw } } - a single snapshot (not history).
    Used as primary source; OpenNEM used for history if available.
    """
    try:
        scada = _fetch_full_scada()
        reg   = _load_registration_list()
        if not scada or not reg:
            return {}

        result: dict[str, dict[str, float]] = {r: {} for r in NEM_REGIONS}
        for duid, mw in scada.items():
            info = reg.get(duid, {})
            region = info.get("region", "")
            raw_fuel = info.get("fuel", "")
            fuel   = raw_fuel if (raw_fuel and raw_fuel != "Other") else _infer_fuel_from_duid(duid)
            if region not in NEM_REGIONS or mw is None or mw <= 0:
                continue
            result[region][fuel] = round(result[region].get(fuel, 0) + mw, 1)

        final = {r: v for r, v in result.items() if v}
        logger.info(f"Live fuel mix: {len(final)} regions")
        return final
    except Exception as e:
        logger.warning(f"Live fuel mix failed: {e}")
        return {}


def scrape_fuel_mix_history_opennem() -> dict:
    """
    Fetch fuel mix history from OpenNEM API.
    Returns { region: [ {interval, Black Coal, Gas, ...}, ... ] }
    Falls back to live SCADA snapshot if OpenNEM is unavailable.
    """
    try:
        url = f"{OPENNEM_API}/v4/stats/power/network/NEM?interval=5m&period=1d"
        r = _get(url, timeout=15)
        if not r or r.status_code != 200:
            raise ValueError(f"OpenNEM returned {r.status_code if r else 'no response'}")
        data = r.json()
        # result: region -> interval_str -> fuel -> mw
        result: dict[str, dict] = {}
        for series in data.get("data", []):
            fuel_raw = series.get("fuel_tech") or series.get("type") or ""
            net_region = series.get("network_region", "")
            if not fuel_raw or not net_region:
                continue
            region = net_region if net_region.endswith("1") else net_region + "1"
            if region not in NEM_REGIONS:
                continue
            fuel = _normalise_fuel(fuel_raw)
            history = series.get("history", {})
            start_str = history.get("start", "")
            interval_mins = history.get("interval", 5)
            values = history.get("data", []) or []
            if not start_str or not values:
                continue
            try:
                start_dt = datetime.fromisoformat(start_str.replace("Z", "+00:00")).astimezone(AEST)
                for i, val in enumerate(values):
                    if val is None:
                        continue
                    slot = (start_dt + timedelta(minutes=i * interval_mins)).strftime("%H:%M")
                    rd = result.setdefault(region, {})
                    sd = rd.setdefault(slot, {})
                    sd[fuel] = sd.get(fuel, 0) + round(float(val), 1)
            except (ValueError, TypeError):
                pass
        final = {}
        for region, series in result.items():
            if series:
                final[region] = [{"interval": k, **v} for k, v in sorted(series.items())]
        if final:
            logger.info(f"Fuel mix (OpenNEM): {sum(len(v) for v in final.values())} pts across {len(final)} regions")
            return final
        raise ValueError("OpenNEM returned empty data")
    except Exception as e:
        logger.warning(f"OpenNEM fuel mix failed: {e} - using live SCADA snapshot")
        # Fall back: wrap live snapshot as single-point history
        live = scrape_fuel_mix_live()
        now_label = datetime.now(AEST).strftime("%H:%M")
        return {r: [{"interval": now_label, **fuels}] for r, fuels in live.items()}


# ---------------------------------------------------------------------------
# In-memory demand history - lightweight fallback used only if TradingIS fetch fails
# ---------------------------------------------------------------------------

_demand_history: dict[str, dict] = {r: {} for r in NEM_REGIONS}

# ---------------------------------------------------------------------------
# In-memory fuel mix history - populated by scrape_gen every 15 min
# { region: { "HH:MM": { fuel: mw } } }
# Keyed by AEST time string; old entries pruned to keep only today's data.
# ---------------------------------------------------------------------------
_fuel_history: dict[str, dict] = {r: {} for r in NEM_REGIONS}
_duid_history: dict[str, dict] = {}  # { duid: { "HH:MM": mw } } - per-unit today history
_rooftop_history: dict[str, dict] = {r: {} for r in NEM_REGIONS}  # TOTALINTERMITTENTGENERATION

def _update_fuel_history(fuel_mix: dict, scada: dict | None = None, pump_load: dict | None = None) -> None:
    """Store a fuel mix snapshot and per-DUID snapshot. Prune to today only."""
    now = datetime.now(AEST)
    # Snap to nearest 5-min boundary so labels align with the 5-min time spine in the frontend
    snapped_min = (now.minute // 5) * 5
    label = now.strftime("%H:") + f"{snapped_min:02d}"
    for region in NEM_REGIONS:
        if region not in fuel_mix:
            continue
        snap = dict(fuel_mix[region])
        # Store pump hydro load as negative value under 'Pump Hydro' key
        if pump_load and pump_load.get(region, 0) < -1:
            snap["Pump Hydro"] = round(pump_load[region], 1)
        _fuel_history[region][label] = snap
        if len(_fuel_history[region]) > 290:
            oldest = sorted(_fuel_history[region].keys())[0]
            del _fuel_history[region][oldest]

    # Store per-DUID history for station drill-down
    if scada:
        for duid, mw in scada.items():
            if mw is None:
                continue
            if duid not in _duid_history:
                _duid_history[duid] = {}
            # Pump-load DUIDs report positive MW when consuming - negate for display
            stored_mw = -round(mw, 1) if duid in PUMP_LOAD_DUIDS else round(mw, 1)
            _duid_history[duid][label] = stored_mw
            if len(_duid_history[duid]) > 290:
                oldest = sorted(_duid_history[duid].keys())[0]
                del _duid_history[duid][oldest]

def _get_fuel_history() -> dict:
    now_label = datetime.now(AEST).strftime("%H:%M")
    result = {}
    for region, series in _fuel_history.items():
        if series:
            result[region] = [{"interval": k, **v} for k, v in sorted(series.items()) if k <= now_label]
    return result


def _update_demand_history(region_summary: dict) -> None:
    label = datetime.now(AEST).strftime("%H:%M")
    for region, data in region_summary.items():
        if region in NEM_REGIONS and "TOTALDEMAND" in data:
            _demand_history[region][label] = data["TOTALDEMAND"]


def _get_demand_history() -> dict:
    result = {}
    for region, series in _demand_history.items():
        if series:
            result[region] = [{"interval": k, "demand": v} for k, v in sorted(series.items())]
    return result


# ---------------------------------------------------------------------------
# In-memory interconnector history + predispatch IC flows
# ---------------------------------------------------------------------------

_ic_history: dict[str, dict] = {}
_bdu_history: dict[str, dict] = {r: {} for r in NEM_REGIONS}  # { region: { HH:MM: {gen,load,storage,max_avail} } }


def _update_ic_history(ic_snapshot: dict) -> None:
    label = datetime.now(AEST).strftime("%H:%M")
    for ic_id, vals in ic_snapshot.items():
        if ic_id not in _ic_history:
            _ic_history[ic_id] = {}
        _ic_history[ic_id][label] = vals.get("flow", 0)


def _get_ic_history() -> dict:
    result = {}
    for ic_id, series in _ic_history.items():
        if series:
            result[ic_id] = [{"interval": k, "flow": v} for k, v in sorted(series.items())]
    return result


def _update_bdu_history(region_summary: dict) -> None:
    label = datetime.now(AEST).strftime("%H:%M")
    for region, d in region_summary.items():
        if region not in NEM_REGIONS:
            continue
        gen  = d.get("BDU_CLEAREDMW_GEN", 0) or 0
        load = d.get("BDU_CLEAREDMW_LOAD", 0) or 0
        soc  = d.get("BDU_ENERGY_STORAGE")
        cap  = d.get("BDU_MAX_AVAIL") or d.get("BDU_MIN_AVAIL")
        # Pump hydro load = total scheduled load - battery load
        disp_load = d.get("DISPATCHABLELOAD", 0) or 0
        pump_load = max(round(disp_load - load, 1), 0)  # never negative
        _bdu_history[region][label] = {
            "net_mw":    round(gen - load, 1),   # positive=discharging, negative=charging
            "gen":       round(gen, 1),
            "load":      round(load, 1),
            "pump_load": pump_load,               # pump hydro MW consuming
            "storage":   round(soc, 1) if soc is not None else None,
            "max_avail": round(cap, 1) if cap is not None else None,
        }


def _update_live_duid_history(scada: dict) -> None:
    """Write current SCADA snapshot into _duid_history so modal charts stay live.
    Called every 5 min from scrape_all. Only writes DUIDs we care about."""
    label = datetime.now(AEST).strftime("%H:%M")
    for duid, mw in scada.items():
        if mw is None:
            continue
        # Negate pump-load DUIDs so they show as negative
        stored_mw = -round(mw, 1) if duid in PUMP_LOAD_DUIDS else round(mw, 1)
        if duid not in _duid_history:
            _duid_history[duid] = {}
        _duid_history[duid][label] = stored_mw
        # Trim to 290 points (~24h of 5-min data)
        if len(_duid_history[duid]) > 290:
            oldest = sorted(_duid_history[duid].keys())[0]
            del _duid_history[duid][oldest]


def _get_bdu_history() -> dict:
    now_label = datetime.now(AEST).strftime("%H:%M")
    result = {}
    for region, series in _bdu_history.items():
        if series:
            result[region] = [{"interval": k, **v} for k, v in sorted(series.items()) if k <= now_label]
    return result


def scrape_predispatch_interconnectors(text: str) -> dict:
    now_aest = datetime.now(AEST).replace(tzinfo=None)
    ic_series: dict[str, dict] = {}
    for tk in ["PREDISPATCH_INTERCONNECTOR_SOLN", "PREDISPATCH_INTERCONNECTORSOLN"]:
        rows = _parse_aemo(text, tk)
        if not rows:
            continue
        for row in rows:
            ic = row.get("INTERCONNECTORID", "").strip()
            dt_str = row.get("DATETIME", row.get("SETTLEMENTDATE", ""))
            flow_str = row.get("MWFLOW", "")
            if not ic or not dt_str or not flow_str:
                continue
            try:
                flow = round(float(flow_str), 1)
                dt = datetime.fromisoformat(dt_str.replace("/", "-"))
                if dt.replace(tzinfo=None) >= now_aest:
                    ic_series.setdefault(ic, {})[dt.strftime("%H:%M")] = flow
            except (ValueError, TypeError):
                pass
        break
    return {ic: [{"interval": k, "flow": v} for k, v in sorted(s.items())]
            for ic, s in ic_series.items() if s}


# ---------------------------------------------------------------------------
# Main scrape_all - parallel fetches
# ---------------------------------------------------------------------------

def scrape_all() -> dict:
    logger.info("scrape_all starting...")

    # Run all IO-bound fetches concurrently with per-future timeout
    with ThreadPoolExecutor(max_workers=4) as ex:
        f_dispatch_is   = ex.submit(_fetch_dispatch_is)
        f_predispatch   = ex.submit(_fetch_predispatch)
        f_trading       = ex.submit(scrape_trading_history)
        f_dispatch_hist = ex.submit(scrape_dispatch_history)
        f_scada         = ex.submit(_fetch_full_scada)

    def _safe_result(fut, default, name):
        try:
            return fut.result(timeout=55)
        except Exception as e:
            logger.error(f"scrape_all: {name} failed - {e}")
            return default

    dispatch_text    = _safe_result(f_dispatch_is,   "",  "dispatch_is")
    predispatch_text = _safe_result(f_predispatch,    "",  "predispatch")
    trading          = _safe_result(f_trading,        {"prices": {}, "fetch_stats": {}}, "trading")
    dispatch_hist    = _safe_result(f_dispatch_hist,  {"demand": {}, "op_demand": {}, "prices": {}}, "dispatch_hist")
    scada_vals       = _safe_result(f_scada,          {}, "scada")

    dispatch_demand       = dispatch_hist.get("demand", {})
    dispatch_op_demand    = dispatch_hist.get("op_demand", {})
    dispatch_price_5min   = dispatch_hist.get("prices", {})
    dispatch_solar        = dispatch_hist.get("solar", {})
    dispatch_wind         = dispatch_hist.get("wind", {})

    # Parse live dispatch snapshot (prices, demand, generation, ICs)
    region_summary  = scrape_region_summary(dispatch_text)
    interconnectors = scrape_interconnectors(dispatch_text)

    # Supplement SCADA with unit solution if needed
    missing = ORIGIN_DUIDS - set(scada_vals.keys())
    if missing:
        unit_sol = scrape_unit_solution(dispatch_text, missing)
        scada_vals.update(unit_sol)

    # Parse predispatch (future forecasts only)
    pd_prices = scrape_predispatch_prices(predispatch_text)
    pd_demand = scrape_predispatch_demand(predispatch_text)
    pd_gen    = scrape_predispatch_generation(predispatch_text)
    pd_units  = {}  # AEMO does not publish unit-level predispatch on public NEMWeb

    # Tomorrow's forecasts from same predispatch file
    tomorrow_prices = scrape_tomorrow_prices(predispatch_text)
    tomorrow_demand = scrape_tomorrow_demand(predispatch_text, {})

    # In-memory accumulators - IC builds up over process lifetime
    _update_demand_history(region_summary)
    _update_ic_history(interconnectors)
    _update_bdu_history(region_summary)
    _update_live_duid_history(scada_vals)   # keep _duid_history current for modal charts

    # Build fuel_mix from live SCADA and update fuel history every 5 min
    # This keeps the fuel stack charts current without waiting for scrape_gen (15 min)
    live_fuel_mix: dict = {r: {} for r in NEM_REGIONS}
    live_pump_load: dict = {}
    for duid, mw_raw in scada_vals.items():
        if mw_raw is None:
            continue
        info = NEM_UNITS.get(duid, {})
        region = info.get("region", "")
        fuel = info.get("fuel", "") or _infer_fuel_from_duid(duid)
        if region not in NEM_REGIONS:
            region = _infer_region_from_duid(duid)
        if region not in NEM_REGIONS:
            continue
        mw = -mw_raw if duid in PUMP_LOAD_DUIDS else mw_raw
        mw_pos = max(mw, 0)
        live_fuel_mix[region][fuel] = round(live_fuel_mix[region].get(fuel, 0) + mw_pos, 1)
        if fuel == "Hydro" and mw < -1:
            live_pump_load[region] = round(live_pump_load.get(region, 0) + mw, 1)
    # Inject rooftop solar from region_summary
    for region, rdata in region_summary.items():
        if region in NEM_REGIONS and "TOTALINTERMITTENTGENERATION" in rdata:
            rooftop = rdata.get("TOTALINTERMITTENTGENERATION", 0)
            if rooftop:
                live_fuel_mix[region]["Rooftop Solar"] = round(float(rooftop), 1)
    _update_fuel_history(live_fuel_mix, None, live_pump_load)
    # Inject live solar/wind into dispatch history for current interval
    now_label = datetime.now(AEST).strftime("%H:%M")
    for region, rdata in region_summary.items():
        if region not in NEM_REGIONS:
            continue
        sol = rdata.get("SS_SOLAR_CLEAREDMW")
        win = rdata.get("SS_WIND_CLEAREDMW")
        if sol is not None:
            dispatch_solar.setdefault(region, [])
            dispatch_solar[region] = [p for p in dispatch_solar[region] if p["interval"] != now_label]
            dispatch_solar[region].append({"interval": now_label, "mw": round(float(sol), 1)})
        if win is not None:
            dispatch_wind.setdefault(region, [])
            dispatch_wind[region] = [p for p in dispatch_wind[region] if p["interval"] != now_label]
            dispatch_wind[region].append({"interval": now_label, "mw": round(float(win), 1)})
    # Use DispatchIS history for demand (full day), fall back to in-memory if empty
    demand_history     = dispatch_demand if dispatch_demand else _get_demand_history()
    ic_history         = _get_ic_history()
    pd_interconnectors = scrape_predispatch_interconnectors(predispatch_text)

    # Live current values from latest dispatch interval
    prices     = {r: d["RRP"]         for r, d in region_summary.items() if "RRP" in d}
    demand     = {r: d["TOTALDEMAND"] for r, d in region_summary.items() if "TOTALDEMAND" in d}
    op_demand  = {r: d["DEMAND_AND_NONSCHEDGEN"] for r, d in region_summary.items() if "DEMAND_AND_NONSCHEDGEN" in d}
    generation = {r: {
        "Scheduled":      d.get("DISPATCHABLEGENERATION", 0),
        "Semi-Scheduled": d.get("SEMISCHEDULE_CLEAREDMW", 0),
        "Net Interchange":d.get("NETINTERCHANGE", 0),
    } for r, d in region_summary.items() if "DISPATCHABLEGENERATION" in d}

    # Build Origin assets output - ORIGIN_DISPLAY_NAMES overrides nem_units station name
    origin_assets_out = {}
    for duid in ORIGIN_DUIDS:
        mw       = scada_vals.get(duid)
        reg_info = NEM_UNITS.get(duid, {})\
        # Solar and wind absent from SCADA = generating 0 (semi-scheduled, not truly unknown)
        fuel     = reg_info.get("fuel")     or "Other"
        if mw is None and fuel in ("Solar", "Wind"):
            mw = 0.0
        station  = ORIGIN_DISPLAY_NAMES.get(duid) or reg_info.get("station") or duid
        region   = reg_info.get("region")   or ""
        capacity = reg_info.get("capacity")
        origin_assets_out[duid] = {
            "station":  station,
            "fuel":     fuel,
            "region":   region,
            "capacity": capacity,
            "mw":       mw,
            "pct":      round(mw / capacity * 100, 1) if mw is not None and capacity else None,
            "status":   "running" if (mw is not None and mw > 5) else ("charging" if (mw is not None and mw < -5) else ("off" if mw is not None else "unknown")),
        }

    # Keep trading (firm 30-min) and dispatch 5-min prices separate
    # so the frontend can style them differently
    trading_prices = trading["prices"]
    # Cap dispatch at now
    now_label = datetime.now(AEST).strftime("%H:%M")
    capped_dispatch_prices = {}
    for r in NEM_REGIONS:
        pts = [p for p in dispatch_price_5min.get(r, []) if p["interval"] <= now_label]
        if pts:
            capped_dispatch_prices[r] = pts

    logger.info(
        f"scrape_all done - prices:{list(prices.keys())} "
        f"trading_pts:{sum(len(v) for v in trading_prices.values())} "
        f"dispatch_5min_pts:{sum(len(v) for v in capped_dispatch_prices.values())} "
        f"origin_duids_found:{len(scada_vals)}"
    )

    return {
        "timestamp":             datetime.now(timezone.utc).isoformat(),
        "prices":                prices,
        "demand":                demand,
        "op_demand":             op_demand,
        "generation":            generation,
        "interconnectors":       interconnectors,
        "raw_summary":           region_summary,
        "historical_prices":     trading_prices,
        "dispatch_prices_5min":  capped_dispatch_prices,
        "price_fetch_stats":     trading.get("fetch_stats", {}),
        "predispatch_prices":    pd_prices,
        "demand_history":        demand_history,
        "op_demand_history":     dispatch_op_demand,
        "dispatch_history":      dispatch_demand,
        "solar_history":         dispatch_solar,
        "wind_history":          dispatch_wind,
        "gen_demand_history":    dispatch_hist.get("gen_demand", {}),
        "predispatch_demand":    pd_demand,
        "predispatch_gen":       pd_gen,
        "predispatch_units":     pd_units,
        "ic_history":            ic_history,
        "predispatch_ic":        pd_interconnectors,
        "bdu_history":           _get_bdu_history(),
        "origin_assets":         origin_assets_out,
        "fuel_colors":           FUEL_COLORS,
        "all_fuels":             ALL_FUELS,
        "tomorrow_prices":       tomorrow_prices,
        "tomorrow_demand":       tomorrow_demand,
    }


# ---------------------------------------------------------------------------
# Registration list cache - DUID -> {station, fuel, region, capacity}
# ---------------------------------------------------------------------------

_reg_cache: dict = {}
_reg_cache_date: str = ""


def _fuel_from_reg(fuel_raw: str) -> str:
    """Map AEMO registration fuel source string to display fuel."""
    f = (fuel_raw or "").upper()
    if "BLACK COAL" in f or "COAL" in f and "BROWN" not in f: return "Black Coal"
    if "BROWN COAL" in f or "BROWN" in f:                      return "Brown Coal"
    if "GAS" in f or "OCGT" in f or "CCGT" in f or "LIQUID FUEL" in f and "GAS" in f: return "Gas"
    if "HYDRO" in f or "WATER" in f:                           return "Hydro"
    if "WIND" in f:                                             return "Wind"
    if "SOLAR" in f or "PHOTOVOLTAIC" in f:                    return "Solar"
    if "BATTERY" in f or "STORAGE" in f:                       return "Battery"
    if "LIQUID" in f or "DISTILLATE" in f or "DIESEL" in f:   return "Liquid"
    return "Other"


def _load_registration_list() -> dict:
    """
    Download and parse AEMO NEM Registration list from NEMWeb Static CSV.
    Falls back to the AEMO website XLS if CSV unavailable.
    Returns { DUID: {station, fuel, region, capacity_mw, participant} }
    Cached once per day.
    """
    global _reg_cache, _reg_cache_date
    today = datetime.now(AEST).strftime("%Y-%m-%d")
    if _reg_cache and _reg_cache_date == today:
        return _reg_cache

    # Primary: NEMWeb Generators and Scheduled Loads CSV (no auth, no XLS)
    CSV_URLS = [
        "https://www.nemweb.com.au/REPORTS/CURRENT/SEMP/PUBLIC_SEMP_REGISTRATION.CSV",
        "https://nemweb.com.au/REPORTS/CURRENT/SEMP/PUBLIC_SEMP_REGISTRATION.CSV",
        f"{NEMWEB_BASE}/REPORTS/CURRENT/SEMP/PUBLIC_SEMP_REGISTRATION.CSV",
    ]
    # Fallback: AEMO static file page (sometimes available without redirect)
    AEMO_URLS = [
        "https://aemo.com.au/-/media/Files/Electricity/NEM/Participant_Information/Current-Participants/NEM-Registration-and-Exemption-List.xls",
        AEMO_REG_LIST_URL,
    ]

    result = _try_load_nemweb_csv()
    if result:
        _reg_cache = result
        _reg_cache_date = today
        return result

    # Fallback: try AEMO XLS
    result = _try_load_aemo_xls()
    if result:
        _reg_cache = result
        _reg_cache_date = today
        return result

    logger.warning("Registration list: all sources failed, using empty cache")
    return _reg_cache


def _try_load_nemweb_csv() -> dict:
    """
    Try multiple NEMWeb sources for DUID registration data.
    Priority:
      1. SEMP registration CSV (direct, updated daily)
      2. MMS DVD DUDETAILSUMMARY zip (monthly snapshot, very complete)
      3. MMS DVD DUDETAIL zip (fallback)
    """
    # Source 1: SEMP registration CSV
    semp_url = f"{NEMWEB_BASE}/REPORTS/CURRENT/SEMP/PUBLIC_SEMP_REGISTRATION.CSV"
    r = _get(semp_url, timeout=15)
    if r and r.status_code == 200 and len(r.content) > 1000:
        logger.info("Registration: loaded from SEMP CSV")
        return _parse_registration_csv(r.text)

    # Source 2: MMS DVD DUDETAILSUMMARY — find latest zip
    for dvd_path in [
        f"{NEMWEB_BASE}/REPORTS/CURRENT/DVD/",
        f"{NEMWEB_BASE}/REPORTS/ARCHIVE/DVD/",
    ]:
        try:
            files = _list_hrefs(dvd_path)
            # Find DUDETAILSUMMARY files
            detail_files = sorted([f for f in files if "DUDETAILSUMMARY" in f.upper()])
            if not detail_files:
                detail_files = sorted([f for f in files if "DUDETAIL" in f.upper()])
            if detail_files:
                url = detail_files[-1]
                logger.info(f"Registration: trying DVD {url}")
                text = _read_zip(url)
                if text:
                    result = _parse_dudetail_csv(text)
                    if result:
                        logger.info(f"Registration: loaded {len(result)} DUIDs from DVD")
                        return result
        except Exception as e:
            logger.debug(f"Registration DVD fetch failed: {e}")

    logger.warning("Registration: all sources failed")
    return {}


def _parse_dudetail_csv(text: str) -> dict:
    """Parse MMS DUDETAILSUMMARY or DUDETAIL CSV into {DUID: info} dict."""
    import csv, io
    result = {}
    try:
        # MMS format: I,DUDETAILSUMMARY,DUDETAILSUMMARY,1,...col headers...
        # D rows: D,DUDETAILSUMMARY,DUDETAILSUMMARY,1,...values...
        reader = csv.reader(io.StringIO(text))
        headers = []
        for row in reader:
            if not row:
                continue
            rec_type = row[0].strip().upper()
            if rec_type == 'I' and len(row) > 4:
                table = row[2].strip().upper()
                if 'DUDETAIL' in table:
                    headers = [c.strip().upper() for c in row[4:]]
            elif rec_type == 'D' and headers:
                if len(row) < 5:
                    continue
                vals = row[4:]
                d = dict(zip(headers, vals))
                duid = d.get('DUID', '').strip().upper()
                if not duid:
                    continue
                # DUDETAILSUMMARY columns: DUID, REGIONID, STATIONID, DISPATCHTYPE,
                #   CONNECTIONPOINTID, REGISTEREDCAPACITY, LASTCHANGED
                region = d.get('REGIONID', '').strip()
                if region and not region.endswith('1'):
                    region += '1'
                capacity_str = d.get('REGISTEREDCAPACITY', '') or d.get('MAXCAPACITY', '')
                try:
                    capacity = round(float(capacity_str)) if capacity_str else 0
                except ValueError:
                    capacity = 0
                station = d.get('STATIONID', duid).strip()
                # Fuel not in DUDETAILSUMMARY — will need to infer or leave for SEMP
                fuel = _infer_fuel_from_duid(duid)
                if region:
                    result[duid] = {
                        "station":  station,
                        "fuel":     fuel,
                        "region":   region,
                        "capacity": capacity,
                    }
    except Exception as e:
        logger.warning(f"_parse_dudetail_csv failed: {e}")
    return result


def _parse_registration_csv(text: str) -> dict:
    """Parse a CSV registration file into {DUID: info} dict."""
    result = {}
    try:
        reader = csv.DictReader(io.StringIO(text))
        headers_upper = {k: k.upper() for k in (reader.fieldnames or [])}

        def get(row, *keys):
            for k in keys:
                for fk, fu in headers_upper.items():
                    if k.upper() in fu:
                        return str(row.get(fk, "") or "").strip()
            return ""

        for row in reader:
            duid = get(row, "DUID").upper()
            if not duid:
                continue
            region = get(row, "REGION", "REGIONID")
            if region and not region.endswith("1"):
                region = region + "1"
            fuel_raw = get(row, "FUEL SOURCE", "FUEL_SOURCE", "FUEL", "TECHNOLOGY")
            cap_str = get(row, "REG CAP", "REGISTERED_CAPACITY", "CAPACITY", "MAX_CAP")
            try:
                capacity = round(float(cap_str), 1) if cap_str else None
            except (ValueError, TypeError):
                capacity = None
            result[duid] = {
                "station":     get(row, "STATION NAME", "STATION", "STATIONNAME") or duid,
                "fuel":        _fuel_from_reg(fuel_raw),
                "fuel_raw":    fuel_raw,
                "region":      region,
                "capacity":    capacity,
                "participant": get(row, "PARTICIPANT", "PARTICIPANTID"),
            }
        logger.info(f"Registration CSV parsed: {len(result)} DUIDs")
    except Exception as e:
        logger.warning(f"Registration CSV parse failed: {e}")
    return result


def _try_load_aemo_xls() -> dict:
    """Try to load AEMO registration XLS using xlrd."""
    try:
        import xlrd
    except ImportError:
        logger.warning("xlrd not installed")
        return {}

    headers_to_try = {
        "User-Agent": "Mozilla/5.0 (compatible; NEM-Dashboard/1.0)",
        "Accept": "application/vnd.ms-excel,*/*",
        "Referer": "https://aemo.com.au/",
    }
    try:
        r = None
        for _url in AEMO_REG_LIST_URLS:
            try:
                r = SESSION.get(_url, timeout=20, headers=headers_to_try, allow_redirects=True)
                if r and r.status_code == 200 and len(r.content) > 10000:
                    logger.info(f"Registration XLS loaded from {_url}")
                    break
            except Exception:
                continue
        if not r or r.status_code != 200 or len(r.content) < 10000:
            logger.warning(f"AEMO XLS fetch failed: status={getattr(r,'status_code','?')} size={len(getattr(r,'content',b''))}")
            return {}

        logger.info(f"AEMO XLS: {len(r.content)} bytes, content-type={r.headers.get('Content-Type','?')}")
        wb = xlrd.open_workbook(file_contents=r.content)
        sheet = None
        for name in wb.sheet_names():
            if "generator" in name.lower() or "scheduled" in name.lower():
                sheet = wb.sheet_by_name(name)
                break
        if sheet is None:
            sheet = wb.sheet_by_index(0)

        # Find DUID header row
        header_row_idx = None
        for i in range(min(20, sheet.nrows)):
            cells = [str(sheet.cell_value(i, j)).upper() for j in range(sheet.ncols)]
            if any("DUID" in c for c in cells):
                header_row_idx = i
                logger.info(f"XLS header at row {i}: {cells[:8]}")
                break
        if header_row_idx is None:
            logger.warning("XLS: no DUID header found")
            return {}

        headers = [str(sheet.cell_value(header_row_idx, j)).strip().upper()
                   for j in range(sheet.ncols)]

        def col(name_parts):
            for i, h in enumerate(headers):
                if any(p.upper() in h for p in name_parts):
                    return i
            return None

        ci_duid     = col(["DUID"])
        ci_station  = col(["STATION NAME", "STATION"])
        ci_fuel     = col(["FUEL SOURCE - DESCRIPTOR", "FUEL SOURCE", "FUEL"])
        ci_region   = col(["REGION"])
        ci_capacity = col(["REG CAP", "REGISTERED CAPACITY", "MAX CAP", "CAPACITY"])
        ci_part     = col(["PARTICIPANT"])

        result = {}
        for i in range(header_row_idx + 1, sheet.nrows):
            if ci_duid is None:
                continue
            duid = str(sheet.cell_value(i, ci_duid)).strip().upper()
            if not duid or duid == "DUID":
                continue
            station  = str(sheet.cell_value(i, ci_station)).strip()  if ci_station  is not None else ""
            fuel_raw = str(sheet.cell_value(i, ci_fuel)).strip()     if ci_fuel     is not None else ""
            region   = str(sheet.cell_value(i, ci_region)).strip()   if ci_region   is not None else ""
            cap_raw  = sheet.cell_value(i, ci_capacity)               if ci_capacity is not None else None
            part     = str(sheet.cell_value(i, ci_part)).strip()     if ci_part     is not None else ""
            try:
                capacity = round(float(cap_raw), 1) if cap_raw not in (None, "") else None
            except (ValueError, TypeError):
                capacity = None
            if region and not region.endswith("1"):
                region = region + "1"
            result[duid] = {
                "station":     station or duid,
                "fuel":        _fuel_from_reg(fuel_raw),
                "fuel_raw":    fuel_raw,
                "region":      region,
                "capacity":    capacity,
                "participant": part,
            }
        logger.info(f"AEMO XLS loaded: {len(result)} DUIDs")
        return result

    except Exception as e:
        logger.warning(f"AEMO XLS load failed: {e}", exc_info=True)
        return {}




# ---------------------------------------------------------------------------
# Full SCADA fetch - all DUIDs (not just Origin)
# ---------------------------------------------------------------------------

def _fetch_full_scada() -> dict:
    """Return { DUID: mw } for every unit in DISPATCH_UNIT_SCADA."""
    url = get_latest_file_url(SCADA_URL, "PUBLIC_DISPATCHSCADA")
    if not url:
        return {}
    text = _read_zip(url)
    result = {}
    for row in _parse_aemo(text, "DISPATCH_UNIT_SCADA"):
        duid = row.get("DUID", "").strip().upper()
        v = row.get("SCADAVALUE", "")
        try:
            result[duid] = round(float(v), 1)
        except (ValueError, TypeError):
            pass
    logger.info(f"Full SCADA: {len(result)} DUIDs")
    return result


ROOFTOP_PV_URL = f"{NEMWEB_BASE}/REPORTS/CURRENT/ROOFTOP_PV/ACTUAL/"

def _scrape_rooftop_pv_latest() -> dict:
    """
    Fetch the latest ROOFTOP_PV_ACTUAL file and return { regionid: mw }.
    Published every 30 min with 30-min actuals by state.
    Table: ROOFTOP_PV,ACTUAL  Cols: INTERVAL_DATETIME, REGIONID, POWER, QI
    """
    try:
        url = get_latest_file_url(ROOFTOP_PV_URL, "PUBLIC_ROOFTOP_PV_ACTUAL")
        if not url:
            return {}
        text = _read_zip(url)
        if not text:
            return {}
        # Take latest interval per region (file may have multiple 30-min rows)
        result = {}
        latest_dt = {}
        for row in _parse_aemo(text, "ROOFTOP_PV_ACTUAL"):
            region = row.get("REGIONID", "").strip().upper()
            if region not in NEM_REGIONS:
                continue
            dt_str  = row.get("INTERVAL_DATETIME", "")
            power_str = row.get("POWER", "")
            try:
                mw = round(float(power_str), 1)
                if mw < 0:
                    continue
                if dt_str > latest_dt.get(region, ""):
                    latest_dt[region] = dt_str
                    result[region] = mw
            except (ValueError, TypeError):
                pass
        logger.info(f"Rooftop PV actual: {result}")
        return result
    except Exception as e:
        logger.warning(f"Rooftop PV fetch failed: {e}")
        return {}


# ---------------------------------------------------------------------------
# ST PASA - 7-day ahead regional demand forecast
# ---------------------------------------------------------------------------

def scrape_stpasa_demand() -> dict:
    """
    Fetch latest ST PASA file and extract STPASA_REGIONSOLUTION.
    Returns { region: [{interval, demand_50, demand_10, solar_uigf, wind_uigf}] }
    """
    try:
        urls = _list_hrefs(ST_PASA_URL)
        pasa_urls = sorted([u for u in urls if "STPASA" in u.upper()])
        if not pasa_urls:
            logger.warning("No ST PASA files found")
            return {}
        url = pasa_urls[-1]
        text = _read_zip(url)

        now_aest = datetime.now(AEST).replace(tzinfo=None)
        region_series: dict = {r: {} for r in NEM_REGIONS}

        for tk in ["STPASA_REGIONSOLUTION", "ST_PASA_REGIONSOLUTION"]:
            rows = _parse_aemo(text, tk)
            if not rows:
                continue
            for row in rows:
                region = row.get("REGIONID", "").strip()
                if region not in NEM_REGIONS:
                    continue
                dt_str = row.get("INTERVAL_DATETIME", row.get("SETTLEMENTDATE", ""))
                d50 = row.get("DEMAND50", row.get("TOTALDEMAND", ""))
                d10 = row.get("DEMAND10", "")
                solar = row.get("SS_SOLAR_UIGF", "")
                wind  = row.get("SS_WIND_UIGF", "")
                if not dt_str:
                    continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-"))
                    if dt.replace(tzinfo=None) < now_aest:
                        continue
                    label = dt.strftime("%Y-%m-%d %H:%M")
                    region_series[region][label] = {
                        "demand_50":  round(float(d50), 1) if d50 else None,
                        "demand_10":  round(float(d10), 1) if d10 else None,
                        "solar_uigf": round(float(solar), 1) if solar else None,
                        "wind_uigf":  round(float(wind), 1) if wind else None,
                    }
                except (ValueError, TypeError):
                    pass
            if any(region_series.values()):
                break

        result = {}
        for region, series in region_series.items():
            if series:
                result[region] = [{"interval": k, **v} for k, v in sorted(series.items())]
        logger.info(f"ST PASA: {sum(len(v) for v in result.values())} pts across {len(result)} regions")
        return result
    except Exception as e:
        logger.warning(f"ST PASA fetch failed: {e}")
        return {}


# ---------------------------------------------------------------------------
# All-generators scrape (for /api/slow endpoint)
# ---------------------------------------------------------------------------

def scrape_all_generators() -> dict:
    """
    Fetch full SCADA output for all NEM DUIDs, join with registration list.
    Returns grouped structure: { region: { fuel: [ {duid, station, mw, capacity, pct} ] } }
    """
    # Fetch SCADA and registration list concurrently
    with ThreadPoolExecutor(max_workers=2) as ex:
        f_scada = ex.submit(_fetch_full_scada)
        f_reg   = ex.submit(_load_registration_list)
    scada  = f_scada.result()
    reg    = f_reg.result()

    timestamp = datetime.now(timezone.utc).isoformat()

    # Group by region → fuel → list of units
    grouped: dict = {}
    for duid, mw in scada.items():
        info = reg.get(duid, {})
        region   = info.get("region", "UNKNOWN")
        fuel     = info.get("fuel", "Other")
        station  = info.get("station", duid)
        capacity = info.get("capacity")
        pct = round(mw / capacity * 100, 1) if (mw is not None and capacity and capacity > 0) else None

        if region not in NEM_REGIONS:
            continue  # skip non-NEM (WA, etc.)

        rg = grouped.setdefault(region, {})
        fg = rg.setdefault(fuel, [])
        fg.append({
            "duid":     duid,
            "station":  station,
            "mw":       round(mw, 1) if mw is not None else None,
            "capacity": capacity,
            "pct":      pct,
            "participant": info.get("participant", ""),
        })

    # Sort within each fuel group by MW descending
    for region in grouped:
        for fuel in grouped[region]:
            grouped[region][fuel].sort(key=lambda x: x["mw"] or 0, reverse=True)

    # Summary stats per region/fuel
    summary: dict = {}
    for region, fuels in grouped.items():
        summary[region] = {}
        for fuel, units in fuels.items():
            total_mw  = sum(u["mw"] or 0 for u in units)
            total_cap = sum(u["capacity"] or 0 for u in units)
            summary[region][fuel] = {
                "total_mw":  round(total_mw, 1),
                "total_cap": round(total_cap, 1),
                "unit_count": len(units),
            }

    logger.info(f"scrape_all_generators: {sum(len(v) for r in grouped.values() for v in r.values())} units across {len(grouped)} regions")
    return {
        "timestamp":      datetime.now(timezone.utc).isoformat(),
        "grouped":        grouped,
        "summary":        summary,
        "fuel_colors":    FUEL_COLORS,
        "reg_list_count": len(reg),
        "scada_count":    len(scada),
    }


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# scrape_scada_history - backfill 24hr fuel mix history from SCADA CURRENT files
# ---------------------------------------------------------------------------

def scrape_scada_history() -> None:
    """
    Fetch all of today's DISPATCH_SCADA files in parallel and populate
    _fuel_history with historical fuel mix snapshots.
    Called once at startup so the gen chart shows 24hr history immediately.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time as _time, random

    now_aest   = datetime.now(AEST)
    today_str  = now_aest.strftime("%Y%m%d")
    today_date = now_aest.date()
    now_label  = now_aest.strftime("%H:%M")

    all_urls = _list_hrefs(SCADA_URL)
    today_urls = sorted([u for u in all_urls if today_str in u and "PUBLIC_DISPATCHSCADA" in u.upper()])
    # Limit to 300 files (25hrs worth at 5min cadence)
    today_urls = today_urls[-300:]

    if not today_urls:
        logger.warning("scrape_scada_history: no SCADA files found for today")
        return

    logger.info(f"scrape_scada_history: fetching {len(today_urls)} SCADA files...")

    def fetch_one(url):
        _time.sleep(random.uniform(0, 0.03))
        try:
            text = _read_zip(url)
            if not text:
                return None
            # Each file: { HH:MM -> { duid: mw } }
            snapshot = {}
            for row in _parse_aemo(text, "DISPATCH_UNIT_SCADA"):
                duid = row.get("DUID", "").strip().upper()
                dt_str = row.get("SETTLEMENTDATE", "")
                v = row.get("SCADAVALUE", "")
                if not dt_str or not v:
                    continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=5)
                    if dt.date() != today_date or dt.strftime("%H:%M") > now_label:
                        continue
                    label = dt.strftime("%H:%M")
                    if label not in snapshot:
                        snapshot[label] = {}
                    snapshot[label][duid] = round(float(v), 1)
                except (ValueError, TypeError):
                    pass
            return snapshot
        except Exception as e:
            logger.debug(f"scrape_scada_history fetch error: {e}")
            return None

    # Parallel fetch
    all_snapshots = {}  # { HH:MM -> { duid: mw } }
    with ThreadPoolExecutor(max_workers=4) as ex:
        futures = {ex.submit(fetch_one, u): u for u in today_urls}
        for fut in as_completed(futures):
            result = fut.result()
            if result:
                for label, duids in result.items():
                    if label not in all_snapshots:
                        all_snapshots[label] = {}
                    all_snapshots[label].update(duids)

    # Build CPID→region map from the latest DispatchIS for unregistered DUIDs
    _CPID_REGION = {"N": "NSW1", "Q": "QLD1", "V": "VIC1", "S": "SA1", "T": "TAS1"}
    cpid_map: dict = {}
    try:
        dispatch_text = _fetch_dispatch_is()
        for row in _parse_aemo(dispatch_text, "DISPATCH_UNIT_SOLUTION"):
            duid = row.get("DUID", "").strip().upper()
            cpid = row.get("CONNECTIONPOINTID", "").strip().upper()
            if duid and cpid and cpid[0] in _CPID_REGION:
                cpid_map[duid] = _CPID_REGION[cpid[0]]
    except Exception as e:
        logger.warning(f"scrape_scada_history: CPID map fetch failed: {e}")

    # Now convert to fuel_mix per timestamp and load into _fuel_history + _duid_history
    reg = NEM_UNITS
    loaded = 0
    for label in sorted(all_snapshots.keys()):
        duids = all_snapshots[label]
        fuel_mix: dict = {r: {} for r in NEM_REGIONS}
        pump_load_hist: dict = {}  # track pump hydro load per region
        for duid, mw_raw in duids.items():
            info   = reg.get(duid, {})
            region = info.get("region", "") or cpid_map.get(duid, "")
            raw_fuel = info.get("fuel", "")
            fuel   = raw_fuel if (raw_fuel and raw_fuel != "Other") else _infer_fuel_from_duid(duid)
            # Negate pump-load DUIDs (positive SCADA = consuming) for all fuel logic
            mw = -mw_raw if (duid in PUMP_LOAD_DUIDS and mw_raw is not None) else mw_raw
            # Always store per-DUID history regardless of region match
            if mw is not None:
                if duid not in _duid_history:
                    _duid_history[duid] = {}
                _duid_history[duid][label] = round(mw, 1)
            if region not in NEM_REGIONS:
                region = _infer_region_from_duid(duid)
            if region not in NEM_REGIONS:
                continue
            mw_val_signed = mw if mw is not None else 0
            mw_pos = max(mw_val_signed, 0)
            fuel_mix[region][fuel] = round(fuel_mix[region].get(fuel, 0) + mw_pos, 1)
            # Track pump hydro load (negative mw on Hydro DUIDs)
            if fuel == "Hydro" and mw_val_signed < -1:
                pump_load_hist[region] = round(pump_load_hist.get(region, 0) + mw_val_signed, 1)
        # Store into _fuel_history with pump hydro
        for region in NEM_REGIONS:
            if fuel_mix[region]:
                snap = dict(fuel_mix[region])
                if pump_load_hist.get(region, 0) < -1:
                    snap["Pump Hydro"] = round(pump_load_hist[region], 1)
                _fuel_history[region][label] = snap
        loaded += 1

    logger.info(f"scrape_scada_history: loaded {loaded} time slots into fuel history")

    # Merge rooftop solar data (from dispatch history backfill) into fuel_history
    for region in NEM_REGIONS:
        for label, val in _rooftop_history.get(region, {}).items():
            if label in _fuel_history.get(region, {}):
                _fuel_history[region][label]["Rooftop Solar"] = val
    logger.info("scrape_scada_history: merged rooftop solar into fuel history")


# scrape_gen - medium speed: fuel mix from SCADA + NEM_UNITS static registry
# Refreshed every 15 min.
# ---------------------------------------------------------------------------

def scrape_gen() -> dict:
    """
    SCADA-first: every DUID in DISPATCH_UNIT_SCADA is included.
    Uses live AEMO registration list as primary source for station/fuel/capacity.
    Falls back to NEM_UNITS static registry, then fuel inference from DUID name.
    """
    logger.info("scrape_gen starting...")

    # Fetch SCADA, DispatchIS and registration list concurrently
    with ThreadPoolExecutor(max_workers=3) as ex:
        f_scada    = ex.submit(_fetch_full_scada)
        f_dispatch = ex.submit(_fetch_dispatch_is)
        f_reg      = ex.submit(_load_registration_list)
    scada         = f_scada.result()
    dispatch_text = f_dispatch.result()
    live_reg      = f_reg.result()

    # Merge: live registration list takes priority, NEM_UNITS fills any gaps
    reg = {**NEM_UNITS, **live_reg} if live_reg else NEM_UNITS
    logger.info(f"scrape_gen: using {len(live_reg)} live reg entries + {len(NEM_UNITS)} static")

    # If we got live reg data, persist any new DUIDs back into nem_units.json
    # so the static file grows over time and covers more DUIDs on future restarts
    if live_reg:
        try:
            new_entries = {k: v for k, v in live_reg.items() if k not in NEM_UNITS and v.get("fuel") and v.get("fuel") != "Other"}
            if new_entries:
                merged = {**NEM_UNITS, **new_entries}
                _json_path = _Path(__file__).parent / "nem_units.json"
                _json_path.write_text(_json.dumps(merged, indent=2, sort_keys=True))
                NEM_UNITS.update(new_entries)
                logger.info(f"scrape_gen: added {len(new_entries)} new DUIDs to nem_units.json (total now {len(merged)})")
        except Exception as _e:
            logger.debug(f"scrape_gen: nem_units.json update failed: {_e}")

    # Build DUID->region from CONNECTIONPOINTID in DISPATCH_UNIT_SOLUTION
    _CPID_REGION = {"N": "NSW1", "Q": "QLD1", "V": "VIC1", "S": "SA1", "T": "TAS1"}
    cpid_map: dict = {}
    for row in _parse_aemo(dispatch_text, "DISPATCH_UNIT_SOLUTION"):
        duid = row.get("DUID", "").strip().upper()
        cpid = row.get("CONNECTIONPOINTID", "").strip().upper()
        if duid and cpid and cpid[0] in _CPID_REGION:
            cpid_map[duid] = _CPID_REGION[cpid[0]]

    fuel_mix:   dict = {r: {} for r in NEM_REGIONS}
    nem_totals: dict = {}
    grouped:    dict = {}
    pump_load:  dict = {}  # { region: negative_mw } for pump hydro display
    unmatched_log: list = []

    for duid, mw_raw in scada.items():
        # Pump-load DUIDs report positive MW when consuming - negate for display
        mw = -mw_raw if (duid in PUMP_LOAD_DUIDS and mw_raw is not None) else mw_raw
        info     = reg.get(duid.upper(), {})
        region   = info.get("region", "") or cpid_map.get(duid.upper(), "")
        raw_fuel = info.get("fuel", "")
        fuel     = raw_fuel if (raw_fuel and raw_fuel != "Other") else _infer_fuel_from_duid(duid)
        station  = info.get("station", duid)
        capacity = info.get("capacity")

        # Last resort: infer region from DUID prefix if still unknown
        if region not in NEM_REGIONS:
            region = _infer_region_from_duid(duid)
        if region not in NEM_REGIONS:
            continue  # truly unknown region — skip

        mw_val = mw if mw is not None else 0
        # fuel_mix is for the generation chart - always positive (generation only)
        # grouped is for DUID/station display - signed (negative = pumping/charging)
        mw_for_chart = max(mw_val, 0)
        # Batteries and pump hydro show negative MW in grouped for display
        mw_pos = mw_val if fuel in ("Battery", "Hydro") else mw_for_chart

        # Log significant unclassified units so we can fix them
        if fuel == "Other" and abs(mw_val) > 50:
            logger.warning(f"OTHER_DUID: {duid} region={region} mw={mw_pos:.0f} raw_fuel={raw_fuel!r} station={station!r}")

        fuel_mix[region][fuel] = round(fuel_mix[region].get(fuel, 0) + mw_for_chart, 1)
        nem_totals[fuel]       = round(nem_totals.get(fuel, 0) + mw_for_chart, 1)
        # Track pump hydro load (negative mw on Hydro DUIDs) for gen chart
        if fuel == "Hydro" and mw_val < -1:
            pump_load[region] = round(pump_load.get(region, 0) + mw_val, 1)

        pct = round(mw_val / capacity * 100, 1) if (mw_val and capacity and capacity > 0) else None
        grouped.setdefault(region, {}).setdefault(fuel, []).append({
            "duid":     duid,
            "station":  station,
            "mw":       round(mw_val, 1),
            "capacity": capacity,
            "pct":      pct,
            "matched":  bool(info),
        })

        if not info:
            unmatched_log.append((duid, mw_val))

    # Sort units within each fuel group by MW desc
    for region in grouped:
        for fuel in grouped[region]:
            grouped[region][fuel].sort(key=lambda x: x["mw"] or 0, reverse=True)

    if unmatched_log:
        top = sorted(unmatched_log, key=lambda x: x[1] or 0, reverse=True)[:20]
        logger.info(f"scrape_gen: {len(unmatched_log)} SCADA DUIDs region-inferred (no registry entry): "
                    + ", ".join(f"{d}={mw:.0f}MW" for d, mw in top if mw and mw > 1))

    # Add real Rooftop Solar from AEMO ROOFTOP_PV/ACTUAL - 30-min intervals by state
    rooftop_pv = _scrape_rooftop_pv_latest()
    for region, mw in rooftop_pv.items():
        if mw > 0:
            fuel_mix[region]["Rooftop Solar"] = mw

    # Accumulate into in-memory history
    _update_fuel_history(fuel_mix, scada, pump_load)

    logger.info(f"scrape_gen done - {len(scada)} SCADA DUIDs, "
                f"reg={len(reg)}, buckets={sum(len(v) for v in fuel_mix.values())}")
    return {
        "timestamp":     datetime.now(timezone.utc).isoformat(),
        "fuel_mix":      fuel_mix,
        "fuel_history":  _get_fuel_history(),
        "nem_totals":    nem_totals,
        "grouped":       grouped,
        "fuel_colors":   FUEL_COLORS,
        "all_fuels":     ALL_FUELS,
        "scada_count":   len(scada),
        "reg_count":     len(reg),
    }


# scrape_slow - background fetch for generators + week-ahead
# ---------------------------------------------------------------------------

def scrape_pasa_duid_availability(source: str = "STPASA") -> dict:
    """
    Fetch PDPASA or STPASA DUID availability.
    source: "PDPASA" (next ~36h, 30-min) or "STPASA" (next ~7d, 30-min)
    Returns { duid: { interval_datetime: pasa_availability_mw } }
    """
    url_map = {"PDPASA": PDPASA_DUID_URL, "STPASA": STPASA_DUID_URL}
    base_url = url_map.get(source, STPASA_DUID_URL)

    try:
        files = _list_hrefs(base_url)
    except Exception as e:
        logger.warning(f"scrape_pasa_duid_availability({source}): listing failed: {e}")
        return {}

    if not files:
        return {}

    latest = sorted(files)[-1]
    logger.info(f"scrape_pasa_duid_availability({source}): fetching {latest}")
    try:
        text = _read_zip(latest)
    except Exception as e:
        logger.warning(f"scrape_pasa_duid_availability({source}): read failed: {e}")
        return {}

    if not text:
        return {}

    result: dict = {}  # { duid: { "YYYY-MM-DD HH:MM": pasa_avail_mw } }
    max_avail: dict = {}  # { duid: max_generation_availability_mw } — use as capacity baseline
    for row in _parse_aemo(text, "DUIDAVAILABILITY"):
        duid = row.get("DUID", "").strip()
        if not duid:
            continue
        dt_str = row.get("INTERVAL_DATETIME", "")
        avail_str = row.get("GENERATION_PASA_AVAILABILITY", "")
        max_str   = row.get("GENERATION_MAX_AVAILABILITY", "")
        if not dt_str or not avail_str:
            continue
        try:
            avail = round(float(avail_str))
            max_a = round(float(max_str)) if max_str else None
            dt = datetime.fromisoformat(dt_str.replace("/", "-"))
            label = dt.strftime("%Y-%m-%d %H:%M")
            if duid not in result:
                result[duid] = {}
            result[duid][label] = avail
            # Track the maximum of GENERATION_MAX_AVAILABILITY seen across all slots
            if max_a is not None:
                if duid not in max_avail or max_a > max_avail[duid]:
                    max_avail[duid] = max_a
        except (ValueError, TypeError):
            continue

    logger.info(f"scrape_pasa_duid_availability({source}): {len(result)} DUIDs")
    return {"slots": result, "max_avail": max_avail}


def scrape_historical_day(date_str: str) -> dict:
    """
    Fetch a full day of historical data for the D-1 page.
    Returns prices (5-min), demand, op_demand, solar, wind from DispatchIS + TradingIS.
    date_str: YYYYMMDD
    """
    from datetime import datetime as _dt, timedelta as _td
    from concurrent.futures import ThreadPoolExecutor as _TPE

    now_aest = datetime.now(AEST)
    today    = now_aest.date()

    try:
        req_date = _dt.strptime(date_str, "%Y%m%d").date()
    except ValueError:
        return {}

    # ── Step 1: Get price data (already handles DispatchIS + TradingIS fallback) ──
    prices_5m = scrape_historical_dispatch_prices(date_str)
    prices_30m_raw = {}
    try:
        from scraper import scrape_historical_prices as _shp
        prices_30m_raw = _shp(date_str)
    except Exception:
        pass

    # ── Step 2: Demand from DispatchIS REGIONSUM ──────────────────────────────
    demand:    dict = {r: {} for r in NEM_REGIONS}  # { region: { HH:MM: mw } }
    op_demand: dict = {r: {} for r in NEM_REGIONS}

    try:
        dispatch_files = _list_hrefs(DISPATCH_IS_URL)
        date_files = sorted([f for f in dispatch_files if date_str in f and "PUBLIC_DISPATCHIS" in f.upper()])
        logger.info(f"scrape_historical_day: {len(date_files)} DispatchIS files")

        for url in date_files:
            try:
                text = _read_zip(url)
                if not text: continue
                for row in _parse_aemo(text, "DISPATCH_REGIONSUM"):
                    region = row.get("REGIONID", "").strip()
                    if region not in NEM_REGIONS: continue
                    dt_str2 = row.get("SETTLEMENTDATE", "")
                    dem_str = row.get("TOTALDEMAND", "")
                    op_str  = row.get("DEMAND_AND_NONSCHEDGEN", "")
                    if not dt_str2: continue
                    try:
                        dt = datetime.fromisoformat(dt_str2.replace("/", "-")) - _td(minutes=5)
                        if dt.date() != req_date: continue
                        label = dt.strftime("%H:%M")
                        if dem_str: demand[region][label]    = round(float(dem_str), 1)
                        if op_str:  op_demand[region][label] = round(float(op_str), 1)
                    except (ValueError, TypeError):
                        continue
            except Exception as e:
                logger.debug(f"scrape_historical_day: DispatchIS error {url}: {e}")
    except Exception as e:
        logger.warning(f"scrape_historical_day: DispatchIS listing failed: {e}")

    # ── Step 3: SCADA-based fuel mix from Dispatch_SCADA files ───────────────
    # Fetch all SCADA files for the date (they're ~5min each, ~288 for a day)
    fuel_history: dict = {r: {} for r in NEM_REGIONS}  # { region: { HH:MM: {fuel: mw} } }

    try:
        scada_files = _list_hrefs(SCADA_URL)
        date_scada = sorted([f for f in scada_files if date_str in f and "PUBLIC_DISPATCHSCADA" in f.upper()])
        logger.info(f"scrape_historical_day: {len(date_scada)} SCADA files")

        def _fetch_scada(url):
            try:
                text = _read_zip(url)
                if not text: return {}
                snapshot = {}  # { duid: mw }
                dt_seen = None
                for row in _parse_aemo(text, "DISPATCH_UNIT_SCADA"):
                    duid = row.get("DUID", "").strip().upper()
                    dt_str2 = row.get("SETTLEMENTDATE", "")
                    v = row.get("SCADAVALUE", "")
                    if not duid or not dt_str2 or not v: continue
                    try:
                        dt = datetime.fromisoformat(dt_str2.replace("/", "-")) - _td(minutes=5)
                        if dt.date() != req_date: continue
                        dt_seen = dt.strftime("%H:%M")
                        snapshot[duid] = round(float(v), 1)
                    except (ValueError, TypeError): continue
                return (dt_seen, snapshot)
            except Exception:
                return None

        with _TPE(max_workers=10) as ex:
            results = list(ex.map(_fetch_scada, date_scada))

        for result in results:
            if not result or not result[0]: continue
            label, snapshot = result
            fuel_mix_snap = {r: {} for r in NEM_REGIONS}
            for duid, mw in snapshot.items():
                info = NEM_UNITS.get(duid, {})
                region = info.get("region", "")
                fuel = info.get("fuel", "") or _infer_fuel_from_duid(duid)
                if region not in NEM_REGIONS: continue
                mw_pos = max(mw, 0)
                fuel_mix_snap[region][fuel] = round(fuel_mix_snap[region].get(fuel, 0) + mw_pos, 1)
            for region in NEM_REGIONS:
                if fuel_mix_snap[region]:
                    fuel_history[region][label] = fuel_mix_snap[region]

    except Exception as e:
        logger.warning(f"scrape_historical_day: SCADA fetch failed: {e}")

    # Convert to list format
    def _to_list(d):
        return {r: [{"interval": k, **v} for k, v in sorted(s.items())] for r, s in d.items() if s}

    def _to_series(d):
        return {r: [{"interval": k, "demand": v} for k, v in sorted(s.items())] for r, s in d.items() if s}

    return {
        "date": date_str,
        "dispatch_prices_5min": prices_5m,
        "historical_prices":   prices_30m_raw,
        "demand_history":      _to_series(demand),
        "op_demand_history":   _to_series(op_demand),
        "fuel_history":        _to_list(fuel_history),
    }


MMSDM_BASE = f"{NEMWEB_BASE}/Data_Archive/Wholesale_Electricity/MMSDM"


def scrape_historical_price_averages(days: int = 90) -> dict:
    """
    Compute daily average prices for the last N days (default 90).

    Sources (in priority order):
    1. GitHub daily price files — 5-min dispatch prices, written every 5 min
       data/prices/YYYY-MM-DD.json
    2. MMSDM monthly DVD — single flat CSV, available ~5-6 weeks after month end
    3. TradingIS weekly archive — ZIP of ZIPs, available next Sunday after week end

    Returns { region: { "YYYY-MM-DD": { avg, min, max, p90, count } } }
    """
    import statistics, json, base64, os
    from datetime import timedelta as _td
    from concurrent.futures import ThreadPoolExecutor as _TPE

    now_aest  = datetime.now(AEST)
    cutoff    = (now_aest - _td(days=days)).date()
    today     = now_aest.date()

    raw: dict = {}  # { (region, "YYYY-MM-DD"): [rrp, ...] }
    days_covered = set()  # dates fully covered by GitHub files

    # ── Source 1: GitHub daily price files ───────────────────────────────────
    GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
    GH_REPO  = os.environ.get("GITHUB_REPO", "")
    if GH_TOKEN and GH_REPO:
        try:
            import requests as _req
            GH_HEADERS = {
                "Authorization": f"token {GH_TOKEN}",
                "Accept": "application/vnd.github.v3+json",
            }
            # Check dates from cutoff to today
            check_date = cutoff
            while check_date <= today:
                date_str = check_date.strftime("%Y-%m-%d")
                path     = f"data/prices/{date_str}.json"
                r = _req.get(
                    f"https://api.github.com/repos/{GH_REPO}/contents/{path}",
                    headers=GH_HEADERS, timeout=8
                )
                if r.status_code == 200:
                    day_data = json.loads(base64.b64decode(r.json()["content"]).decode())
                    day_data.pop("date", None)
                    for region, intervals in day_data.items():
                        if region not in NEM_REGIONS:
                            continue
                        for hhmm, rrp in intervals.items():
                            key = (region, date_str)
                            raw.setdefault(key, []).append(float(rrp))
                    days_covered.add(check_date)
                if check_date == today:
                    break
                # advance
                check_date = check_date + _td(days=1)
            logger.info(f"scrape_historical_price_averages: {len(days_covered)} days from GitHub files")
        except Exception as e:
            logger.warning(f"scrape_historical_price_averages: GitHub files failed: {e}")

    # Work out which months still need MMSDM/archive coverage
    months_needed = set()
    d = cutoff
    while d < today:
        if d not in days_covered:
            months_needed.add((d.year, d.month))
        if d.month == 12:
            d = d.replace(year=d.year+1, month=1, day=1)
        else:
            d = d.replace(month=d.month+1, day=1)

    if not months_needed:
        logger.info("scrape_historical_price_averages: all days covered by GitHub files")
    else:
        logger.info(f"scrape_historical_price_averages: {len(months_needed)} months need MMSDM/archive")

    def _parse_text(text: str):
        """Parse AEMO CSV text and accumulate into raw."""
        # Try both TRADINGPRICE (MMSDM) and TRADING_PRICE (TradingIS) table names
        rows = _parse_aemo(text, "TRADINGPRICE") or _parse_aemo(text, "TRADING_PRICE")
        for row in rows:
            region = row.get("REGIONID", "").strip()
            if region not in NEM_REGIONS:
                continue
            if row.get("INVALIDFLAG", "0") not in ("0", ""):
                continue
            dt_str  = row.get("SETTLEMENTDATE", "")
            rrp_str = row.get("RRP", "")
            if not dt_str or not rrp_str:
                continue
            try:
                dt       = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                day_date = dt.date()
                if day_date < cutoff or day_date >= today:
                    continue
                key = (region, dt.strftime("%Y-%m-%d"))
                raw.setdefault(key, []).append(round(float(rrp_str), 2))
            except (ValueError, TypeError):
                continue

    # ── Step 1: try MMSDM monthly DVDs ───────────────────────────────────────
    mmsdm_covered = set()  # months successfully fetched from MMSDM
    for (year, month) in sorted(months_needed):
        ym     = f"{year}{month:02d}"
        url    = (f"{MMSDM_BASE}/{year}/MMSDM_{ym}/"
                  f"MMSDM_Historical_Data_SQLLoader/DATA/"
                  f"PUBLIC_DVD_TRADINGPRICE_{ym}010000.zip")
        text = _read_zip(url)
        if text:
            logger.info(f"scrape_historical_price_averages: MMSDM {ym} OK ({len(text)//1024}KB)")
            _parse_text(text)
            mmsdm_covered.add((year, month))
        else:
            logger.info(f"scrape_historical_price_averages: MMSDM {ym} not available, will use TradingIS archive")

    # ── Step 2: for months not in MMSDM, use TradingIS weekly archive ────────
    months_needing_archive = months_needed - mmsdm_covered
    if months_needing_archive:
        try:
            all_hrefs = _list_hrefs(TRADING_ARCHIVE)
            # Find weekly ZIPs that overlap with needed months
            relevant = []
            for href in all_hrefs:
                fname = href.split('/')[-1]
                parts = fname.replace('.zip', '').split('_')
                if len(parts) >= 4:
                    try:
                        start_date = datetime.strptime(parts[-2], "%Y%m%d").date()
                        end_date   = datetime.strptime(parts[-1], "%Y%m%d").date()
                        # Include if any needed month overlaps this week
                        for (yr, mo) in months_needing_archive:
                            from calendar import monthrange
                            month_start = datetime(yr, mo, 1).date()
                            last_day    = monthrange(yr, mo)[1]
                            month_end   = datetime(yr, mo, last_day).date()
                            if start_date <= month_end and end_date >= month_start and end_date >= cutoff:
                                relevant.append(href)
                                break
                    except ValueError:
                        continue
            relevant = list(set(relevant))  # deduplicate
            logger.info(f"scrape_historical_price_averages: {len(relevant)} TradingIS archive ZIPs to fetch")

            def _fetch_archive(u):
                try: return _read_zip_of_zips(u)
                except: return ""

            with _TPE(max_workers=3) as ex:
                for text in ex.map(_fetch_archive, relevant):
                    if text:
                        _parse_text(text)
        except Exception as e:
            logger.warning(f"scrape_historical_price_averages: TradingIS archive failed: {e}")

    # ── Step 3: compute daily stats ───────────────────────────────────────────
    results: dict = {r: {} for r in NEM_REGIONS}
    for (region, day_str), prices in raw.items():
        if not prices:
            continue
        prices_sorted = sorted(prices)
        p90_idx = min(int(len(prices_sorted) * 0.90), len(prices_sorted) - 1)
        results[region][day_str] = {
            "avg":   round(statistics.mean(prices), 2),
            "min":   round(min(prices), 2),
            "max":   round(max(prices), 2),
            "p90":   round(prices_sorted[p90_idx], 2),
            "count": len(prices),
        }

    total = sum(len(v) for v in results.values())
    logger.info(f"scrape_historical_price_averages: {total} region-days computed")
    return results

    # ── Step 4: compute stats ────────────────────────────────────────────────
    results: dict = {r: {} for r in NEM_REGIONS}
    for (region, day_str), prices in raw.items():
        if not prices:
            continue
        prices_sorted = sorted(prices)
        p90_idx = min(int(len(prices_sorted) * 0.90), len(prices_sorted) - 1)
        results[region][day_str] = {
            "avg":   round(statistics.mean(prices), 2),
            "min":   round(min(prices), 2),
            "max":   round(max(prices), 2),
            "p90":   round(prices_sorted[p90_idx], 2),
            "count": len(prices),
        }

    total = sum(len(v) for v in results.values())
    logger.info(f"scrape_historical_price_averages: {total} region-days computed")
    return results


def scrape_mtpasa_outages() -> list:
    """
    Find units that are currently offline/derated OR have upcoming availability changes.
    Sources (in priority order):
      - PDPASA: current availability (30-min, most fresh)
      - STPASA: next 7 days at 30-min resolution
      - MTPASA: next 2 years at weekly change-points
    A unit appears if:
      (a) current availability < nameplate, OR
      (b) availability changes by any amount in STPASA or MTPASA horizon
    """
    from concurrent.futures import ThreadPoolExecutor

    # Fetch all three sources in parallel
    with ThreadPoolExecutor(max_workers=3) as ex:
        f_stpasa = ex.submit(scrape_pasa_duid_availability, "STPASA")
        f_pdpasa = ex.submit(scrape_pasa_duid_availability, "PDPASA")
        f_mtpasa_files = ex.submit(_list_hrefs, MTPASA_DUID_URL)

    stpasa_result = {}
    pdpasa_result = {}
    mtpasa_files = []
    try: stpasa_result = f_stpasa.result(timeout=25)
    except Exception as e: logger.warning(f"STPASA fetch failed: {e}")
    try: pdpasa_result = f_pdpasa.result(timeout=25)
    except Exception as e: logger.warning(f"PDPASA fetch failed: {e}")
    try: mtpasa_files = f_mtpasa_files.result(timeout=25)
    except Exception as e: logger.warning(f"MTPASA listing failed: {e}")

    # Unpack slots and max_avail from new return format
    stpasa_avail = stpasa_result.get("slots", {}) if isinstance(stpasa_result, dict) and "slots" in stpasa_result else stpasa_result
    pdpasa_avail = pdpasa_result.get("slots", {}) if isinstance(pdpasa_result, dict) and "slots" in pdpasa_result else pdpasa_result
    stpasa_max   = stpasa_result.get("max_avail", {}) if isinstance(stpasa_result, dict) else {}
    pdpasa_max   = pdpasa_result.get("max_avail", {}) if isinstance(pdpasa_result, dict) else {}

    # Parse MTPASA file
    duid_data: dict = {}  # { duid: { "YYYY/MM/DD": {avail, state} } }
    if mtpasa_files:
        latest = sorted(mtpasa_files)[-1]
        logger.info(f"scrape_mtpasa_outages: fetching {latest}")
        try:
            text = _read_zip(latest)
            if text:
                for row in _parse_aemo(text, "MTPASA_DUIDAVAILABILITY"):
                    duid = row.get("DUID", "").strip()
                    if not duid:
                        continue
                    unit = NEM_UNITS.get(duid, {})
                    if not unit or not unit.get("capacity"):
                        continue
                    day = row.get("DAY", "").strip()[:10]
                    try:
                        avail = float(row.get("PASAAVAILABILITY") or 0)
                    except ValueError:
                        avail = 0.0
                    state = row.get("UNITSTATE", "").strip() or "Unknown"
                    if duid not in duid_data:
                        duid_data[duid] = {}
                    duid_data[duid][day] = {"avail": avail, "state": state}
        except Exception as e:
            logger.warning(f"MTPASA parse failed: {e}")

    now_aest  = datetime.now(AEST)
    today_str = now_aest.strftime("%Y/%m/%d")
    now_label = now_aest.strftime("%Y-%m-%d %H:%M")

    # Combined DUID set — any DUID in any source
    all_duids = set(duid_data.keys()) | set(stpasa_avail.keys()) | set(pdpasa_avail.keys())

    results = []

    # ── Strategy ──────────────────────────────────────────────────────────────
    # For each scheduled DUID, find the MINIMUM availability across:
    #   - All PDPASA slots (next ~36h, 30-min) — coal only
    #   - All STPASA daily 00:30 snapshots (next ~7d) — all scheduled fuels
    #   - MTPASA change-points (full horizon) — fallback
    # This catches: current outages, outages starting later today,
    # outages starting in the next 7 days.
    # Return date = first future slot/day where avail recovers above 70%.
    # ──────────────────────────────────────────────────────────────────────────
    THRESHOLD = 0.70
    SCHEDULED_FUELS = {"Black Coal", "Brown Coal", "Gas", "Hydro", "Battery", "Liquid", "Other"}

    # Pre-build STPASA daily 00:30 snapshots per DUID
    stpasa_daily: dict = {}  # { duid: { "YYYY-MM-DD": avail } }
    for duid, slots in stpasa_avail.items():
        daily = {}
        for slot, avail in slots.items():
            if slot[11:] == "00:30":
                daily[slot[:10]] = avail
        if daily:
            stpasa_daily[duid] = daily

    all_duids = set(stpasa_daily.keys()) | set(pdpasa_avail.keys()) | set(duid_data.keys())

    for duid in all_duids:
        unit = NEM_UNITS.get(duid, {})
        capacity = unit.get("capacity") or 0
        if capacity <= 0:
            continue
        fuel = unit.get("fuel", "Other")
        if fuel not in SCHEDULED_FUELS:
            continue

        threshold_mw = capacity * THRESHOLD
        mtpasa_days = duid_data.get(duid, {})
        sorted_days = sorted(mtpasa_days.keys())
        pdpasa_slots = pdpasa_avail.get(duid, {})
        stpasa_slots_raw = stpasa_avail.get(duid, {})
        daily_st = stpasa_daily.get(duid, {})

        # ── Find minimum availability and when it occurs ───────────────────────
        min_avail = capacity
        min_avail_date = None
        min_avail_source = None
        state_now = "Unknown"

        # PDPASA — coal only: find FIRST slot below threshold and overall minimum
        pdpasa_first_below = None
        pdpasa_first_avail = None   # avail at very first PDPASA slot (regardless of threshold)
        if fuel in ("Black Coal", "Brown Coal") and pdpasa_slots:
            first_pd_slot = sorted(pdpasa_slots.keys())[0]
            pdpasa_first_avail = pdpasa_slots[first_pd_slot]
            for slot in sorted(pdpasa_slots.keys()):
                avail = pdpasa_slots[slot]
                if avail < threshold_mw and pdpasa_first_below is None:
                    pdpasa_first_below = slot[:10].replace("-", "/")
                if avail < min_avail:
                    min_avail = avail
                    min_avail_date = slot[:10].replace("-", "/")
                    min_avail_source = "PDPASA"

        # STPASA — all fuels: find FIRST day below threshold and overall minimum
        stpasa_first_below = None
        stpasa_first_avail = None   # avail at very first STPASA daily slot
        if daily_st:
            first_st_day = sorted(daily_st.keys())[0]
            stpasa_first_avail = daily_st[first_st_day]
        for day in sorted(daily_st.keys()):
            avail = daily_st[day]
            if avail < threshold_mw and stpasa_first_below is None:
                stpasa_first_below = day.replace("-", "/")
            if avail < min_avail:
                min_avail = avail
                min_avail_date = day.replace("-", "/")
                min_avail_source = "STPASA"

        # is_current: first PDPASA slot (coal) or first STPASA slot (gas/hydro) is below threshold
        # MTPASA-only units are never current — they are future planned outages.
        # Battery excluded — 0MW at night is normal dispatch, not an outage.
        is_current = False
        if fuel in ("Black Coal", "Brown Coal"):
            if pdpasa_first_avail is not None:
                is_current = pdpasa_first_avail < threshold_mw
        elif fuel in ("Gas", "Hydro", "Liquid", "Other"):
            if stpasa_first_avail is not None:
                is_current = stpasa_first_avail < threshold_mw

        # MTPASA — fallback: scan all change-points
        mtpasa_first_below = None
        for d in sorted_days:
            entry = mtpasa_days[d]
            if entry["avail"] < threshold_mw and mtpasa_first_below is None:
                mtpasa_first_below = d
            if entry["avail"] < min_avail:
                min_avail = entry["avail"]
                min_avail_date = d
                min_avail_source = "MTPASA"
                state_now = entry["state"]

        # outage_start = earliest date across all sources where avail drops below threshold
        candidates = [d for d in [pdpasa_first_below, stpasa_first_below, mtpasa_first_below] if d]
        outage_start = min(candidates) if candidates else min_avail_date

        # Skip if always above threshold
        if min_avail >= threshold_mw:
            continue

        if state_now in ("Mothballed", "Retired", "Decommissioned"):
            continue

        # ── Find return date: first future date where avail >= threshold ─────────
        # Scan from outage_start onwards — first date where avail recovers above 70%
        return_date = None
        return_source = None
        outage_start_cmp = outage_start.replace("/", "-") if outage_start else ""

        # STPASA daily snapshots from outage_start onwards
        for day in sorted(daily_st.keys()):
            if day < outage_start_cmp:
                continue
            if daily_st[day] >= threshold_mw:
                return_date = day.replace("-", "/")
                return_source = "STPASA"
                break

        # MTPASA: scan all change-points from outage_start onwards
        if not return_date:
            for d in sorted_days:
                d_cmp = d.replace("/", "-")
                if d_cmp < outage_start_cmp:
                    continue
                if mtpasa_days[d]["avail"] >= threshold_mw:
                    return_date = d
                    return_source = "MTPASA"
                    break

        # ── Outage minimum: minimum avail between outage_start and return_date ──
        # This is what the card displays — how low it drops during the outage window
        return_label = return_date.replace("/", "-") if return_date else "9999-99-99"
        outage_start_label = outage_start.replace("/", "-") if outage_start else ""
        outage_min = min_avail  # default to global min

        if outage_start_label:
            window_min = capacity
            # Scan PDPASA slots within outage window (coal only)
            if fuel in ("Black Coal", "Brown Coal"):
                for slot, avail in pdpasa_slots.items():
                    d = slot[:10]
                    if d >= outage_start_label and d <= return_label:
                        if avail < window_min:
                            window_min = avail
            # Scan STPASA daily slots within outage window
            for day, avail in daily_st.items():
                if day >= outage_start_label and day <= return_label:
                    if avail < window_min:
                        window_min = avail
            # Scan MTPASA change-points within outage window
            for d in sorted_days:
                if d >= outage_start and d <= return_label.replace("-", "/"):
                    if mtpasa_days[d]["avail"] < window_min:
                        window_min = mtpasa_days[d]["avail"]
            outage_min = window_min

        # ── State label ───────────────────────────────────────────────────────
        if min_avail == 0:
            label = "Forced" if state_now == "Forced" else "Planned" if state_now == "Planned" else "Offline"
        else:
            label = "Derated"

        # avail_now = current reading from first PDPASA/STPASA slot (for offline/derated toggle)
        avail_now = None
        if fuel in ("Black Coal", "Brown Coal") and pdpasa_first_avail is not None:
            avail_now = pdpasa_first_avail
        elif stpasa_first_avail is not None:
            avail_now = stpasa_first_avail
        if avail_now is None:
            avail_now = min_avail  # fallback

        results.append({
            "duid":          duid,
            "station":       unit.get("station", duid),
            "fuel":          fuel,
            "region":        unit.get("region", ""),
            "capacity":      int(capacity),
            "avail_today":   int(outage_min),
            "avail_now":     int(avail_now),   # current reading for offline/derated toggle
            "avail_source":  min_avail_source,
            "state":         label,
            "pasa_state":    state_now,
            "change_mw":     int(min_avail - capacity),
            "outage_start":  outage_start,
            "return_date":   return_date,
            "return_source": return_source,
            "is_current":    is_current,
        })

    results.sort(key=lambda x: x["change_mw"])
    logger.info(f"scrape_mtpasa_outages: {len(results)} units (current+upcoming, MTPASA+STPASA+PDPASA)")
    return results


def scrape_weather() -> dict:
    """
    Fetch 7-day daily forecast using Open-Meteo API (free, no key, uses BOM ACCESS-G model).
    Returns { region: { name, days: [{day_of_week, date_label, temp_max, temp_min}] } }
    """
    # Exact BOM station coordinates for each NEM region
    LOCATIONS = {
        "QLD1": {"name": "Archerfield",       "lat": -27.57,   "lon": 153.01},
        "NSW1": {"name": "Bankstown",          "lat": -33.92,   "lon": 150.98},
        "VIC1": {"name": "Melbourne (Olympic Park)", "lat": -37.8136, "lon": 144.9631},
        "SA1":  {"name": "Adelaide (West Terrace)",  "lat": -34.95,   "lon": 138.52},
        "TAS1": {"name": "Hobart",             "lat": -42.8821, "lon": 147.3272},
    }

    result = {}
    session = requests.Session()
    session.headers.update({"User-Agent": "nem-dashboard/1.0"})

    for region, info in LOCATIONS.items():
        try:
            url = (
                f"https://api.open-meteo.com/v1/forecast"
                f"?latitude={info['lat']}&longitude={info['lon']}"
                f"&daily=temperature_2m_max,temperature_2m_min"
                f"&timezone=Australia%2FSydney&forecast_days=7"
            )
            r = session.get(url, timeout=10)
            if r.status_code != 200:
                logger.warning(f"scrape_weather: {region} status {r.status_code}")
                continue
            j = r.json()
            daily = j.get("daily", {})
            dates    = daily.get("time", [])
            temp_max = daily.get("temperature_2m_max", [])
            temp_min = daily.get("temperature_2m_min", [])

            days = []
            for i, date_str in enumerate(dates[:7]):
                try:
                    from datetime import datetime as _dt
                    dt = _dt.strptime(date_str, "%Y-%m-%d")
                    day_of_week = dt.strftime("%a")
                    date_label  = dt.strftime("%-d %b")
                except Exception:
                    day_of_week = ""
                    date_label  = date_str
                days.append({
                    "date":        date_str,
                    "day_of_week": day_of_week,
                    "date_label":  date_label,
                    "temp_max":    round(temp_max[i], 1) if i < len(temp_max) and temp_max[i] is not None else None,
                    "temp_min":    round(temp_min[i], 1) if i < len(temp_min) and temp_min[i] is not None else None,
                })

            if days:
                result[region] = {"name": info["name"], "days": days}
                logger.info(f"scrape_weather: {region} ({info['name']}) got {len(days)} days")
            else:
                logger.warning(f"scrape_weather: {region} no days in response")
        except Exception as e:
            logger.warning(f"scrape_weather: {region} error: {e}")

    return result


STTM_BASE   = f"{NEMWEB_BASE}/REPORTS/CURRENT/STTM"
VICGAS_BASE = f"{NEMWEB_BASE}/REPORTS/CURRENT/VicGas"

# Facility type classification for STTM demand breakdown
def _sttm_facility_type(facility_name: str) -> str:
    """Classify an STTM facility into a demand segment."""
    n = facility_name.upper()
    if any(x in n for x in ["POWER", "GENERATION", "GENERAT", "ENERGY", "TURBINE", "GAS FIRED"]):
        return "Generation"
    if any(x in n for x in ["DISTRIBUTION", "DISTRIB", "NETWORK", "RETAIL"]):
        return "Distribution"
    if any(x in n for x in ["INDUSTRIAL", "INDUSTRY", "MANUFACTURING", "PROCESS"]):
        return "Industrial"
    if any(x in n for x in ["PIPELINE", "TRANSMISSION", "HAUL"]):
        return "Pipeline"
    return "Other"


def scrape_gas(days: int = 14) -> dict:
    """
    Scrape gas market data:
    - STTM: ex ante prices + scheduled withdrawals for Adelaide, Brisbane, Sydney
      Source: nemweb REPORTS/CURRENT/STTM/CURRENTDAY.ZIP + DAY01..DAY14.ZIP
    - VicGas DWGM: indicative intraday price + scheduled price + withdrawals + history
      Source: nemweb REPORTS/CURRENT/VicGas/ (flat CSVs)

    Returns {
      "sttm": {
        "today": { hub: { "price": float, "demand_tj": float, "gas_date": str } },
        "history": [ { "gas_date": str, hub: { "price": float, "demand_tj": float } } ]  # 14 days
      },
      "vicgas": {
        "today_price": float,          # scheduled price $/GJ
        "intraday": [ {"time": "HH:MM", "price": float} ],  # indicative, 30-min
        "today_demand_tj": float,      # scheduled withdrawals TJ
        "history": [ {"gas_date": str, "price": float, "demand_tj": float} ]  # 14 days
      },
      "last_updated": isoformat
    }
    """
    from concurrent.futures import ThreadPoolExecutor as _TPE

    result = {
        "sttm":    {"today": {}, "history": []},
        "vicgas":  {"today_price": None, "intraday": [], "today_demand_tj": None, "history": []},
        "last_updated": datetime.now(timezone.utc).isoformat(),
    }

    STTM_HUBS = ["Adelaide", "Brisbane", "Sydney"]

    # ── Helper: parse STTM ZIP (CURRENTDAY or DAY##) ─────────────────────────
    def _parse_sttm_zip(zip_bytes: bytes) -> dict:
        """Parse STTM ZIP → { hub: {price, demand_tj, gas_date, facilities: {name: {type, tj}}} }"""
        out = {}
        try:
            with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
                names = z.namelist()
                # int651 = ex ante market price, int652 = ex ante schedule quantity
                price_csvs = [n for n in names if "int651" in n.lower() and n.endswith(".csv")]
                qty_csvs   = [n for n in names if "int652" in n.lower() and n.endswith(".csv")]

                # Parse prices (one row per hub per schedule run — take latest)
                for csv_name in sorted(price_csvs):
                    with z.open(csv_name) as f:
                        reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8", errors="replace"))
                        for row in reader:
                            hub_name = (row.get("hub_name") or "").strip().title()
                            price    = (row.get("ex_ante_market_price") or "").strip()
                            gas_date_raw = (row.get("gas_date") or "").strip()
                            if hub_name and price:
                                try:
                                    from datetime import datetime as _dt
                                    try:
                                        gas_date = _dt.strptime(gas_date_raw, "%d %b %Y").strftime("%Y-%m-%d")
                                    except ValueError:
                                        gas_date = gas_date_raw[:10]  # fallback
                                    out.setdefault(hub_name, {})["price"]    = round(float(price), 4)
                                    out.setdefault(hub_name, {})["gas_date"] = gas_date
                                except (ValueError, TypeError):
                                    pass

                # Parse withdrawals — flow_direction "F" = from hub (withdrawal/demand)
                for csv_name in sorted(qty_csvs):
                    with z.open(csv_name) as f:
                        reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8", errors="replace"))
                        for row in reader:
                            hub_name = (row.get("hub_name") or "").strip().title()
                            facility = (row.get("facility_name") or "").strip()
                            qty      = (row.get("scheduled_qty") or "").strip()
                            direction = (row.get("flow_direction") or "").strip().upper()
                            if hub_name and qty and direction == "F":  # F = from hub = withdrawal
                                try:
                                    tj = round(float(qty) / 1000, 3)  # GJ → TJ
                                    ftype = _sttm_facility_type(facility)
                                    out.setdefault(hub_name, {}).setdefault("facilities", {})
                                    out[hub_name]["facilities"][facility] = {"type": ftype, "tj": tj}
                                    out[hub_name]["demand_tj"] = round(
                                        out[hub_name].get("demand_tj", 0) + tj, 3
                                    )
                                except (ValueError, TypeError):
                                    pass
        except Exception as e:
            logger.warning(f"scrape_gas: STTM ZIP parse failed: {e}")
        return out

    # ── STTM: fetch CURRENTDAY ZIP only ──────────────────────────────────────
    # CURRENTDAY.ZIP contains:
    #   int651 — ex ante market prices (today + last 8 days = 9 days, 3 hubs each)
    #   int652 — ex ante schedule quantities (today's demand only, flow_direction F)
    # No daily demand history available from STTM sources.

    def _fetch_sttm_current():
        url = f"{STTM_BASE}/CURRENTDAY.ZIP"
        r = _get(url, timeout=30)
        if not r:
            return {}
        return r.content

    sttm_bytes = {}
    with _TPE(max_workers=1) as ex:
        sttm_bytes = ex.submit(_fetch_sttm_current).result(timeout=45)

    if sttm_bytes:
        try:
            history_prices = {}  # { "YYYY-MM-DD": { hub: price } }
            today_qty      = {}  # { hub: demand_tj }
            today_date_raw = {}  # { hub: gas_date str for today }

            with zipfile.ZipFile(io.BytesIO(sttm_bytes)) as z:
                names = z.namelist()

                # int651 — ex ante prices, multiple rows per hub per day
                # Take the latest schedule_identifier (highest number) per hub per date
                int651_best = {}  # { (hub, date_raw): (sched_id, price) }
                for fname in sorted(n for n in names if "int651" in n.lower()):
                    with z.open(fname) as f:
                        reader = csv.DictReader(io.TextIOWrapper(f, errors="replace"))
                        for row in reader:
                            hub      = (row.get("hub_name") or "").strip().title()
                            date_raw = (row.get("gas_date") or "").strip()
                            price    = (row.get("ex_ante_market_price") or "").strip()
                            sched_id = int(row.get("schedule_identifier") or 0)
                            if hub and date_raw and price:
                                key = (hub, date_raw)
                                if key not in int651_best or sched_id > int651_best[key][0]:
                                    int651_best[key] = (sched_id, price)

                for (hub, date_raw), (_, price) in int651_best.items():
                    try:
                        from datetime import datetime as _dt
                        gas_date = _dt.strptime(date_raw, "%d %b %Y").strftime("%Y-%m-%d")
                        history_prices.setdefault(gas_date, {})[hub] = round(float(price), 4)
                    except (ValueError, TypeError):
                        pass

                # int652 — today's withdrawals: take the LATEST schedule run per hub
                # Multiple int652 files = multiple schedule runs; use highest schedule_identifier
                qty_best = {}  # { hub: (sched_id, total_qty) }
                for fname in sorted(n for n in names if "int652" in n.lower()):
                    with z.open(fname) as f:
                        reader = csv.DictReader(io.TextIOWrapper(f, errors="replace"))
                        # Sum F-direction qty per hub within this schedule file
                        sched_qty  = {}
                        sched_id   = 0
                        date_raw_h = {}
                        for row in reader:
                            hub       = (row.get("hub_name") or "").strip().title()
                            direction = (row.get("flow_direction") or "").strip().upper()
                            qty       = (row.get("scheduled_qty") or "").strip()
                            date_raw  = (row.get("gas_date") or "").strip()
                            sid       = int(row.get("schedule_identifier") or 0)
                            if hub and direction == "F" and qty:
                                try:
                                    sched_qty[hub]  = sched_qty.get(hub, 0.0) + float(qty)
                                    date_raw_h[hub] = date_raw
                                    sched_id = max(sched_id, sid)
                                except (ValueError, TypeError):
                                    pass
                        # Keep this file's totals only if schedule_id is newer
                        for hub, total in sched_qty.items():
                            if hub not in qty_best or sched_id > qty_best[hub][0]:
                                qty_best[hub] = (sched_id, total)
                                today_date_raw[hub] = date_raw_h.get(hub, "")

                today_qty = {hub: v[1] for hub, v in qty_best.items()}

            # Determine today's date (most recent in int651)
            from datetime import datetime as _dt
            sorted_dates = sorted(history_prices.keys())
            today_iso = sorted_dates[-1] if sorted_dates else None

            # Populate today
            if today_iso:
                for hub in STTM_HUBS:
                    hub_price = history_prices.get(today_iso, {}).get(hub)
                    demand_tj = round(today_qty.get(hub, 0) / 1000, 3) if hub in today_qty else None
                    result["sttm"]["today"][hub] = {
                        "price":    hub_price,
                        "demand_tj": demand_tj,
                        "gas_date": today_iso,
                    }

            # Populate history (all dates, sorted ascending, last `days` entries)
            for gas_date in sorted_dates[-days:]:
                entry = {"gas_date": gas_date}
                for hub in STTM_HUBS:
                    p = history_prices.get(gas_date, {}).get(hub)
                    if p is not None:
                        entry[hub] = {"price": p}
                if len(entry) > 1:
                    result["sttm"]["history"].append(entry)

            logger.info(f"scrape_gas: STTM history {len(result['sttm']['history'])} days, today={today_iso}")
        except Exception as e:
            logger.warning(f"scrape_gas: STTM parse failed: {e}")

    # ── VicGas: fetch flat CSVs ───────────────────────────────────────────────
    def _fetch_vicgas_csv(filename: str) -> str:
        url = f"{VICGAS_BASE}/{filename}"
        r = _get(url, timeout=15)
        return r.text if r else ""

    vicgas_files = [
        "INT037B_V4_INDICATIVE_MKT_PRICE_1.CSV",      # intraday indicative price
        "INT041_V4_MARKET_AND_REFERENCE_PRICES_1.CSV", # 14-day price history
        "INT050_V4_SCHED_WITHDRAWALS_1.CSV",           # today's withdrawals by zone
    ]
    with _TPE(max_workers=4) as ex:
        vicgas_texts = {f: t for f, t in zip(vicgas_files, ex.map(_fetch_vicgas_csv, vicgas_files))}

    # INT037B — intraday indicative price
    # Columns: demand_type_name, price_value_gst_ex, transmission_group_id, schedule_type_id,
    #          transmission_id, gas_date, approval_datetime, current_date
    try:
        text = vicgas_texts.get("INT037B_V4_INDICATIVE_MKT_PRICE_1.CSV", "")
        if text:
            reader = csv.DictReader(io.StringIO(text))
            pts = []
            for row in reader:
                demand_type = (row.get("demand_type_name") or "").strip()
                price       = (row.get("price_value_gst_ex") or "").strip()
                approval    = (row.get("approval_datetime") or "").strip()
                if demand_type.lower() == "normal" and price and approval:
                    try:
                        from datetime import datetime as _dt
                        dt = _dt.strptime(approval, "%d %b %Y %H:%M:%S")
                        pts.append({"time": dt.strftime("%H:%M"), "price": round(float(price), 4)})
                    except (ValueError, TypeError):
                        pass
            result["vicgas"]["intraday"] = sorted(pts, key=lambda x: x["time"])
            logger.info(f"scrape_gas: VicGas intraday {len(pts)} pts")
    except Exception as e:
        logger.warning(f"scrape_gas: INT037B parse failed: {e}")

    # INT041 — 14-day price history
    # Columns: gas_date, price_bod_gst_ex, price_10am_gst_ex, ..., imb_wtd_ave_price_gst_ex, current_date
    try:
        text = vicgas_texts.get("INT041_V4_MARKET_AND_REFERENCE_PRICES_1.CSV", "")
        if text:
            from datetime import datetime as _dt
            reader = csv.DictReader(io.StringIO(text))
            hist = []
            for row in reader:
                date_raw = (row.get("gas_date") or "").strip()
                price    = (row.get("price_bod_gst_ex") or "").strip()
                if date_raw and price:
                    try:
                        gas_date = _dt.strptime(date_raw, "%d %b %Y").strftime("%Y-%m-%d")
                        hist.append({"gas_date": gas_date, "price": round(float(price), 4)})
                    except (ValueError, TypeError):
                        pass
            hist.sort(key=lambda x: x["gas_date"])
            result["vicgas"]["history"] = hist[-days:]
            # Today's price = last row
            if hist:
                result["vicgas"]["today_price"] = hist[-1]["price"]
            logger.info(f"scrape_gas: VicGas history {len(result['vicgas']['history'])} days")
    except Exception as e:
        logger.warning(f"scrape_gas: INT041 parse failed: {e}")

    # INT050 — today's scheduled withdrawals by zone (filter to today's date only)
    # Columns: gas_date, withdrawal_zone_name, scheduled_qty, transmission_id, current_date
    try:
        text = vicgas_texts.get("INT050_V4_SCHED_WITHDRAWALS_1.CSV", "")
        if text:
            from datetime import datetime as _dt, date as _date
            today_str = _date.today().strftime("%d %b %Y")  # e.g. "29 Mar 2026"
            reader = csv.DictReader(io.StringIO(text))
            total = 0.0
            for row in reader:
                row_date = (row.get("gas_date") or "").strip()
                qty      = (row.get("scheduled_qty") or "").strip()
                if row_date == today_str and qty:
                    try:
                        total += float(qty)
                    except (ValueError, TypeError):
                        pass
            result["vicgas"]["today_demand_tj"] = round(total / 1000, 3) if total else None
            logger.info(f"scrape_gas: VicGas demand {result['vicgas']['today_demand_tj']} TJ")
    except Exception as e:
        logger.warning(f"scrape_gas: INT050 parse failed: {e}")

    # ── VicGas demand history: AEMO DWGM Prices and Demand Excel ────────────
    # Demand sheet columns: Gas_Date, System Demand, GPG, Total Demand (already in TJ)
    try:
        import openpyxl
        excel_url = "https://www.aemo.com.au/-/media/files/gas/dwgm/dwgm-prices-and-demand.xlsx"
        r = _get(excel_url, timeout=60)
        if r and len(r.content) > 10000:
            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            demand_sheet = next((ws for ws in wb.worksheets if ws.title == "Demand"), None)
            if demand_sheet:
                from datetime import datetime as _dt, date as _date
                demand_hist = []
                header_done = False
                date_col = demand_col = gpg_col = None
                for row in demand_sheet.iter_rows(values_only=True):
                    if not header_done:
                        # Header row
                        headers = [str(v).lower().strip() if v else "" for v in row]
                        date_col   = next((i for i, h in enumerate(headers) if "date" in h), None)
                        demand_col = next((i for i, h in enumerate(headers) if "system demand" in h), None)
                        gpg_col    = next((i for i, h in enumerate(headers) if "gpg" in h), None)
                        header_done = True
                        continue
                    if date_col is None or demand_col is None:
                        continue
                    dv  = row[date_col]
                    dem = row[demand_col]
                    if not dv or not dem:
                        continue
                    try:
                        gas_date = dv.strftime("%Y-%m-%d") if hasattr(dv, "strftime") else str(dv)[:10]
                        entry = {"gas_date": gas_date, "demand_tj": round(float(dem), 3)}
                        if gpg_col is not None and row[gpg_col]:
                            entry["gpg_tj"] = round(float(row[gpg_col]), 3)
                        demand_hist.append(entry)
                    except (ValueError, TypeError):
                        pass
                wb.close()

                # Last `days` entries are already in date order (file is chronological)
                demand_by_date = {d["gas_date"]: d for d in demand_hist[-days:]}

                # Merge into vicgas history price entries
                for entry in result["vicgas"]["history"]:
                    d = entry["gas_date"]
                    if d in demand_by_date:
                        entry["demand_tj"] = demand_by_date[d]["demand_tj"]
                        if "gpg_tj" in demand_by_date[d]:
                            entry["gpg_tj"] = demand_by_date[d]["gpg_tj"]

                # Today's demand fallback (INT050 is more current)
                today_iso = _date.today().strftime("%Y-%m-%d")
                if result["vicgas"]["today_demand_tj"] is None and today_iso in demand_by_date:
                    result["vicgas"]["today_demand_tj"] = demand_by_date[today_iso]["demand_tj"]

                logger.info(f"scrape_gas: DWGM Excel {len(demand_hist)} total days, last={demand_hist[-1]['gas_date'] if demand_hist else 'none'}")
        else:
            logger.warning("scrape_gas: DWGM Excel fetch failed or empty")
    except Exception as e:
        logger.warning(f"scrape_gas: DWGM Excel demand failed: {e}")

    logger.info(
        f"scrape_gas done — STTM today hubs={list(result['sttm']['today'].keys())} "
        f"history={len(result['sttm']['history'])} days, "
        f"VicGas price={result['vicgas']['today_price']} "
        f"intraday={len(result['vicgas']['intraday'])} pts"
    )
    return result


def scrape_gbb() -> dict:
    """
    Scrape GasBBActualFlowStorageLast31.CSV from NEMWeb.
    Returns storage levels (31-day history) and state supply/demand summary.
    """
    url = "https://www.nemweb.com.au/Reports/Current/GBB/GasBBActualFlowStorageLast31.CSV"
    result = {
        "storage": {},       # { facility_name: [{gas_date, held_tj, demand_tj, supply_tj}] }
        "state_summary": {}, # { state: {supply, demand, net} } for latest date
        "latest_date": None,
        "last_updated": datetime.now(timezone.utc).isoformat(),
    }
    try:
        r = _get(url, timeout=30)
        if not r:
            logger.warning("scrape_gbb: fetch failed")
            return result

        rows = list(csv.DictReader(r.text.splitlines()))
        all_dates = sorted({row["GasDate"] for row in rows})
        latest_date = all_dates[-1] if all_dates else None
        result["latest_date"] = latest_date

        # ── Storage facilities ──────────────────────────────────────────────
        stor_rows = [row for row in rows if row.get("FacilityType") == "STOR"]
        storage_hist = {}  # { facility: { date: {held, demand, supply} } }
        for row in stor_rows:
            name     = row["FacilityName"]
            gas_date = row["GasDate"].replace("/", "-")  # → YYYY-MM-DD
            held     = row.get("HeldInStorage", "")
            demand   = row.get("Demand", "")
            supply   = row.get("Supply", "")
            if not held:
                continue
            try:
                storage_hist.setdefault(name, {})[gas_date] = {
                    "held_tj":   round(float(held), 1),
                    "demand_tj": round(float(demand), 1) if demand else 0,
                    "supply_tj": round(float(supply), 1) if supply else 0,
                }
            except (ValueError, TypeError):
                pass

        # Convert to sorted lists
        for name, dates in storage_hist.items():
            result["storage"][name] = [
                {"gas_date": d, **v}
                for d, v in sorted(dates.items())
            ]

        # ── State summary for latest date ───────────────────────────────────
        latest_rows = [row for row in rows if row["GasDate"] == latest_date]
        state_agg = {}
        for row in latest_rows:
            st = row.get("State", "")
            if not st:
                continue
            state_agg.setdefault(st, {"supply": 0.0, "demand": 0.0})
            try:
                state_agg[st]["supply"] += float(row.get("Supply") or 0)
                state_agg[st]["demand"] += float(row.get("Demand") or 0)
            except (ValueError, TypeError):
                pass

        for st, vals in state_agg.items():
            s, d = round(vals["supply"], 1), round(vals["demand"], 1)
            result["state_summary"][st] = {
                "supply": s, "demand": d, "net": round(s - d, 1)
            }

        # ── Demand by sector: GPG, LNG Export, Large Industrial, Residential ─
        sector_hist = {}  # { sector: { date: demand_tj } }
        for row in rows:
            ft     = row.get("FacilityType", "")
            gd     = row.get("GasDate", "").replace("/", "-")
            demand = float(row.get("Demand") or 0)
            if not gd or demand == 0:
                continue
            if ft == "BBGPG":
                key = "Gas Power Generation"
            elif ft == "LNGEXPORT":
                key = "LNG Export"
            elif ft == "BBLARGE":
                key = "Large Industrial"
            elif ft == "PIPE":
                key = "Residential & Commercial"
            else:
                continue
            sector_hist.setdefault(key, {}).setdefault(gd, 0.0)
            sector_hist[key][gd] = round(sector_hist[key][gd] + demand, 1)

        result["demand_by_sector"] = {}
        for key, dates in sector_hist.items():
            sorted_dates = sorted(dates.keys())[-14:]
            result["demand_by_sector"][key] = [
                {"gas_date": d, "demand_tj": dates[d]}
                for d in sorted_dates
            ]

        # ── Production by region: VIC/SA individually, QLD+NT aggregated ──────
        # Key facilities to show individually
        VIC_PROD = {'Longford', 'Otway', 'Orbost', 'ATHENA', 'Lang Lang'}
        SA_PROD  = {'Moomba'}
        NT_PROD  = {'Mereenie', 'Palm Valley', 'Yelcherr'}

        prod_rows = [row for row in rows if row.get("FacilityType") == "PROD"]
        prod_hist = {}  # { series_name: { date: supply_tj } }

        for row in prod_rows:
            name   = row.get("FacilityName", "")
            st     = row.get("State", "")
            gd     = row.get("GasDate", "").replace("/", "-")
            supply = float(row.get("Supply") or 0)
            if not gd or supply == 0:
                continue

            if name in VIC_PROD:
                key = f"VIC: {name}"
            elif name in SA_PROD:
                key = "SA: Moomba"
            elif st == "QLD":
                key = "QLD (CSG+LNG)"
            elif st == "NT":
                key = "NT"
            else:
                continue

            prod_hist.setdefault(key, {}).setdefault(gd, 0.0)
            prod_hist[key][gd] = round(prod_hist[key][gd] + supply, 1)

        # Convert to sorted lists, last 14 days
        result["production_history"] = {}
        for key, dates in prod_hist.items():
            sorted_dates = sorted(dates.keys())[-14:]
            result["production_history"][key] = [
                {"gas_date": d, "supply_tj": dates[d]}
                for d in sorted_dates
            ]

        logger.info(
            f"scrape_gbb: {len(rows)} rows, {len(all_dates)} dates, "
            f"storage={list(result['storage'].keys())}, latest={latest_date}"
        )
    except Exception as e:
        logger.warning(f"scrape_gbb: failed: {e}")

    return result


def scrape_slow() -> dict:
    """
    Combined slow scrape: STPASA demand forecast + BOM weather.
    Called by _run_slow in main.py every 30 minutes.
    Gas data is fetched on-demand via /api/gas endpoint.
    """
    logger.info("scrape_slow starting...")
    result = {
        "timestamp":     datetime.now(timezone.utc).isoformat(),
        "stpasa_demand": {},
        "fuel_colors":   FUEL_COLORS,
        "all_fuels":     ALL_FUELS,
        "fuel_mix_today": {},
        "weather":       {},
    }

    # ST PASA demand forecast
    try:
        result["stpasa_demand"] = scrape_stpasa_demand()
    except Exception as e:
        logger.warning(f"scrape_slow: stpasa_demand failed: {e}")

    # BOM weather forecast
    try:
        result["weather"] = scrape_weather()
    except Exception as e:
        logger.warning(f"scrape_slow: weather failed: {e}")

    logger.info("scrape_slow done")
    return result
