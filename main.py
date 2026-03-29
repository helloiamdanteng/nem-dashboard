"""
NEM Dashboard - FastAPI backend
Fast cache: prices, demand, gen, IC, Origin — refreshed every 5 min
Slow cache: all generators, ST PASA week-ahead — refreshed every 30 min
"""

import asyncio
import io
import logging
import time
import traceback
from datetime import datetime, timezone
from pathlib import Path
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

from scraper import scrape_all, scrape_gen, scrape_slow, scrape_scada_history, scrape_mtpasa_outages

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

fast_cache   = {"data": None, "last_updated": None, "error": None}
mtpasa_cache = {"data": [],   "last_updated": None, "error": None}
gen_cache  = {"data": None, "last_updated": None, "error": None}
slow_cache = {"data": None, "last_updated": None, "error": None}

FAST_INTERVAL = 300    # 5 min
GEN_INTERVAL  = 900    # 15 min
SLOW_INTERVAL = 1800   # 30 min


_FAST_EMPTY = {
    "prices": {}, "demand": {}, "op_demand": {}, "region_summary": {},
    "historical_prices": {}, "predispatch_prices": {}, "predispatch_demand": {},
    "predispatch_gen": {}, "predispatch_ic": {}, "ic_flows": {}, "ic_history": {},
    "origin_assets": {}, "tomorrow_prices": {}, "tomorrow_demand": {},
    "dispatch_history": {}, "fuel_mix": {}, "predispatch_units": {},
    "dispatch_prices_5min": {}, "demand_history": {}, "op_demand_history": {}, "predispatch_sensitivity": {},
    "interconnectors": {}, "generation": {}, "fuel_colors": {}, "all_fuels": [],
    "bdu_history": {},
}


async def _append_prices_to_github(prices_5min: dict):
    """
    Append latest 5-min dispatch prices to today's daily file in GitHub.
    File: data/prices/YYYY-MM-DD.json
    Structure: { "date": "YYYY-MM-DD", "QLD1": {"HH:MM": rrp, ...}, ... }
    """
    import json, base64, os
    from datetime import datetime, timezone, timedelta
    import httpx

    AEST     = timezone(timedelta(hours=10))
    today    = datetime.now(AEST).strftime("%Y-%m-%d")
    GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
    GH_REPO  = os.environ.get("GITHUB_REPO", "")
    GH_PATH  = f"data/prices/{today}.json"
    GH_HEADERS = {
        "Authorization": f"token {GH_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }

    if not GH_TOKEN or not GH_REPO:
        return

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            # Read existing file
            r = await client.get(
                f"https://api.github.com/repos/{GH_REPO}/contents/{GH_PATH}",
                headers=GH_HEADERS,
            )
            if r.status_code == 200:
                resp_json = r.json()
                file_sha  = resp_json.get("sha", "")
                existing  = json.loads(base64.b64decode(resp_json["content"]).decode())
            else:
                file_sha = ""
                existing = {"date": today}

            # Merge new prices in
            for region, pts in prices_5min.items():
                if region not in existing:
                    existing[region] = {}
                for pt in pts:
                    existing[region][pt["interval"]] = pt["rrp"]

            # Write back
            encoded = base64.b64encode(
                json.dumps(existing, separators=(',', ':')).encode()
            ).decode()
            payload = {
                "message": f"prices: {today}",
                "content": encoded,
            }
            if file_sha:
                payload["sha"] = file_sha
            await client.put(
                f"https://api.github.com/repos/{GH_REPO}/contents/{GH_PATH}",
                headers=GH_HEADERS,
                json=payload,
            )
            logger.debug(f"prices: wrote {today}.json")
    except Exception as e:
        logger.warning(f"prices: GitHub write failed: {e}")


async def _run_fast():
    t0 = time.time()
    try:
        data = await asyncio.wait_for(
            asyncio.get_running_loop().run_in_executor(None, scrape_all),
            timeout=60,
        )
        fast_cache["data"] = data
        fast_cache["last_updated"] = datetime.now(timezone.utc).isoformat()
        fast_cache["error"] = None
        logger.info(f"Fast scrape done in {time.time()-t0:.1f}s")
        # Append prices to GitHub rolling file (fire and forget)
        prices_5min = data.get("dispatch_prices_5min", {})
        if prices_5min:
            asyncio.create_task(_append_prices_to_github(prices_5min))
    except asyncio.TimeoutError:
        logger.error("Fast scrape timed out after 60s")
        fast_cache["error"] = "timeout"
        if fast_cache["data"] is None:
            fast_cache["data"] = dict(_FAST_EMPTY)
            fast_cache["last_updated"] = datetime.now(timezone.utc).isoformat()
    except Exception as e:
        logger.error(f"Fast scrape error: {e}\n{traceback.format_exc()}")
        fast_cache["error"] = str(e)
        if fast_cache["data"] is None:
            fast_cache["data"] = dict(_FAST_EMPTY)
            fast_cache["last_updated"] = datetime.now(timezone.utc).isoformat()


async def _gap_fill_prices():
    """
    Every slow scrape (30 min): check today's and yesterday's GitHub price files
    and patch any missing 5-min intervals from TradingIS CURRENT.
    Runs sequentially with delays to avoid 403s.
    """
    import json, base64, os, zipfile as _zf, io as _io
    from datetime import timedelta
    from scraper import (TRADING_CURRENT, NEM_REGIONS, AEST,
                         _list_hrefs, _parse_aemo, _get as _scraper_get)
    import httpx

    AEST_TZ  = timezone(timedelta(hours=10))
    now_aest = datetime.now(AEST_TZ)
    GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
    GH_REPO  = os.environ.get("GITHUB_REPO", "")
    GH_HEADERS = {
        "Authorization": f"token {GH_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }

    if not GH_TOKEN or not GH_REPO:
        return

    # Build full set of expected 5-min intervals for a day up to a cutoff time
    def _expected_intervals(up_to_hhmm: str) -> set:
        intervals = set()
        h, m = 0, 5
        while True:
            s = f"{h:02d}:{m:02d}"
            if s > up_to_hhmm:
                break
            intervals.add(s)
            m += 5
            if m >= 60:
                m = 0
                h += 1
            if h >= 24:
                break
        return intervals

    loop = asyncio.get_running_loop()

    # Determine which dates to gap-fill: today + yesterday (near midnight)
    today_str     = now_aest.strftime("%Y-%m-%d")
    yesterday_str = (now_aest - timedelta(days=1)).strftime("%Y-%m-%d")
    now_hhmm      = now_aest.strftime("%H:%M")
    is_near_midnight = now_aest.hour == 0 and now_aest.minute < 35

    dates_to_check = [today_str]
    if is_near_midnight:
        dates_to_check.append(yesterday_str)

    # List CURRENT files once
    try:
        current_hrefs = await loop.run_in_executor(None, _list_hrefs, TRADING_CURRENT)
    except Exception as e:
        logger.warning(f"gap-fill: CURRENT listing failed: {e}")
        return

    # Build map: date → { hhmm → url } (last sequence per timestamp)
    current_map: dict = {}
    for href in current_hrefs:
        fname = href.split('/')[-1]
        parts = fname.split('_')
        if len(parts) >= 3 and len(parts[2]) >= 12:
            try:
                fd   = datetime.strptime(parts[2][:8], "%Y%m%d").strftime("%Y-%m-%d")
                hhmm = f"{parts[2][8:10]}:{parts[2][10:12]}"
                current_map.setdefault(fd, {})[hhmm] = href
            except ValueError:
                pass

    for date_str in dates_to_check:
        is_today = date_str == today_str
        up_to    = now_hhmm if is_today else "23:55"

        # Read existing GitHub file
        path = f"data/prices/{date_str}.json"
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                r = await client.get(
                    f"https://api.github.com/repos/{GH_REPO}/contents/{path}",
                    headers=GH_HEADERS
                )
                if r.status_code == 200:
                    rj       = r.json()
                    file_sha = rj.get("sha", "")
                    existing = json.loads(base64.b64decode(rj["content"]).decode())
                else:
                    file_sha = ""
                    existing = {"date": date_str}
        except Exception as e:
            logger.warning(f"gap-fill: read {date_str} failed: {e}")
            continue

        # Find missing intervals for any region
        expected    = _expected_intervals(up_to)
        have        = set()
        for region in NEM_REGIONS:
            have |= set(existing.get(region, {}).keys())
        missing = expected - have

        if not missing:
            logger.debug(f"gap-fill: {date_str} complete ({len(expected)} intervals)")
            continue

        logger.info(f"gap-fill: {date_str} missing {len(missing)} intervals: {sorted(missing)[:5]}...")

        # Fetch missing intervals from CURRENT
        day_current = current_map.get(date_str, {})
        fetched_any = False

        for hhmm in sorted(missing):
            url = day_current.get(hhmm)
            if not url:
                continue
            try:
                await asyncio.sleep(1.5)  # gentle delay between requests
                def _fetch(u=url):
                    r = _scraper_get(u, timeout=15)
                    if not r:
                        return ""
                    try:
                        with _zf.ZipFile(_io.BytesIO(r.content)) as z:
                            csvs = [n for n in z.namelist() if n.lower().endswith(".csv")]
                            if csvs:
                                with z.open(csvs[0]) as f:
                                    return f.read().decode("utf-8", errors="replace")
                    except Exception:
                        return ""
                    return ""

                text = await loop.run_in_executor(None, _fetch)
                if not text:
                    continue

                for row in _parse_aemo(text, "TRADING_PRICE"):
                    region = row.get("REGIONID", "").strip()
                    if region not in NEM_REGIONS: continue
                    if row.get("INVALIDFLAG", "0") not in ("0", ""): continue
                    dt_str  = row.get("SETTLEMENTDATE", "")
                    rrp_str = row.get("RRP", "")
                    if not dt_str or not rrp_str: continue
                    try:
                        dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                        existing.setdefault(region, {})[dt.strftime("%H:%M")] = round(float(rrp_str), 2)
                        fetched_any = True
                    except (ValueError, TypeError):
                        continue
            except Exception as e:
                logger.warning(f"gap-fill: fetch {date_str} {hhmm} failed: {e}")

        # Write back if we got anything
        if fetched_any:
            try:
                encoded = base64.b64encode(
                    json.dumps(existing, separators=(',', ':')).encode()
                ).decode()
                payload = {"message": f"gap-fill: {date_str}", "content": encoded}
                if file_sha: payload["sha"] = file_sha
                async with httpx.AsyncClient(timeout=10) as client:
                    wr = await client.put(
                        f"https://api.github.com/repos/{GH_REPO}/contents/{path}",
                        headers=GH_HEADERS, json=payload
                    )
                logger.info(f"gap-fill: {date_str} patched, status={wr.status_code}")
            except Exception as e:
                logger.warning(f"gap-fill: write {date_str} failed: {e}")
        else:
            logger.info(f"gap-fill: {date_str} nothing fetched from CURRENT")


async def _run_slow():
    t0 = time.time()
    try:
        loop = asyncio.get_running_loop()
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_slow),
            timeout=240  # MTPASA(45s) + BOM(15s) + STPASA + gas(30s) + margin
        )
        slow_cache["data"] = data
        slow_cache["last_updated"] = datetime.now(timezone.utc).isoformat()
        slow_cache["error"] = None
        logger.info(f"Slow scrape done in {time.time()-t0:.1f}s")
        # Gap-fill prices in background (fire and forget)
        asyncio.create_task(_gap_fill_prices())
    except asyncio.TimeoutError:
        logger.error("Slow scrape timed out after 120s")
        slow_cache["error"] = "timeout"
        # Populate with whatever partial data we have rather than staying empty
        if slow_cache["data"] is None:
            slow_cache["data"] = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "generators": {"grouped": {}, "summary": {}, "fuel_colors": {}, "reg_list_count": 0, "scada_count": 0},
                "stpasa_demand": {},
                "fuel_mix_today": {},
                "fuel_colors": {},
                "all_fuels": [],
                "weather": {},
                "gas": {},
            }
            slow_cache["last_updated"] = datetime.now(timezone.utc).isoformat()


async def fast_loop():
    while True:
        try:
            await _run_fast()
        except Exception as e:
            logger.error(f"Fast scrape error: {e}\n{traceback.format_exc()}")
            fast_cache["error"] = str(e)
        await asyncio.sleep(FAST_INTERVAL)


async def _run_gen():
    t0 = time.time()
    try:
        loop = asyncio.get_running_loop()
        data = await asyncio.wait_for(loop.run_in_executor(None, scrape_gen), timeout=60)
        gen_cache["data"] = data
        gen_cache["last_updated"] = datetime.now(timezone.utc).isoformat()
        gen_cache["error"] = None
        logger.info(f"Gen scrape done in {time.time()-t0:.1f}s — "
                    f"scada={data.get('scada_count',0)} reg={data.get('reg_count',0)}")
    except asyncio.TimeoutError:
        logger.error("Gen scrape timed out")
        gen_cache["error"] = "timeout"
        if gen_cache["data"] is None:
            gen_cache["data"] = {"fuel_mix": {}, "nem_totals": {}, "grouped": {},
                                 "fuel_colors": {}, "all_fuels": [], "scada_count": 0, "reg_count": 0}
    except Exception as e:
        logger.error(f"Gen scrape error: {e}")
        gen_cache["error"] = str(e)


async def gen_loop():
    await asyncio.sleep(5)   # let fast scrape finish first
    # Backfill 24hr SCADA history once at startup so chart is immediately populated
    try:
        loop = asyncio.get_running_loop()
        logger.info("Starting SCADA history backfill…")
        await asyncio.wait_for(
            loop.run_in_executor(None, scrape_scada_history), timeout=120
        )
        logger.info("SCADA history backfill complete")
    except Exception as e:
        logger.warning(f"SCADA history backfill failed: {e}")
    while True:
        try:
            await _run_gen()
        except Exception as e:
            logger.error(f"Gen loop error: {e}")
            gen_cache["error"] = str(e)
        await asyncio.sleep(GEN_INTERVAL)


async def mtpasa_loop():
    await asyncio.sleep(10)  # let fast/gen start first
    while True:
        try:
            loop = asyncio.get_running_loop()
            data = await asyncio.wait_for(
                loop.run_in_executor(None, scrape_mtpasa_outages),
                timeout=120
            )
            mtpasa_cache["data"] = data
            mtpasa_cache["last_updated"] = datetime.now(timezone.utc).isoformat()
            mtpasa_cache["error"] = None
            logger.info(f"MTPASA scrape done: {len(data)} units")
        except asyncio.TimeoutError:
            logger.warning("MTPASA scrape timed out")
            mtpasa_cache["error"] = "timeout"
        except Exception as e:
            logger.warning(f"MTPASA scrape error: {e}")
            mtpasa_cache["error"] = str(e)
        await asyncio.sleep(1800)  # refresh every 30 min


async def slow_loop():
    while True:
        try:
            await _run_slow()
        except Exception as e:
            logger.error(f"Slow scrape error: {e}\n{traceback.format_exc()}")
            slow_cache["error"] = str(e)
        await asyncio.sleep(SLOW_INTERVAL)


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Kick off all scrapes in background — don't block startup
    # Server binds immediately; /api/data returns 202 until data arrives
    fast_task   = asyncio.create_task(fast_loop())
    gen_task    = asyncio.create_task(gen_loop())
    slow_task   = asyncio.create_task(slow_loop())
    mtpasa_task = asyncio.create_task(mtpasa_loop())
    yield
    fast_task.cancel()
    gen_task.cancel()
    slow_task.cancel()
    mtpasa_task.cancel()


app = FastAPI(title="NEM Dashboard", lifespan=lifespan)

static_dir = Path(__file__).parent / "static"
static_dir.mkdir(exist_ok=True)
(static_dir / "data").mkdir(exist_ok=True)  # ensure views.json directory exists
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


@app.get("/api/data")
async def get_data():
    if fast_cache["data"] is None:
        return JSONResponse(
            content={"error": fast_cache.get("error", "Data not yet available"), "loading": True},
            status_code=503 if fast_cache["error"] else 202,
        )
    return JSONResponse(content={
        **fast_cache["data"],
        "last_updated": fast_cache["last_updated"],
        "cache_error":  fast_cache.get("error"),
    })


@app.get("/api/gen")
async def get_gen():
    if gen_cache["data"] is None:
        return JSONResponse(
            content={"loading": True, "error": gen_cache.get("error")},
            status_code=202,
        )
    from scraper import _get_fuel_history
    return JSONResponse(content={
        **gen_cache["data"],
        "fuel_history":  _get_fuel_history(),   # always fresh — gen cache is 15min stale
        "last_updated": gen_cache["last_updated"],
        "cache_error":  gen_cache.get("error"),
    })


@app.get("/api/slow")
async def get_slow():
    if slow_cache["data"] is None:
        return JSONResponse(
            content={"loading": True, "error": slow_cache.get("error")},
            status_code=202,
        )
    return JSONResponse(content={
        **slow_cache["data"],
        "mtpasa_outages": mtpasa_cache["data"],
        "last_updated": slow_cache["last_updated"],
        "cache_error":  slow_cache.get("error"),
    })


@app.get("/api/rescrape")
async def rescrape():
    """Force a full rescrape: backfill history + fast + gen + slow caches."""
    import asyncio
    from scraper import scrape_scada_history, scrape_dispatch_history

    async def _run():
        loop = asyncio.get_running_loop()
        # Run backfills in parallel
        await asyncio.gather(
            loop.run_in_executor(None, scrape_scada_history),
            loop.run_in_executor(None, scrape_dispatch_history),
        )
        # Then fast scrape
        await _run_fast()
        # Then gen scrape
        await _run_gen()
        # Then slow scrape (includes weather)
        await _run_slow()

    asyncio.create_task(_run())
    return {"status": "rescrape started — allow ~60s for data to refresh"}


@app.get("/api/health")
async def health():
    return {
        "status":           "ok" if fast_cache["data"] else "loading",
        "fast_updated":     fast_cache["last_updated"],
        "slow_updated":     slow_cache["last_updated"],
        "fast_has_data":    fast_cache["data"] is not None,
        "slow_has_data":    slow_cache["data"] is not None,
        "fast_error":       fast_cache.get("error"),
        "slow_error":       slow_cache.get("error"),
    }


@app.get("/api/debug")
async def debug():
    import csv, io
    from scraper import _list_hrefs, _read_zip, _parse_aemo, DISPATCH_IS_URL, PREDISPATCH_URL, TRADING_CURRENT, ST_PASA_URL
    from zoneinfo import ZoneInfo

    result = {}
    aest = ZoneInfo("Australia/Brisbane")  # UTC+10 fixed, matches AEMO
    now = datetime.now(aest)
    today_str = now.strftime("%Y%m%d")
    result["now_aest"] = now.isoformat()

    # Direct test of scrape_trading_history
    try:
        from scraper import scrape_trading_history
        th = scrape_trading_history()
        result["trading_history_test"] = {
            "price_regions": list(th["prices"].keys()),
            "price_counts": {r: len(v) for r, v in th["prices"].items()},
            "price_sample": {r: v[:2] for r, v in th["prices"].items()},
            "fetch_stats": th.get("fetch_stats", {}),
        }
    except Exception:
        result["trading_history_error"] = traceback.format_exc()

    # Trading archive timing
    try:
        t0 = time.time()
        all_zips = _list_hrefs(TRADING_CURRENT)
        today_zips = [u for u in all_zips if today_str in u]
        result["trading_total_zips"] = len(all_zips)
        result["trading_today_zips"] = len(today_zips)
        result["trading_list_ms"] = round((time.time()-t0)*1000)
        test_url = today_zips[-1] if today_zips else (all_zips[-1] if all_zips else None)
        if test_url:
            result["trading_test_url"] = test_url
            t1 = time.time()
            text = _read_zip(test_url)
            result["trading_fetch_ms"] = round((time.time()-t1)*1000)
            # Show all table names AND their columns
            tables = {}
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if row and row[0].strip().upper() == "I" and len(row) >= 5:
                    key = f"{row[1].strip()}_{row[2].strip()}".upper()
                    tables[key] = [c.strip().upper() for c in row[4:] if c.strip()]
            result["trading_tables_with_cols"] = tables
            # Sample actual rows
            for tbl in ["TRADING_PRICE", "TRADING_REGIONSUM"]:
                rows = _parse_aemo(text, tbl)
                result[f"{tbl}_count"] = len(rows)
                result[f"{tbl}_sample"] = rows[:2] if rows else []
    except Exception:
        result["trading_error"] = traceback.format_exc()

    # Also check DispatchIS (the live file)
    try:
        dispatch_urls = _list_hrefs(DISPATCH_IS_URL)
        if dispatch_urls:
            result["dispatch_latest_url"] = dispatch_urls[-1]
            t1 = time.time()
            text = _read_zip(dispatch_urls[-1])
            result["dispatch_fetch_ms"] = round((time.time()-t1)*1000)
            tables = {}
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if row and row[0].strip().upper() == "I" and len(row) >= 5:
                    key = f"{row[1].strip()}_{row[2].strip()}".upper()
                    tables[key] = [c.strip().upper() for c in row[4:] if c.strip()]
            result["dispatch_tables_with_cols"] = tables
            for tbl in ["DISPATCH_PRICE", "DISPATCH_REGIONSUM"]:
                rows = _parse_aemo(text, tbl)
                result[f"{tbl}_count"] = len(rows)
                result[f"{tbl}_sample"] = rows[:2] if rows else []
    except Exception:
        result["dispatch_error"] = traceback.format_exc()

    # ST PASA listing
    try:
        t0 = time.time()
        pasa_urls = _list_hrefs(ST_PASA_URL)
        result["stpasa_total"] = len(pasa_urls)
        result["stpasa_list_ms"] = round((time.time()-t0)*1000)
        result["stpasa_sample"] = pasa_urls[-2:] if pasa_urls else []
    except Exception:
        result["stpasa_error"] = traceback.format_exc()

    # Cache summary
    if fast_cache["data"]:
        d = fast_cache["data"]
        result["fast_cache"] = {
            "prices":        d.get("prices", {}),
            "demand":        d.get("demand", {}),
            "hist_prices":   {r: len(v) for r, v in d.get("historical_prices", {}).items()},
            "hist_prices_sample": {r: v[:2] for r, v in d.get("historical_prices", {}).items()},
            "hist_fetch_stats": d.get("price_fetch_stats", {}),
            "pd_prices":     {r: len(v) for r, v in d.get("predispatch_prices", {}).items()},
            "dispatch_hist": {r: len(v) for r, v in d.get("dispatch_history", {}).items()},
            "dispatch_hist_sample": {r: v[:2] for r, v in d.get("dispatch_history", {}).items()},
            "demand_hist":   {r: len(v) for r, v in d.get("demand_history", {}).items()},
            "demand_fetch_stats": d.get("demand_fetch_stats", {}),
            "fuel_mix":      {r: len(v) for r, v in d.get("fuel_mix_history", {}).items()},
            "ic_history":    {k: len(v) for k, v in d.get("ic_history", {}).items()},
        }
    if gen_cache["data"]:
        gd = gen_cache["data"]
        result["gen_cache"] = {
            "scada_count":   gd.get("scada_count", 0),
            "reg_count":     gd.get("reg_count", 0),
            "fuel_mix_regions": list(gd.get("fuel_mix", {}).keys()),
            "fuel_mix_sample": gd.get("fuel_mix", {}),
            "fuel_history_pts": {r: len(v) for r, v in gd.get("fuel_history", {}).items()},
            "gen_error":     gen_cache.get("error"),
        }
    else:
        result["gen_cache"] = {"status": "no data yet", "error": gen_cache.get("error")}

    if slow_cache["data"]:
        d = slow_cache["data"]
        result["slow_cache"] = {
            "stpasa_pts":    {r: len(v) for r, v in d.get("stpasa_demand", {}).items()},
        }

    return JSONResponse(content=result)


@app.get("/api/reg-sources-debug")
async def reg_sources_debug():
    """Test all registration list sources to find what works from Render."""
    import asyncio, time as _time
    from scraper import _get, _list_hrefs, NEMWEB_BASE

    loop = asyncio.get_running_loop()
    def _fetch():
        results = {}
        # Test SEMP CSV
        url = f"{NEMWEB_BASE}/REPORTS/CURRENT/SEMP/PUBLIC_SEMP_REGISTRATION.CSV"
        r = _get(url, timeout=10)
        results["semp_csv"] = {"status": r.status_code if r else "failed", "size": len(r.content) if r else 0}

        # Test DVD directory listing
        for path in [f"{NEMWEB_BASE}/REPORTS/CURRENT/DVD/", f"{NEMWEB_BASE}/REPORTS/ARCHIVE/DVD/"]:
            try:
                files = _list_hrefs(path)
                detail = [f for f in files if "DUDETAIL" in f.upper()]
                results[f"dvd_{path.split('/')[-2]}"] = {"file_count": len(files), "dudetail_count": len(detail), "sample": detail[-2:] if detail else []}
            except Exception as e:
                results[f"dvd_{path.split('/')[-2]}"] = {"error": str(e)}
        return results

    return await loop.run_in_executor(None, _fetch)

@app.get("/api/reg-test")
async def reg_test():
    """Directly test AEMO registration list download and parse."""
    import time as _time
    result = {}
    try:
        from scraper import AEMO_REG_LIST_URL, SESSION
        result["url"] = AEMO_REG_LIST_URL
        t0 = _time.time()
        r = SESSION.get(AEMO_REG_LIST_URL, timeout=30, allow_redirects=True)
        result["status_code"] = r.status_code
        result["content_type"] = r.headers.get("Content-Type", "")
        result["content_length"] = len(r.content)
        result["fetch_ms"] = round((_time.time() - t0) * 1000)
        result["final_url"] = r.url

        if r.status_code != 200:
            result["error"] = f"HTTP {r.status_code}"
            return JSONResponse(content=result)

        # Try xlrd
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=r.content)
            result["xlrd_sheets"] = wb.sheet_names()
            # Try each sheet
            sheet_info = {}
            for name in wb.sheet_names():
                sh = wb.sheet_by_name(name)
                # First 5 rows, first 8 cols
                rows_sample = []
                for i in range(min(5, sh.nrows)):
                    rows_sample.append([str(sh.cell_value(i, j)) for j in range(min(8, sh.ncols))])
                sheet_info[name] = {"nrows": sh.nrows, "ncols": sh.ncols, "sample": rows_sample}
            result["sheets"] = sheet_info
        except Exception as e:
            result["xlrd_error"] = str(e)
            # Try openpyxl as fallback diagnostic
            try:
                import openpyxl
                wb2 = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
                result["openpyxl_sheets"] = wb2.sheetnames
            except Exception as e2:
                result["openpyxl_error"] = str(e2)

    except Exception as e:
        result["exception"] = str(e)
        import traceback as tb
        result["traceback"] = tb.format_exc()

    return JSONResponse(content=result)



@app.get("/api/scada-debug")
async def scada_debug():
    """Show raw SCADA DUIDs and which ones match/miss the registry."""
    from scraper import _fetch_full_scada, NEM_UNITS
    loop = asyncio.get_running_loop()
    scada = await loop.run_in_executor(None, _fetch_full_scada)
    matched, unmatched = {}, {}
    for duid, mw in scada.items():
        info = NEM_UNITS.get(duid.upper())
        if info:
            matched[duid] = {**info, "mw": mw}
        else:
            unmatched[duid] = mw
    # Sort unmatched by MW desc to see biggest missing generators
    top_unmatched = dict(sorted(unmatched.items(), key=lambda x: abs(x[1] or 0), reverse=True)[:100])
    return JSONResponse(content={
        "total_scada": len(scada),
        "matched": len(matched),
        "unmatched": len(unmatched),
        "top_unmatched_by_mw": top_unmatched,
        "matched_fuel_summary": {
            fuel: round(sum(v["mw"] or 0 for v in matched.values() if v["fuel"] == fuel), 1)
            for fuel in set(v["fuel"] for v in matched.values())
        }
    })


@app.get("/api/gen-debug")
async def gen_debug():
    """Show what DUIDs are classified as 'Other' in the current gen cache, plus top unmatched."""
    from scraper import _fetch_full_scada, NEM_UNITS, _infer_fuel_from_duid
    data = gen_cache.get("data") or {}
    grouped = data.get("grouped", {})
    other_by_region = {}
    for region, fuels in grouped.items():
        others = fuels.get("Other", [])
        if others:
            other_by_region[region] = sorted(others, key=lambda x: x.get("mw") or 0, reverse=True)[:20]

    # Live check: fetch SCADA and show top unmatched DUIDs
    loop = asyncio.get_running_loop()
    scada = await loop.run_in_executor(None, _fetch_full_scada)
    unmatched = {}
    for duid, mw in scada.items():
        if not NEM_UNITS.get(duid.upper()):
            inferred = _infer_fuel_from_duid(duid)
            unmatched[duid] = {"mw": mw, "inferred_fuel": inferred}
    top_unmatched = dict(sorted(unmatched.items(), key=lambda x: abs(x[1]["mw"] or 0), reverse=True)[:30])

    # Check fuel_history for QLD Black Coal specifically
    from scraper import _fuel_history
    qld_hist = _fuel_history.get("QLD1", {})
    sample_slot = sorted(qld_hist.keys())[-1] if qld_hist else None
    qld_fuels = qld_hist.get(sample_slot, {}) if sample_slot else {}

    # Check which SCADA DUIDs map to QLD Black Coal
    qld_coal_scada = {}
    for duid, mw in scada.items():
        info = NEM_UNITS.get(duid.upper(), {})
        if info.get("region") == "QLD1" and info.get("fuel") == "Black Coal":
            qld_coal_scada[duid] = {"mw": mw, "station": info.get("station")}

    # Show fuel_mix totals from gen_cache vs what SCADA should give
    gen_data = gen_cache.get("data") or {}
    fuel_mix_totals = {}
    for region, fuels in gen_data.get("fuel_mix", {}).items():
        fuel_mix_totals[region] = {f: round(mw,1) for f,mw in fuels.items() if mw > 0}
    fuel_mix_grand_total = sum(sum(f.values()) for f in fuel_mix_totals.values())

    # SCADA total vs fuel_mix total
    scada_total = sum(abs(v) for v in scada.values() if v and v > 0)

    return JSONResponse(content={
        "other_by_region": other_by_region,
        "top_unmatched_scada_duids": top_unmatched,
        "nem_units_count": len(NEM_UNITS),
        "scada_total_mw": round(scada_total, 1),
        "fuel_mix_total_mw": round(fuel_mix_grand_total, 1),
        "fuel_mix_by_region": fuel_mix_totals,
        "qld_fuel_history_slot": sample_slot,
        "qld_fuel_history_fuels": qld_fuels,
        "qld_coal_scada_duids": qld_coal_scada,
    })


@app.get("/api/station/{duid}")
async def station_detail(duid: str):
    from scraper import _duid_history, NEM_UNITS
    duid = duid.strip().upper()
    info = NEM_UNITS.get(duid, {})
    from datetime import datetime
    from zoneinfo import ZoneInfo
    now_label = datetime.now(ZoneInfo("Australia/Brisbane")).strftime("%H:%M")
    history = _duid_history.get(duid, {})
    history_series = [{"interval": k, "mw": v} for k, v in sorted(history.items()) if k <= now_label]
    return JSONResponse(content={
        "duid": duid, "station": info.get("station", duid),
        "fuel": info.get("fuel", "Other"), "region": info.get("region", ""),
        "capacity": info.get("capacity"), "history": history_series, "predispatch": [],
    })


@app.get("/api/stations/batch")
async def station_batch(duids: str):
    """Return history for multiple DUIDs in one request. duids=DUID1,DUID2,..."""
    from scraper import _duid_history, NEM_UNITS
    result = []
    for duid in duids.upper().split(","):
        duid = duid.strip()
        if not duid:
            continue
        info = NEM_UNITS.get(duid, {})
        history = _duid_history.get(duid, {})
        from datetime import datetime as _dt2
        from zoneinfo import ZoneInfo as _ZI
        _now = _dt2.now(_ZI("Australia/Brisbane")).strftime("%H:%M")
        result.append({
            "duid": duid, "station": info.get("station", duid),
            "fuel": info.get("fuel", "Other"), "region": info.get("region", ""),
            "capacity": info.get("capacity"),
            "history": [{"interval": k, "mw": v} for k, v in sorted(history.items()) if k <= _now],
            "predispatch": [],
        })
    return JSONResponse(content=result)


@app.get("/api/historical_prices")
async def historical_prices(date: str):
    """Fetch 30-min trading prices for a given date (YYYYMMDD)."""
    import re
    if not re.match(r'^\d{8}$', date):
        return JSONResponse(status_code=400, content={"error": "date must be YYYYMMDD"})
    from scraper import scrape_historical_prices
    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_historical_prices, date),
            timeout=60
        )
        return data
    except asyncio.TimeoutError:
        return JSONResponse(status_code=504, content={"error": "timeout"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})



@app.get("/api/pasa-structure-debug")
async def pasa_structure_debug():
    """Inspect PDPASA and STPASA DUID availability file structures."""
    import asyncio
    from scraper import _read_zip, _parse_aemo, NEMWEB_BASE

    loop = asyncio.get_running_loop()

    def _fetch():
        import csv, io, requests
        results = {}
        targets = {
            "PDPASA": f"{NEMWEB_BASE}/REPORTS/CURRENT/PDPASA_DUIDAvailability/PUBLIC_PDPASA_DUIDAVAILABILITY_202603232100_0000000509352436.zip",
            "STPASA": f"{NEMWEB_BASE}/REPORTS/CURRENT/STPASA_DUIDAvailability/PUBLIC_STPASA_DUIDAVAILABILITY_202603232000_0000000509348868.zip",
        }
        for name, url in targets.items():
            text = _read_zip(url)
            if not text:
                results[name] = {"error": "failed to read"}
                continue
            # Get all table names and columns
            tables = {}
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if row and row[0].strip().upper() == 'I' and len(row) >= 3:
                    tbl = row[2].strip().upper()
                    cols = [c.strip() for c in row[4:] if c.strip()]
                    tables[tbl] = cols
            # Sample rows from DUID tables
            samples = {}
            for tbl in tables:
                if 'DUID' in tbl:
                    rows = list(_parse_aemo(text, tbl))
                    samples[tbl] = rows[:3]
            results[name] = {"tables": tables, "samples": samples}
        return results

    return await loop.run_in_executor(None, _fetch)

@app.get("/api/pasa-probe-debug")
async def pasa_probe_debug():
    """Probe known PDPASA and STPASA DUID availability paths on NEMWeb."""
    import asyncio
    from scraper import _get, _list_hrefs, NEMWEB_BASE

    loop = asyncio.get_running_loop()

    def _fetch():
        # Try known paths for DUID-level availability data
        candidates = [
            f"{NEMWEB_BASE}/REPORTS/CURRENT/PDPASA_DUIDAvailability/",
            f"{NEMWEB_BASE}/REPORTS/CURRENT/PDPASA_DUID/",
            f"{NEMWEB_BASE}/REPORTS/CURRENT/Short_Term_PASA_DUID/",
            f"{NEMWEB_BASE}/REPORTS/CURRENT/STPASA_DUIDAvailability/",
            f"{NEMWEB_BASE}/REPORTS/CURRENT/PD_PASA/",
            f"{NEMWEB_BASE}/REPORTS/CURRENT/Predispatch_PASA/",
            f"{NEMWEB_BASE}/REPORTS/CURRENT/MTPASA_DUIDAvailability/",  # known good - control
        ]
        results = {}
        for url in candidates:
            r = _get(url, timeout=10)
            if r:
                import re
                zips = re.findall(r'href="([^"]+\.zip)"', r.text, re.IGNORECASE)
                results[url] = {"status": r.status_code, "zip_count": len(zips), "sample": zips[-1] if zips else None}
            else:
                results[url] = {"status": "failed"}
        return results

    result = await loop.run_in_executor(None, _fetch)
    return result

@app.get("/api/pasa-dirs-debug")
async def pasa_dirs_debug():
    """List all PASA-related directories on NEMWeb to find DUID availability."""
    import asyncio
    from scraper import _get, NEMWEB_BASE
    import re

    loop = asyncio.get_running_loop()

    def _fetch():
        # List the CURRENT directory to find all PASA-related folders
        r = _get(f"{NEMWEB_BASE}/REPORTS/CURRENT/")
        if not r:
            return {"error": "failed to list CURRENT"}
        folders = re.findall(r'href="([^"]+)"', r.text)
        pasa = [f for f in folders if 'PASA' in f.upper() or 'AVAIL' in f.upper()]
        return {"pasa_folders": pasa, "all_count": len(folders)}

    result = await loop.run_in_executor(None, _fetch)
    return result

@app.get("/api/yallourn-debug")
async def yallourn_debug():
    """Check what STPASA, PDPASA and MTPASA show for Yallourn units."""
    import asyncio
    from scraper import scrape_pasa_duid_availability, _list_hrefs, _read_zip, _parse_aemo, NEM_UNITS, MTPASA_DUID_URL, AEST
    from datetime import datetime

    loop = asyncio.get_running_loop()

    def _fetch():
        now_aest = datetime.now(AEST)
        today_str = now_aest.strftime("%Y/%m/%d")
        now_label = now_aest.strftime("%Y-%m-%d %H:%M")

        duids = ["YWPS1","YWPS2","YWPS3","YWPS4"]

        # STPASA
        st = scrape_pasa_duid_availability("STPASA")
        st_slots = st.get("slots", {})
        st_max = st.get("max_avail", {})

        # PDPASA
        pd = scrape_pasa_duid_availability("PDPASA")
        pd_slots = pd.get("slots", {})

        # MTPASA
        files = _list_hrefs(MTPASA_DUID_URL)
        text = _read_zip(sorted(files)[-1])
        mtpasa = {}
        for row in _parse_aemo(text, "MTPASA_DUIDAVAILABILITY"):
            duid = row.get("DUID","").strip()
            if duid in duids:
                day = row.get("DAY","").strip()[:10]
                avail = float(row.get("PASAAVAILABILITY") or 0)
                state = row.get("UNITSTATE","").strip()
                if duid not in mtpasa: mtpasa[duid] = []
                mtpasa[duid].append({"day": day, "avail": avail, "state": state})

        result = {}
        for duid in duids:
            cap = NEM_UNITS.get(duid, {}).get("capacity", 0)
            # STPASA first interval
            duid_st = st_slots.get(duid, {})
            st_first_dt = sorted(duid_st.keys())[0] if duid_st else None
            st_first_val = duid_st[st_first_dt] if st_first_dt else None
            # PDPASA first interval
            duid_pd = pd_slots.get(duid, {})
            pd_first_dt = sorted(duid_pd.keys())[0] if duid_pd else None
            pd_first_val = duid_pd[pd_first_dt] if pd_first_dt else None
            # MTPASA
            mt = sorted(mtpasa.get(duid, []), key=lambda x: x["day"])
            past = [e for e in mt if e["day"] <= today_str]
            cur_mt = past[-1] if past else (mt[0] if mt else None)

            result[duid] = {
                "capacity": cap,
                "threshold_70": round(cap * 0.70),
                "stpasa_first_dt": st_first_dt,
                "stpasa_first_avail": st_first_val,
                "pdpasa_first_dt": pd_first_dt,
                "pdpasa_first_avail": pd_first_val,
                "mtpasa_current": cur_mt,
                "mtpasa_next_3": [e for e in mt if e["day"] > today_str][:3],
            }
        return result

    return await loop.run_in_executor(None, _fetch)

@app.get("/api/tnps1-debug2")
async def tnps1_debug2():
    """Check return date logic for TNPS1."""
    import asyncio
    from scraper import scrape_pasa_duid_availability, _list_hrefs, _read_zip, _parse_aemo, NEM_UNITS, MTPASA_DUID_URL, AEST
    from datetime import datetime
    loop = asyncio.get_running_loop()
    def _fetch():
        now = datetime.now(AEST)
        today_str = now.strftime("%Y/%m/%d")
        cap = NEM_UNITS.get("TNPS1", {}).get("capacity", 480)
        threshold = cap * 0.70

        pd = scrape_pasa_duid_availability("PDPASA")
        st = scrape_pasa_duid_availability("STPASA")
        pd_slots = pd.get("slots", {}).get("TNPS1", {})
        st_slots = st.get("slots", {}).get("TNPS1", {})

        # Build STPASA daily 00:30
        daily_st = {s[:10]: st_slots[s] for s in st_slots if s[11:] == "00:30"}

        # MTPASA
        files = _list_hrefs(MTPASA_DUID_URL)
        text = _read_zip(sorted(files)[-1])
        mt = {}
        for row in _parse_aemo(text, "MTPASA_DUIDAVAILABILITY"):
            if row.get("DUID","").strip() == "TNPS1":
                d = row.get("DAY","").strip()[:10]
                mt[d] = float(row.get("PASAAVAILABILITY") or 0)

        # Find outage start
        pd_first_below = next((s[:10].replace("-","/") for s in sorted(pd_slots) if pd_slots[s] < threshold), None)
        st_first_below = next((d.replace("-","/") for d in sorted(daily_st) if daily_st[d] < threshold), None)
        mt_first_below = next((d for d in sorted(mt) if mt[d] < threshold), None)
        outage_start = min([d for d in [pd_first_below, st_first_below, mt_first_below] if d]) if any([pd_first_below, st_first_below, mt_first_below]) else None

        # Find return - scan STPASA then MTPASA for first slot >= threshold after outage_start
        outage_label = outage_start.replace("/","-") if outage_start else ""
        stpasa_return = next((d.replace("-","/") for d in sorted(daily_st) if d > outage_label and daily_st[d] >= threshold), None)
        mtpasa_return = next((d for d in sorted(mt) if d.replace("/","-") > outage_label and mt[d] >= threshold), None)

        return {
            "capacity": cap, "threshold_70": threshold,
            "pdpasa_first_5": sorted(pd_slots.items())[:5],
            "stpasa_daily_first_5": sorted(daily_st.items())[:5],
            "mtpasa_first_5": sorted(mt.items())[:5],
            "pd_first_below_threshold": pd_first_below,
            "st_first_below_threshold": st_first_below,
            "mt_first_below_threshold": mt_first_below,
            "outage_start": outage_start,
            "stpasa_return": stpasa_return,
            "mtpasa_return": mtpasa_return,
        }
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/tnps1-debug")
async def tnps1_debug():
    """Check exactly what PDPASA and STPASA show for TNPS1."""
    import asyncio
    from scraper import scrape_pasa_duid_availability, NEM_UNITS, AEST
    from datetime import datetime
    loop = asyncio.get_running_loop()
    def _fetch():
        now = datetime.now(AEST).strftime("%Y-%m-%d %H:%M")
        pd = scrape_pasa_duid_availability("PDPASA")
        st = scrape_pasa_duid_availability("STPASA")
        pd_slots = pd.get("slots", {}).get("TNPS1", {})
        st_slots = st.get("slots", {}).get("TNPS1", {})
        cap = NEM_UNITS.get("TNPS1", {}).get("capacity", 0)
        pd_sorted = sorted(pd_slots.items())
        st_sorted = sorted(st_slots.items())
        return {
            "now": now,
            "capacity": cap,
            "threshold_70": round(cap * 0.70),
            "pdpasa_first_5": pd_sorted[:5],
            "pdpasa_in_file": bool(pd_slots),
            "stpasa_first_5": st_sorted[:5],
            "stpasa_in_file": bool(st_slots),
        }
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/trading-window-debug")
async def trading_window_debug():
    """Check how far back TradingIS CURRENT goes."""
    import asyncio
    from scraper import _list_hrefs, TRADING_CURRENT
    loop = asyncio.get_running_loop()
    def _fetch():
        files = _list_hrefs(TRADING_CURRENT)
        dates = sorted(set(
            f.split("PUBLIC_TRADINGIS_")[1][:8]
            for f in files if "PUBLIC_TRADINGIS_" in f
        ))
        return {
            "total_files": len(files),
            "earliest_date": dates[0] if dates else None,
            "latest_date": dates[-1] if dates else None,
            "all_dates": dates,
        }
    return await loop.run_in_executor(None, _fetch)


@app.get("/api/trading-archive-range")
async def trading_archive_range():
    """Check full date range of TradingIS archive weekly files."""
    import asyncio
    from scraper import _list_hrefs, TRADING_ARCHIVE
    loop = asyncio.get_running_loop()
    def _fetch():
        files = sorted(_list_hrefs(TRADING_ARCHIVE))
        # Extract date ranges from filenames
        import re
        ranges = []
        for f in files:
            m = re.search(r'(\d{8})_(\d{8})', f)
            if m:
                ranges.append({"start": m.group(1), "end": m.group(2), "url": f})
        return {
            "total_files": len(files),
            "oldest": ranges[0] if ranges else None,
            "newest": ranges[-1] if ranges else None,
            "all_ranges": [(r["start"], r["end"]) for r in ranges],
        }
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/trading-archive-debug")
async def trading_archive_debug():
    """Check how far back TradingIS archive goes."""
    import asyncio
    from scraper import _list_hrefs, TRADING_ARCHIVE, NEMWEB_BASE
    loop = asyncio.get_running_loop()
    def _fetch():
        results = {}
        # Check archive root for available months
        try:
            dirs = _list_hrefs(TRADING_ARCHIVE)
            results["archive_months"] = sorted(dirs)[-6:] if dirs else []
            results["total_months"] = len(dirs)
        except Exception as e:
            results["archive_error"] = str(e)

        # Try fetching a specific old month to see if it works
        for ym in ["202603", "202601", "202501", "202401", "202301"]:
            url = f"{TRADING_ARCHIVE}{ym}/"
            try:
                files = _list_hrefs(url)
                results[f"month_{ym}"] = {"count": len(files), "sample": sorted(files)[:2]}
            except Exception as e:
                results[f"month_{ym}"] = {"error": str(e)}
        return results
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/mmsdm-debug")
async def mmsdm_debug():
    """Check MMSDM archive for historical dispatch prices."""
    import asyncio
    from scraper import _list_hrefs, NEMWEB_BASE
    loop = asyncio.get_running_loop()
    def _fetch():
        results = {}
        # MMSDM is published monthly with ~1 month lag
        # Try several possible URL patterns
        urls_to_try = {
            "mmsdm_2026_02": f"{NEMWEB_BASE}/Data/MMSDM/2026/MMSDM_2026_02/MMSDM_Historical_Data_SQLLoader/DATA/PUBLIC_DVD_DISPATCHPRICE_202602010000.zip",
            "mmsdm_root_2026": f"{NEMWEB_BASE}/Data/MMSDM/2026/",
            "mmsdm_root": f"{NEMWEB_BASE}/Data/MMSDM/",
            "nemweb_data": f"{NEMWEB_BASE}/Data/",
            # Alternative: Electricity Statement of Opportunities or other
            "archive_5min": f"{NEMWEB_BASE}/REPORTS/ARCHIVE/5MIN/",
            "current_5min": f"{NEMWEB_BASE}/REPORTS/CURRENT/5MIN/",
        }
        for name, url in urls_to_try.items():
            try:
                import requests
                r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
                results[name] = {"status": r.status_code, "url": url, "content_length": len(r.content), "preview": r.text[:200]}
            except Exception as e:
                results[name] = {"error": str(e), "url": url}
        return results
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/tradingis-structure-debug")
async def tradingis_structure_debug():
    """Check tables and columns inside a TradingIS file."""
    import asyncio, csv, io
    from scraper import _list_hrefs, _read_zip, NEMWEB_BASE
    loop = asyncio.get_running_loop()
    def _fetch():
        url = f"{NEMWEB_BASE}/REPORTS/CURRENT/TradingIS_Reports/"
        files = sorted(_list_hrefs(url))
        # Pick a file from a few days ago
        old_files = [f for f in files if "20260320" in f]
        sample = old_files[0] if old_files else files[10]
        text = _read_zip(sample)
        tables = {}
        current_table = None
        current_headers = []
        sample_rows = {}
        for row in csv.reader(io.StringIO(text)):
            if not row: continue
            if row[0].strip().upper() == 'I' and len(row) >= 3:
                current_table = row[2].strip().upper()
                current_headers = [c.strip() for c in row[4:] if c.strip()]
                tables[current_table] = current_headers
            if row[0].strip().upper() == 'D' and current_table and current_table not in sample_rows:
                d = dict(zip(current_headers, row[4:]))
                sample_rows[current_table] = d
        return {"file": sample, "tables": tables, "sample_rows": sample_rows}
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/trading-files-debug")
async def trading_files_debug():
    """Check what's in TradingIS CURRENT and what date range it covers."""
    import asyncio
    from scraper import _list_hrefs, NEMWEB_BASE
    loop = asyncio.get_running_loop()
    def _fetch():
        url = f"{NEMWEB_BASE}/REPORTS/CURRENT/TradingIS_Reports/"
        files = _list_hrefs(url)
        # Get date range from filenames
        dated = sorted([f for f in files if "PUBLIC_TRADING" in f.upper() or "PUBLIC_DISPATCHPRICE" in f.upper()])
        # Sample file names to see structure
        sample = dated[:3] + dated[-3:]
        # Extract unique dates
        import re
        dates = sorted(set(re.search(r'_(\d{8})', f).group(1)
                          for f in dated if re.search(r'_(\d{8})', f)))
        return {
            "total": len(files),
            "sample_files": sample,
            "date_range": {"first": dates[0] if dates else None, "last": dates[-1] if dates else None},
            "unique_dates_count": len(dates),
            "all_prefixes": sorted(set(f.split('/')[-1][:30] for f in files[:20])),
        }
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/dispatch-sources-debug")
async def dispatch_sources_debug():
    """Check alternative sources for historical 5-min dispatch prices."""
    import asyncio
    from scraper import _list_hrefs, NEMWEB_BASE
    loop = asyncio.get_running_loop()
    def _fetch():
        results = {}
        # AEMO keeps historical dispatch prices in the MMS DVD format
        # Try MMSDM archive which has PUBLIC_DVD_DISPATCHPRICE files
        sources = {
            "mmsdm_202603": f"{NEMWEB_BASE}/Data/MMSDM/2026/MMSDM_2026_03/MMSDM_Historical_Data_SQLLoader/DATA/",
            "mmsdm_202602": f"{NEMWEB_BASE}/Data/MMSDM/2026/MMSDM_2026_02/MMSDM_Historical_Data_SQLLoader/DATA/",
            "archive_trading": f"{NEMWEB_BASE}/REPORTS/ARCHIVE/TradingIS_Reports/202603/",
            "current_trading": f"{NEMWEB_BASE}/REPORTS/CURRENT/TradingIS_Reports/",
        }
        for name, url in sources.items():
            try:
                files = _list_hrefs(url)
                dispatch = [f for f in files if "DISPATCHPRICE" in f.upper() or "DISPATCH_PRICE" in f.upper()]
                results[name] = {
                    "url": url,
                    "total_files": len(files),
                    "dispatch_price_files": len(dispatch),
                    "sample": dispatch[:2],
                }
            except Exception as e:
                results[name] = {"url": url, "error": str(e)}
        return results
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/dispatch-archive-debug")
async def dispatch_archive_debug():
    """Check DispatchIS archive URL structure and how far back it goes."""
    import asyncio
    from scraper import _list_hrefs, NEMWEB_BASE, DISPATCH_IS_URL
    loop = asyncio.get_running_loop()
    def _fetch():
        results = {}
        # 1. Check CURRENT window
        try:
            current = _list_hrefs(DISPATCH_IS_URL)
            current_dates = sorted(set(f.split("PUBLIC_DISPATCHIS_")[1][:8]
                for f in current if "PUBLIC_DISPATCHIS_" in f))
            results["current_dates"] = current_dates
        except Exception as e:
            results["current_error"] = str(e)

        # 2. Check archive - try recent months
        for ym in ["202603", "202602", "202601", "202512"]:
            url = f"{NEMWEB_BASE}/REPORTS/ARCHIVE/DispatchIS_Reports/{ym}/"
            try:
                files = _list_hrefs(url)
                dates = sorted(set(f.split("PUBLIC_DISPATCHIS_")[1][:8]
                    for f in files if "PUBLIC_DISPATCHIS_" in f))
                results[f"archive_{ym}"] = {"count": len(files), "dates_sample": dates[:3] + dates[-3:]}
            except Exception as e:
                results[f"archive_{ym}"] = {"error": str(e)}
        return results
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/dispatch-price-debug3")
async def dispatch_price_debug3():
    """Check PRICE table columns and sample rows."""
    import asyncio, csv, io
    from scraper import _list_hrefs, _read_zip, DISPATCH_IS_URL
    loop = asyncio.get_running_loop()
    def _fetch():
        files = _list_hrefs(DISPATCH_IS_URL)
        matching = sorted([f for f in files if "20260324" in f and "PUBLIC_DISPATCHIS" in f.upper()])
        if not matching:
            return {"error": "no files"}
        text = _read_zip(matching[0])
        headers, rows = [], []
        for row in csv.reader(io.StringIO(text)):
            if not row: continue
            if row[0].strip().upper() == 'I' and len(row) >= 3 and row[2].strip().upper() == 'PRICE':
                headers = [c.strip() for c in row[4:] if c.strip()]
            if row[0].strip().upper() == 'D' and len(row) >= 3 and row[2].strip().upper() == 'PRICE':
                d = dict(zip(headers, row[4:]))
                rows.append(d)
                if len(rows) >= 3: break
        return {"headers": headers, "sample_rows": rows}
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/dispatch-price-debug2")
async def dispatch_price_debug2():
    """Check what tables are inside a DispatchIS file."""
    import asyncio
    from scraper import _list_hrefs, _read_zip, DISPATCH_IS_URL
    loop = asyncio.get_running_loop()
    def _fetch():
        files = _list_hrefs(DISPATCH_IS_URL)
        # Get a file from yesterday
        matching = sorted([f for f in files if "20260324" in f and "PUBLIC_DISPATCHIS" in f.upper()])
        if not matching:
            return {"error": "no files found"}
        url = matching[0]
        text = _read_zip(url)
        # Find all table names (I rows)
        import csv, io
        tables = set()
        for row in csv.reader(io.StringIO(text)):
            if row and row[0].strip().upper() == 'I' and len(row) >= 3:
                tables.add(row[2].strip().upper())
        return {"file": url, "tables": sorted(tables)}
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/dispatch-price-debug")
async def dispatch_price_debug(date: str = "20260324"):
    """Debug historical dispatch prices fetch."""
    import asyncio
    from scraper import _list_hrefs, NEM_REGIONS, NEMWEB_BASE, DISPATCH_IS_URL, AEST
    from datetime import datetime, timedelta
    loop = asyncio.get_running_loop()
    def _fetch():
        now = datetime.now(AEST)
        today = now.date()
        from datetime import datetime as _dt
        req_date = _dt.strptime(date, "%Y%m%d").date()
        ym = req_date.strftime("%Y%m")

        # Try both URLs
        current_url = DISPATCH_IS_URL
        archive_url = f"{NEMWEB_BASE}/REPORTS/ARCHIVE/DispatchIS_Reports/{ym}/"

        current_files, archive_files = [], []
        try:
            all_current = _list_hrefs(current_url)
            current_files = sorted([f for f in all_current if date in f and "PUBLIC_DISPATCHIS" in f.upper()])[:3]
        except Exception as e:
            current_files = [f"ERROR: {e}"]
        try:
            all_archive = _list_hrefs(archive_url)
            archive_files = sorted([f for f in all_archive if date in f and "PUBLIC_DISPATCHIS" in f.upper()])[:3]
        except Exception as e:
            archive_files = [f"ERROR: {e}"]

        return {
            "date": date,
            "today": str(today),
            "req_date": str(req_date),
            "current_url": current_url,
            "archive_url": archive_url,
            "current_matching_files": current_files,
            "archive_matching_files": archive_files,
        }
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/outage-debug")
async def outage_debug():
    """Show first 20 outages with is_current flag to debug filter."""
    import asyncio
    from scraper import scrape_mtpasa_outages
    loop = asyncio.get_running_loop()
    def _fetch():
        results = scrape_mtpasa_outages()
        # Show all outages, flag is_current
        return [{
            "duid": r["duid"],
            "station": r["station"],
            "fuel": r["fuel"],
            "avail_today": r["avail_today"],
            "capacity": r["capacity"],
            "avail_source": r["avail_source"],
            "is_current": r.get("is_current"),
            "outage_start": r.get("outage_start"),
        } for r in results[:30]]
    return await loop.run_in_executor(None, _fetch)

@app.get("/api/stpasa-snapshot")
async def stpasa_snapshot():
    """Check the first STPASA interval - what is available/unavailable right now."""
    import asyncio
    from scraper import _list_hrefs, _read_zip, _parse_aemo, NEM_UNITS, STPASA_DUID_URL, AEST
    from datetime import datetime

    loop = asyncio.get_running_loop()

    def _fetch():
        files = _list_hrefs(STPASA_DUID_URL)
        if not files:
            return {"error": "no files"}
        latest = sorted(files)[-1]
        text = _read_zip(latest)

        # Get all rows, find the first (earliest) INTERVAL_DATETIME
        rows = list(_parse_aemo(text, "DUIDAVAILABILITY"))
        if not rows:
            return {"error": "no rows"}

        # Find the first interval datetime
        first_dt = min(r.get("INTERVAL_DATETIME","") for r in rows if r.get("INTERVAL_DATETIME"))
        last_dt  = max(r.get("INTERVAL_DATETIME","") for r in rows if r.get("INTERVAL_DATETIME"))

        # Get all DUIDs at the first interval
        first_rows = [r for r in rows if r.get("INTERVAL_DATETIME","") == first_dt]

        # Find unavailable ones (< 70% of nameplate)
        unavailable = []
        for r in first_rows:
            duid = r.get("DUID","").strip()
            unit = NEM_UNITS.get(duid, {})
            cap = unit.get("capacity", 0)
            if cap <= 0:
                continue
            avail = float(r.get("GENERATION_PASA_AVAILABILITY") or 0)
            if avail < cap * 0.70:
                unavailable.append({
                    "duid": duid,
                    "station": unit.get("station", duid),
                    "fuel": unit.get("fuel","?"),
                    "region": unit.get("region","?"),
                    "capacity": cap,
                    "avail": avail,
                    "pct": round(avail/cap*100, 1),
                })

        unavailable.sort(key=lambda x: x["avail"] - x["capacity"])
        return {
            "file": latest,
            "first_interval": first_dt,
            "last_interval": last_dt,
            "total_duids_in_file": len(set(r.get("DUID") for r in rows)),
            "unavailable_count": len(unavailable),
            "unavailable": unavailable[:50],
        }

    return await loop.run_in_executor(None, _fetch)

@app.get("/api/eraring-debug2")
async def eraring_debug2():
    """Check earliest/latest STPASA slots for Eraring."""
    import asyncio
    from scraper import scrape_pasa_duid_availability, NEM_UNITS, AEST
    from datetime import datetime

    loop = asyncio.get_running_loop()

    def _fetch():
        now_aest = datetime.now(AEST)
        now_label = now_aest.strftime("%Y-%m-%d %H:%M")

        stpasa_result = scrape_pasa_duid_availability("STPASA")
        stpasa_slots = stpasa_result.get("slots", {})

        result = {}
        for duid in ["ER01","ER02","ER03","ER04"]:
            duid_slots = stpasa_slots.get(duid, {})
            all_slots = sorted(duid_slots.keys())
            past = [k for k in all_slots if k <= now_label]
            future = [k for k in all_slots if k > now_label]
            result[duid] = {
                "now": now_label,
                "total_slots": len(all_slots),
                "earliest": all_slots[0] if all_slots else None,
                "latest": all_slots[-1] if all_slots else None,
                "past_count": len(past),
                "future_count": len(future),
                "first_future": future[0] if future else None,
                "first_future_avail": duid_slots[future[0]] if future else None,
                "sample_future": {k: duid_slots[k] for k in future[:5]},
            }
        return result

    return await loop.run_in_executor(None, _fetch)

@app.get("/api/eraring-debug")
async def eraring_debug():
    """Check what MTPASA and STPASA show for Eraring units right now."""
    import asyncio
    from scraper import scrape_pasa_duid_availability, _list_hrefs, _read_zip, _parse_aemo, NEM_UNITS, MTPASA_DUID_URL, AEST
    from datetime import datetime

    loop = asyncio.get_running_loop()

    def _fetch():
        now_aest = datetime.now(AEST)
        today_str = now_aest.strftime("%Y/%m/%d")
        now_label = now_aest.strftime("%Y-%m-%d %H:%M")

        # Fetch MTPASA
        files = _list_hrefs(MTPASA_DUID_URL)
        latest = sorted(files)[-1]
        text = _read_zip(latest)

        eraring_duids = ["ER01","ER02","ER03","ER04"]
        mtpasa = {}
        for row in _parse_aemo(text, "MTPASA_DUIDAVAILABILITY"):
            duid = row.get("DUID","").strip()
            if duid in eraring_duids:
                day = row.get("DAY","").strip()[:10]
                avail = float(row.get("PASAAVAILABILITY") or 0)
                state = row.get("UNITSTATE","").strip()
                if duid not in mtpasa:
                    mtpasa[duid] = []
                mtpasa[duid].append({"day": day, "avail": avail, "state": state})

        # Fetch STPASA
        stpasa_result = scrape_pasa_duid_availability("STPASA")
        stpasa_slots = stpasa_result.get("slots", {})

        result = {}
        for duid in eraring_duids:
            cap = NEM_UNITS.get(duid, {}).get("capacity", 750)
            mt = sorted(mtpasa.get(duid, []), key=lambda x: x["day"])
            # Current MTPASA: last entry on or before today
            past = [e for e in mt if e["day"] <= today_str]
            cur_mt = past[-1] if past else (mt[0] if mt else None)
            # STPASA current
            duid_st = stpasa_slots.get(duid, {})
            past_st = sorted([k for k in duid_st if k <= now_label])
            cur_st = duid_st[past_st[-1]] if past_st else None
            result[duid] = {
                "capacity": cap,
                "threshold_70pct": round(cap * 0.70),
                "mtpasa_current": cur_mt,
                "mtpasa_next_5": [e for e in mt if e["day"] > today_str][:5],
                "stpasa_current": cur_st,
                "stpasa_in_file": duid in stpasa_slots,
            }
        return {"today": today_str, "duids": result}

    return await loop.run_in_executor(None, _fetch)

@app.get("/api/stpasa-debug")
async def stpasa_debug():
    """Inspect STPASA file tables and columns."""
    import asyncio
    from scraper import _list_hrefs, _read_zip, _parse_aemo, ST_PASA_URL

    loop = asyncio.get_running_loop()

    def _fetch():
        urls = _list_hrefs(ST_PASA_URL)
        pasa_urls = sorted([u for u in urls if "STPASA" in u.upper()])
        if not pasa_urls:
            return {"error": "no STPASA files"}
        url = pasa_urls[-1]
        text = _read_zip(url)

        import csv, io
        tables = {}
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if row and row[0].strip().upper() == 'I' and len(row) >= 3:
                tbl = row[2].strip().upper()
                cols = [c.strip() for c in row[4:] if c.strip()]
                tables[tbl] = cols

        # Sample first 2 rows from DUIDAVAILABILITY if it exists
        samples = {}
        for tbl in tables:
            if 'DUID' in tbl or 'AVAIL' in tbl:
                rows = list(_parse_aemo(text, tbl))
                samples[tbl] = rows[:2]

        return {"file": url, "tables": tables, "samples": samples}

    result = await loop.run_in_executor(None, _fetch)
    return result

@app.get("/api/mtpasa-debug")
async def mtpasa_debug():
    from scraper import _list_hrefs, MTPASA_DUID_URL
    files = _list_hrefs(MTPASA_DUID_URL)
    latest = sorted(files)[-1] if files else None
    return {"total_files": len(files), "latest_file": latest, "all_files": files[-5:]}

@app.get("/api/rescrape")
async def rescrape():
    """Force immediate re-run of all scrape jobs including history backfill."""
    import asyncio
    from scraper import scrape_all, scrape_gen, scrape_slow, scrape_scada_history, scrape_mtpasa_outages
    loop = asyncio.get_running_loop()

    async def _run():
        # Kick off all jobs concurrently
        await asyncio.gather(
            loop.run_in_executor(None, scrape_scada_history),
            return_exceptions=True
        )
        # Then fast + gen
        data = await loop.run_in_executor(None, scrape_all)
        fast_cache["data"] = data
        fast_cache["last_updated"] = datetime.now(timezone.utc).isoformat()
        gen_data = await loop.run_in_executor(None, scrape_gen)
        gen_cache["data"] = gen_data
        gen_cache["last_updated"] = datetime.now(timezone.utc).isoformat()

    asyncio.create_task(_run())
    return {"status": "rescrape triggered — history backfill + fast + gen running in background"}

@app.get("/api/pd-sens-debug")
async def pd_sens_debug():
    from scraper import _fetch_predispatch, _parse_aemo
    text = _fetch_predispatch()
    tables = set()
    import csv, io
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if row and row[0].strip().upper() == 'I' and len(row) >= 3:
            tables.add(row[2].strip().upper())
    # Check PRICESENSITIVITIES columns
    sens_rows = list(_parse_aemo(text, 'PREDISPATCH_PRICESENSITIVITIES'))[:2]
    return {"tables": sorted(tables), "sens_sample": sens_rows}

@app.get("/api/sens-debug2")
async def sens_debug2():
    """Read scenario definitions from the sensitivities file."""
    from scraper import _list_hrefs, _read_zip, _parse_aemo, PREDISPATCH_SENS_URL
    import csv, io
    files = _list_hrefs(PREDISPATCH_SENS_URL)
    if not files: return {"error": "no files"}
    text = _read_zip(files[-1])
    
    # Dump ALL I-rows (table headers) to see every table and its columns
    result = {"file": files[-1], "all_tables": {}}
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if row and row[0].strip().upper() == 'I' and len(row) >= 3:
            tbl = row[2].strip().upper()
            cols = [c.strip() for c in row[4:] if c.strip()]
            result["all_tables"][tbl] = cols
    
    # Sample first 3 D-rows from every table
    result["samples"] = {}
    for tbl in result["all_tables"]:
        rows = list(_parse_aemo(text, tbl))
        result["samples"][tbl] = rows[:3]
    
    return result

@app.get("/api/sens-debug")
async def sens_debug():
    from scraper import _list_hrefs, _read_zip, _parse_aemo, PREDISPATCH_SENS_URL, get_latest_file_url
    import csv, io
    result = {}
    try:
        files = _list_hrefs(PREDISPATCH_SENS_URL)
        result['total_files'] = len(files)
        result['sample_files'] = files[-3:] if files else []
    except Exception as e:
        result['list_error'] = str(e)
        return result
    if not files:
        return result
    url = files[-1]
    result['fetching'] = url
    text = _read_zip(url)
    result['text_len'] = len(text)
    # List tables
    tables = {}
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if row and row[0].strip().upper() == 'I' and len(row) >= 3:
            tbl = row[2].strip().upper()
            cols = [c.strip() for c in row[4:] if c.strip()]
            tables[tbl] = cols[:20]
    result['tables'] = tables
    # Sample first NSW1 row from any sensitivity table
    for tbl_key in tables:
        rows = list(_parse_aemo(text, tbl_key))
        nsw = [r for r in rows if r.get('REGIONID','').strip() == 'NSW1'][:1]
        if nsw:
            result['nsw_sample_' + tbl_key] = nsw[0]
            break
    return result

# Cache for MTPASA calendar data (refreshed with slow cache)
_mtpasa_cal_cache: dict = {"data": None, "last_updated": None}

def _build_mtpasa_calendar() -> dict:
    """Blocking function — run in executor. Fetches and parses MTPASA data."""
    import requests, zipfile, io
    from scraper import _list_hrefs, _parse_aemo, NEM_UNITS, MTPASA_DUID_URL, AEST
    from datetime import datetime

    # Use a fresh session — the global SESSION is not thread-safe under concurrent load
    def _get_zip(url):
        try:
            s = requests.Session()
            s.headers.update({"User-Agent": "NEM-Dashboard/1.0"})
            r = s.get(url, timeout=30)
            r.raise_for_status()
            with zipfile.ZipFile(io.BytesIO(r.content)) as z:
                csvs = [n for n in z.namelist() if n.lower().endswith(".csv")]
                if not csvs: return ""
                with z.open(csvs[0]) as f:
                    return f.read().decode("utf-8", errors="replace")
        except Exception as e:
            logger.warning(f"_build_mtpasa_calendar: fetch failed {url}: {e}")
            return ""

    files = _list_hrefs(MTPASA_DUID_URL)
    if not files:
        return {}
    latest = sorted(files)[-1]
    text   = _get_zip(latest)
    if not text:
        return {}

    duid_days: dict = {}
    for row in _parse_aemo(text, "MTPASA_DUIDAVAILABILITY"):
        duid = row.get("DUID", "").strip()
        if not duid:
            continue
        unit = NEM_UNITS.get(duid, {})
        if not unit or not unit.get("capacity"):
            continue
        day   = row.get("DAY", "").strip()[:10]
        try:
            avail = round(float(row.get("PASAAVAILABILITY") or 0))
        except ValueError:
            avail = 0
        state = row.get("UNITSTATE", "").strip() or "Unknown"
        if duid not in duid_days:
            duid_days[duid] = {}
        duid_days[duid][day] = {"avail": avail, "state": state}

    result = {}
    for duid, days in duid_days.items():
        unit = NEM_UNITS.get(duid, {})
        result[duid] = {
            "station":  unit.get("station", duid),
            "fuel":     unit.get("fuel", "Other"),
            "region":   unit.get("region", ""),
            "capacity": int(unit.get("capacity") or 0),
            "days":     [{"day": k, "avail": v["avail"], "state": v["state"]}
                         for k, v in sorted(days.items())]
        }
    return {"source": latest, "as_of": datetime.now(AEST).isoformat(), "duids": result}

@app.get("/api/mtpasa-calendar")
async def mtpasa_calendar():
    """Return cached MTPASA calendar data. Fetches in background if stale."""
    import asyncio
    # Return cached data immediately if available
    if _mtpasa_cal_cache["data"]:
        return _mtpasa_cal_cache["data"]
    # First call: fetch in executor so we don't block the event loop
    loop = asyncio.get_running_loop()
    data = await loop.run_in_executor(None, _build_mtpasa_calendar)
    if data:
        _mtpasa_cal_cache["data"] = data
    return data or {"error": "Failed to fetch MTPASA data"}

# Cache for historical price averages — expensive to compute (~53 file fetches)
_price_avg_cache = {"data": None, "last_updated": None}

@app.get("/api/prices-rolling")
async def prices_rolling(days: int = 14):
    """
    Return daily price files from GitHub for the last N days.
    Each day: { region: { "HH:MM": rrp } }
    Returns: { "YYYY-MM-DD": { region: { "HH:MM": rrp } } }
    """
    import json, base64, os
    from datetime import datetime, timezone, timedelta
    import httpx

    AEST     = timezone(timedelta(hours=10))
    GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
    GH_REPO  = os.environ.get("GITHUB_REPO", "")
    GH_HEADERS = {
        "Authorization": f"token {GH_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }

    if not GH_TOKEN or not GH_REPO:
        return JSONResponse(content={"error": "GitHub not configured"})

    now_aest = datetime.now(AEST)
    dates = [
        (now_aest - timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range(days)
    ]

    result = {}
    async with httpx.AsyncClient(timeout=10) as client:
        for date_str in dates:
            path = f"data/prices/{date_str}.json"
            try:
                r = await client.get(
                    f"https://api.github.com/repos/{GH_REPO}/contents/{path}",
                    headers=GH_HEADERS,
                )
                if r.status_code == 200:
                    data = json.loads(base64.b64decode(r.json()["content"]).decode())
                    data.pop("date", None)
                    result[date_str] = data
            except Exception:
                pass

    return JSONResponse(content=result)


@app.get("/api/historical_price_averages")
async def historical_price_averages(refresh: bool = False):
    """Monthly average electricity prices from TradingIS archive (~12 months)."""
    import asyncio
    from datetime import datetime, timezone, timedelta
    from scraper import scrape_historical_price_averages

    # Serve cache if fresh (< 24h old) and not forced refresh
    if not refresh and _price_avg_cache["data"] and _price_avg_cache["last_updated"]:
        age = datetime.now(timezone.utc) - _price_avg_cache["last_updated"]
        if age < timedelta(hours=6):
            return JSONResponse(content=_price_avg_cache["data"])

    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, lambda: scrape_historical_price_averages(days=90)),
            timeout=300.0
        )
        _price_avg_cache["data"] = data
        _price_avg_cache["last_updated"] = datetime.now(timezone.utc)
        return JSONResponse(content=data)
    except asyncio.TimeoutError:
        return JSONResponse(status_code=504, content={"error": "timeout"})
    except Exception as e:
        logger.error(f"historical_price_averages error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/historical_day")
async def historical_day(date: str):
    """Fetch full day of historical data for D-1 page: prices, demand, fuel mix."""
    import re
    if not re.match(r'^\d{8}$', date):
        return JSONResponse(status_code=400, content={"error": "date must be YYYYMMDD"})
    from scraper import scrape_historical_day
    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_historical_day, date),
            timeout=120.0
        )
        return JSONResponse(content=data)
    except asyncio.TimeoutError:
        return JSONResponse(status_code=504, content={"error": "timeout"})
    except Exception as e:
        logger.error(f"historical_day error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/api/historical_dispatch_prices")
async def historical_dispatch_prices(date: str):
    """Fetch 5-min dispatch prices for a given date (YYYYMMDD)."""
    import re
    if not re.match(r'^\d{8}$', date):
        return JSONResponse(status_code=400, content={"error": "date must be YYYYMMDD"})
    from scraper import scrape_historical_dispatch_prices
    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_historical_dispatch_prices, date),
            timeout=60.0
        )
        return JSONResponse(content=data)
    except asyncio.TimeoutError:
        return JSONResponse(status_code=504, content={"error": "timeout fetching dispatch prices"})
    except Exception as e:
        logger.error(f"historical_dispatch_prices error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/origin")
async def origin_history():
    """Return Origin asset history — separate from fast cache to keep /api/data lean."""
    from scraper import _duid_history, ORIGIN_DUIDS
    from datetime import datetime
    from zoneinfo import ZoneInfo
    now_label = datetime.now(ZoneInfo("Australia/Brisbane")).strftime("%H:%M")
    result = {}
    for duid in ORIGIN_DUIDS:
        history = _duid_history.get(duid, {})
        if history:
            result[duid] = [{"interval": k, "mw": v} for k, v in sorted(history.items()) if k <= now_label]
    return JSONResponse(content=result)


@app.get("/api/price-avg-debug")
async def price_avg_debug():
    """Debug endpoint — tests each step of scrape_historical_price_averages."""
    import time
    from scraper import (TRADING_ARCHIVE, _list_hrefs, _read_zip_all,
                         _parse_aemo, NEM_REGIONS, AEST)
    from datetime import datetime, timedelta

    results = {}
    now_aest = datetime.now(AEST)
    cutoff   = (now_aest - timedelta(days=30)).date()
    loop     = asyncio.get_running_loop()

    # Step 1: list archive
    t0 = time.time()
    try:
        hrefs = await loop.run_in_executor(None, _list_hrefs, TRADING_ARCHIVE)
        relevant = []
        for href in hrefs:
            fname = href.split('/')[-1]
            parts = fname.replace('.zip','').split('_')
            if len(parts) >= 4:
                try:
                    end_date = datetime.strptime(parts[-1], "%Y%m%d").date()
                    if end_date >= cutoff:
                        relevant.append(href)
                except ValueError:
                    pass
        results["step1_listing"] = {
            "total_hrefs": len(hrefs),
            "relevant": len(relevant),
            "relevant_files": [h.split('/')[-1] for h in relevant],
            "ms": round((time.time()-t0)*1000)
        }
    except Exception as e:
        results["step1_listing"] = {"error": str(e)}
        return JSONResponse(content=results)

    if not relevant:
        results["error"] = "no relevant ZIPs"
        return JSONResponse(content=results)

    # Step 2: download most recent ZIP and inspect
    t0 = time.time()
    test_url = relevant[-1]
    try:
        import zipfile as _zf, io as _io
        from scraper import _get, _read_zip_of_zips
        r = await loop.run_in_executor(None, lambda: _get(test_url, timeout=60))
        if not r:
            results["step2_zip"] = {"error": "no response", "url": test_url}
            return JSONResponse(content=results)
        raw_bytes = len(r.content)
        # Inspect ZIP contents
        zip_files = []
        try:
            with _zf.ZipFile(_io.BytesIO(r.content)) as z:
                zip_files = z.namelist()
        except Exception as ze:
            results["step2_zip"] = {"error": f"zip open failed: {ze}", "url": test_url, "bytes": raw_bytes}
            return JSONResponse(content=results)

        results["step2_zip"] = {
            "url": test_url,
            "bytes": raw_bytes,
            "zip_file_count": len(zip_files),
            "zip_files_sample": zip_files[:5],
            "ms": round((time.time()-t0)*1000)
        }
    except Exception as e:
        results["step2_zip"] = {"error": str(e)}
        return JSONResponse(content=results)

    # Step 3: read one inner ZIP and parse
    t0 = time.time()
    try:
        text = await loop.run_in_executor(None, _read_zip_of_zips, test_url)
        lines = text.split('\n') if text else []
        d_count = sum(1 for l in lines if l.startswith('D,'))
        rows = list(_parse_aemo(text, "TRADING_PRICE")) if text else []
        results["step3_parse"] = {
            "kb": len(text)//1024 if text else 0,
            "total_lines": len(lines),
            "d_rows": d_count,
            "trading_price_rows": len(rows),
            "sample": rows[:2],
            "ms": round((time.time()-t0)*1000)
        }
    except Exception as e:
        results["step3_parse"] = {"error": str(e)}

    return JSONResponse(content=results)






@app.get("/api/backfill-prices")
async def backfill_prices(days: int = 90, secret: str = ""):
    """
    One-off backfill: fetch TradingIS archive + CURRENT files and write
    daily price files to GitHub. Runs as a background task.
    Protect with ?secret=yourtoken to avoid accidental triggers.
    """
    import os
    if secret != os.environ.get("BACKFILL_SECRET", "backfill"):
        return JSONResponse(status_code=403, content={"error": "wrong secret"})

    async def _run():
        import json, base64, zipfile as _zf, csv as _csv, io as _io, time as _time, re as _re
        from datetime import timedelta
        from scraper import (TRADING_ARCHIVE, NEMWEB_BASE, NEM_REGIONS, AEST,
                             _list_hrefs, _read_zip_of_zips, _read_zip, _parse_aemo)
        import httpx

        now_aest = datetime.now(timezone(timedelta(hours=10)))
        cutoff   = (now_aest - timedelta(days=days)).date()
        today    = now_aest.date()
        GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
        GH_REPO  = os.environ.get("GITHUB_REPO", "")
        GH_HEADERS = {
            "Authorization": f"token {GH_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
        }

        if not GH_TOKEN or not GH_REPO:
            logger.error("backfill: GitHub not configured")
            return

        logger.info(f"backfill: starting {cutoff} → {today}")
        daily: dict = {}
        loop = asyncio.get_running_loop()

        # ── Phase 1: weekly archive ZIPs ─────────────────────────────────────
        logger.info("backfill: Phase 1 — listing TradingIS archive")
        all_hrefs = await loop.run_in_executor(None, _list_hrefs, TRADING_ARCHIVE)
        relevant = []
        for href in all_hrefs:
            fname = href.split('/')[-1]
            parts = fname.replace('.zip', '').split('_')
            if len(parts) >= 4:
                try:
                    sd = datetime.strptime(parts[-2], "%Y%m%d").date()
                    ed = datetime.strptime(parts[-1], "%Y%m%d").date()
                    if ed >= cutoff and sd < today:
                        relevant.append((sd, ed, href))
                except ValueError:
                    pass
        relevant.sort()
        logger.info(f"backfill: {len(relevant)} weekly ZIPs to fetch")

        archive_end = cutoff
        for i, (sd, ed, url) in enumerate(relevant):
            logger.info(f"backfill: [{i+1}/{len(relevant)}] {url.split('/')[-1]}")
            text = await loop.run_in_executor(None, _read_zip_of_zips, url)
            if not text:
                logger.warning(f"backfill: skip {url.split('/')[-1]} — no data")
                continue
            rows = _parse_aemo(text, "TRADING_PRICE")
            logger.info(f"backfill:   {len(rows)} rows")
            for row in rows:
                region = row.get("REGIONID", "").strip()
                if region not in NEM_REGIONS: continue
                if row.get("INVALIDFLAG", "0") not in ("0", ""): continue
                dt_str = row.get("SETTLEMENTDATE", "")
                rrp_str = row.get("RRP", "")
                if not dt_str or not rrp_str: continue
                try:
                    dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                    dd = dt.date()
                    if dd < cutoff or dd >= today: continue
                    daily.setdefault(dt.strftime("%Y-%m-%d"), {}).setdefault(region, {})[dt.strftime("%H:%M")] = round(float(rrp_str), 2)
                    if dd > archive_end: archive_end = dd
                except (ValueError, TypeError): continue
            logger.info(f"backfill:   {len(daily)} days accumulated")

        logger.info(f"backfill: archive covers up to {archive_end}")

        # ── Phase 2: CURRENT files for gap ───────────────────────────────────
        gap_start = archive_end + timedelta(days=1)
        gap_end   = today - timedelta(days=1)

        if gap_start <= gap_end:
            logger.info(f"backfill: Phase 2 — CURRENT files {gap_start} → {gap_end}")
            TRADING_CURRENT = f"{NEMWEB_BASE}/REPORTS/CURRENT/TradingIS_Reports/"
            current_hrefs = await loop.run_in_executor(None, _list_hrefs, TRADING_CURRENT)
            seen: dict = {}
            for href in current_hrefs:
                fname = href.split('/')[-1]
                parts = fname.split('_')
                if len(parts) >= 3 and len(parts[2]) >= 8:
                    try:
                        fd = datetime.strptime(parts[2][:8], "%Y%m%d").date()
                        if gap_start <= fd <= gap_end:
                            seen[parts[2]] = href
                    except ValueError: pass
            gap_urls = sorted(seen.values())
            logger.info(f"backfill: {len(gap_urls)} CURRENT files for gap")

            ok = fail = 0
            for j, url in enumerate(gap_urls):
                try:
                    def _fetch_one(u=url):
                        import time as _t
                        _t.sleep(2)
                        return _read_zip(u)
                    text = await loop.run_in_executor(None, _fetch_one)
                    if text:
                        rows = _parse_aemo(text, "TRADING_PRICE")
                        for row in rows:
                            region = row.get("REGIONID", "").strip()
                            if region not in NEM_REGIONS: continue
                            if row.get("INVALIDFLAG", "0") not in ("0", ""): continue
                            dt_str = row.get("SETTLEMENTDATE", "")
                            rrp_str = row.get("RRP", "")
                            if not dt_str or not rrp_str: continue
                            try:
                                dt = datetime.fromisoformat(dt_str.replace("/", "-")) - timedelta(minutes=30)
                                daily.setdefault(dt.strftime("%Y-%m-%d"), {}).setdefault(row.get("REGIONID","").strip(), {})[dt.strftime("%H:%M")] = round(float(rrp_str), 2)
                            except (ValueError, TypeError): continue
                        ok += 1
                    else:
                        fail += 1
                    if j % 50 == 0:
                        logger.info(f"backfill: CURRENT [{j+1}/{len(gap_urls)}] ok={ok} fail={fail} days={len(daily)}")
                except Exception as e:
                    logger.warning(f"backfill: CURRENT error: {e}")
                    fail += 1
            logger.info(f"backfill: Phase 2 done ok={ok} fail={fail}")

        # ── Phase 3: upload to GitHub ─────────────────────────────────────────
        dates = sorted(daily.keys())
        logger.info(f"backfill: Phase 3 — uploading {len(dates)} files to GitHub")

        async with httpx.AsyncClient(timeout=15) as client:
            for j, date_str in enumerate(dates):
                path = f"data/prices/{date_str}.json"
                day_data = {"date": date_str, **daily[date_str]}
                try:
                    # Read existing
                    r = await client.get(
                        f"https://api.github.com/repos/{GH_REPO}/contents/{path}",
                        headers=GH_HEADERS
                    )
                    sha = ""
                    if r.status_code == 200:
                        rj = r.json()
                        sha = rj.get("sha", "")
                        existing = json.loads(base64.b64decode(rj["content"]).decode())
                        for region, intervals in daily[date_str].items():
                            existing.setdefault(region, {}).update(intervals)
                        day_data = existing

                    encoded = base64.b64encode(json.dumps(day_data, separators=(',',':')).encode()).decode()
                    payload = {"message": f"backfill: {date_str}", "content": encoded}
                    if sha: payload["sha"] = sha
                    wr = await client.put(
                        f"https://api.github.com/repos/{GH_REPO}/contents/{path}",
                        headers=GH_HEADERS, json=payload
                    )
                    status = "ok" if wr.status_code in (200,201) else f"err={wr.status_code}"
                    if j % 10 == 0:
                        logger.info(f"backfill: [{j+1}/{len(dates)}] {date_str} {status}")
                    await asyncio.sleep(0.5)
                except Exception as e:
                    logger.warning(f"backfill: upload {date_str} failed: {e}")

        logger.info(f"backfill: COMPLETE — {len(dates)} days uploaded")

    asyncio.create_task(_run())
    return JSONResponse(content={
        "status": "backfill started",
        "days": days,
        "message": "Watch Render logs for progress. Will take 30-120 minutes."
    })


# Cache for gas data — refreshed on demand, valid for 6 hours
_gas_cache = {"data": None, "last_updated": None}

@app.get("/api/gas")
async def gas_data(refresh: bool = False):
    """Return gas market data. Fetched on demand, cached for 6 hours."""
    from scraper import scrape_gas
    from datetime import timezone, timedelta

    # Serve cache if fresh
    if not refresh and _gas_cache["data"] and _gas_cache["last_updated"]:
        age = datetime.now(timezone.utc) - _gas_cache["last_updated"]
        if age < timedelta(hours=6):
            return JSONResponse(content=_gas_cache["data"])

    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_gas),
            timeout=60.0
        )
        _gas_cache["data"] = data
        _gas_cache["last_updated"] = datetime.now(timezone.utc)
        return JSONResponse(content=data)
    except asyncio.TimeoutError:
        if _gas_cache["data"]:
            return JSONResponse(content=_gas_cache["data"])
        return JSONResponse(status_code=504, content={"error": "timeout"})
    except Exception as e:
        logger.error(f"gas_data error: {e}")
        if _gas_cache["data"]:
            return JSONResponse(content=_gas_cache["data"])
        return JSONResponse(status_code=500, content={"error": str(e)})


_gbb_cache: dict = {"data": None, "last_updated": None}

@app.get("/api/gbb")
async def gbb_data(refresh: bool = False):
    """Return GBB storage levels and state summary. Cached for 1 hour."""
    from scraper import scrape_gbb
    from datetime import timezone, timedelta

    if not refresh and _gbb_cache["data"] and _gbb_cache["last_updated"]:
        age = datetime.now(timezone.utc) - _gbb_cache["last_updated"]
        if age < timedelta(hours=1):
            return JSONResponse(content=_gbb_cache["data"])

    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_gbb),
            timeout=30.0
        )
        _gbb_cache["data"] = data
        _gbb_cache["last_updated"] = datetime.now(timezone.utc)
        return JSONResponse(content=data)
    except asyncio.TimeoutError:
        if _gbb_cache["data"]:
            return JSONResponse(content=_gbb_cache["data"])
        return JSONResponse(status_code=504, content={"error": "timeout"})
    except Exception as e:
        logger.error(f"gbb_data error: {e}")
        if _gbb_cache["data"]:
            return JSONResponse(content=_gbb_cache["data"])
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/gas-debug")
async def gas_debug():
    """Inspect STTM and VicGas CSV contents in detail."""
    import zipfile as _zf, io as _io, csv as _csv
    from scraper import STTM_BASE, VICGAS_BASE, _get

    loop = asyncio.get_running_loop()

    def _inspect():
        out = {}
        # STTM CURRENTDAY — int676 all rows, int678 all rows
        url = f"{STTM_BASE}/CURRENTDAY.ZIP"
        r = _get(url, timeout=30)
        if r:
            with _zf.ZipFile(_io.BytesIO(r.content)) as z:
                out["sttm_files"] = z.namelist()
                for prefix, key in [("int676", "int676_rows"), ("int678", "int678_rows"), ("int651", "int651_rows")]:
                    matches = [n for n in z.namelist() if prefix in n.lower()]
                    if matches:
                        with z.open(matches[0]) as f:
                            out[key] = list(_csv.DictReader(_io.TextIOWrapper(f, errors="replace")))
        # VicGas INT041 all rows
        r041 = _get(f"{VICGAS_BASE}/INT041_V4_MARKET_AND_REFERENCE_PRICES_1.CSV", timeout=15)
        if r041:
            out["vicgas_int041_rows"] = list(_csv.DictReader(_io.StringIO(r041.text)))
        # VicGas INT037B first 5 lines
        r037 = _get(f"{VICGAS_BASE}/INT037B_V4_INDICATIVE_MKT_PRICE_1.CSV", timeout=15)
        if r037:
            out["vicgas_int037b_first5"] = r037.text.splitlines()[:5]
        # VicGas INT050 all rows
        r050 = _get(f"{VICGAS_BASE}/INT050_V4_SCHED_WITHDRAWALS_1.CSV", timeout=15)
        if r050:
            out["vicgas_int050_rows"] = list(_csv.DictReader(_io.StringIO(r050.text)))
        return out

    result = await loop.run_in_executor(None, _inspect)
    return JSONResponse(content=result)


@app.get("/api/gbb-debug")
async def gbb_debug():
    """Check nomination demand types and pipeline direction data."""
    import csv as _csv
    from scraper import _get
    loop = asyncio.get_running_loop()

    def _inspect():
        out = {}
        # 1. Nominations file - what demand exists?
        nom_r = _get("https://www.nemweb.com.au/Reports/Current/GBB/GasBBNominationAndForecastNext7.CSV", timeout=20)
        if nom_r:
            nom_rows = list(_csv.DictReader(nom_r.text.splitlines()))
            # Show demand by facility type for first nomination date
            nom_dates = sorted({row.get("Gasdate","") for row in nom_rows if row.get("Gasdate","")})
            out["nomination_dates"] = nom_dates
            first_date = nom_dates[0] if nom_dates else ""
            by_type = {}
            for row in nom_rows:
                if row.get("Gasdate","") != first_date: continue
                ft = row.get("FacilityType","")
                by_type.setdefault(ft, {"count":0,"supply":0.0,"demand":0.0})
                by_type[ft]["count"] += 1
                try:
                    by_type[ft]["supply"] += float(row.get("Supply") or 0)
                    by_type[ft]["demand"] += float(row.get("Demand") or 0)
                except: pass
            out["nom_first_date_by_type"] = {k:{"count":v["count"],"supply":round(v["supply"],1),"demand":round(v["demand"],1)} for k,v in by_type.items()}

        # 2. Actuals - show EGP pipeline rows with TransferIn/Out for direction
        act_r = _get("https://www.nemweb.com.au/Reports/Current/GBB/GasBBActualFlowStorageLast31.CSV", timeout=30)
        if act_r:
            act_rows = list(_csv.DictReader(act_r.text.splitlines()))
            all_dates = sorted({r["GasDate"] for r in act_rows})
            latest = all_dates[-2]  # use second-to-last (complete day)
            # Show key interstate pipeline rows
            key_pipes = ["EGP", "TGP", "MSP", "SWQP", "VTS", "MAPS", "RBP"]
            pipe_rows = [r for r in act_rows if r["GasDate"] == latest and r["FacilityName"] in key_pipes and r["FacilityType"] == "PIPE"]
            out["key_pipe_rows_latest"] = [{"pipe": r["FacilityName"], "loc": r["LocationName"],
                "state": r["State"], "supply": r["Supply"], "demand": r["Demand"],
                "transferIn": r["TransferIn"], "transferOut": r["TransferOut"]} for r in pipe_rows]

            # Check april 5 data completeness
            apr5 = "2026/04/05"
            apr5_rows = [r for r in act_rows if r["GasDate"] == apr5]
            if apr5_rows:
                by_type = {}
                for r in apr5_rows:
                    ft = r["FacilityType"]
                    by_type.setdefault(ft, {"count":0,"supply":0.0})
                    by_type[ft]["count"] += 1
                    try: by_type[ft]["supply"] += float(r.get("Supply") or 0)
                    except: pass
                out["apr5_by_type"] = {k:{"count":v["count"],"supply":round(v["supply"],1)} for k,v in by_type.items()}
            else:
                out["apr5_by_type"] = "not in file"

        return out

    result = await loop.run_in_executor(None, _inspect)
    return JSONResponse(content=result)





async def gas_excel_debug():
    """Inspect AEMO DWGM Excel file structure."""
    import io as _io, openpyxl as _xl
    from scraper import _get
    loop = asyncio.get_running_loop()

    def _inspect():
        out = {}
        try:
            url = "https://www.aemo.com.au/-/media/files/gas/dwgm/dwgm-prices-and-demand.xlsx"
            r = _get(url, timeout=60)
            if not r:
                return {"error": "fetch failed"}
            out["bytes"] = len(r.content)
            wb = _xl.load_workbook(_io.BytesIO(r.content), read_only=True, data_only=True)
            out["sheet_names"] = wb.sheetnames
            for ws in wb.worksheets[:3]:
                first_rows, last_rows = [], []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    r2 = [str(c) if c is not None else None for c in row[:6]]
                    if i < 4:
                        first_rows.append(r2)
                    last_rows.append(r2)
                    if len(last_rows) > 5:
                        last_rows.pop(0)
                out[ws.title] = {"first": first_rows, "last": last_rows}
            wb.close()
        except Exception as e:
            import traceback
            out["error"] = str(e)
            out["traceback"] = traceback.format_exc()
        return out

    result = await loop.run_in_executor(None, _inspect)
    return JSONResponse(content=result)


@app.get("/api/weather-debug")
async def weather_debug():
    """Test weather scraping directly and return raw result."""
    from scraper import scrape_weather
    loop = asyncio.get_running_loop()
    try:
        data = await asyncio.wait_for(
            loop.run_in_executor(None, scrape_weather),
            timeout=30.0
        )
        return JSONResponse(content={"ok": True, "data": data})
    except Exception as e:
        return JSONResponse(content={"ok": False, "error": str(e)})


@app.get("/api/pd-debug")
async def pd_debug():
    """Find where AEMO publishes unit-level dispatch forecasts."""
    from scraper import _list_hrefs, _read_zip, NEMWEB_BASE
    import csv, io, re
    loop = asyncio.get_running_loop()

    dirs_to_probe = [
        f"{NEMWEB_BASE}/Reports/CURRENT/P5_Reports/",
        f"{NEMWEB_BASE}/Reports/CURRENT/PredispatchIS_Reports/",
        f"{NEMWEB_BASE}/Reports/CURRENT/DispatchIS_Reports/",
    ]

    results = {}
    for dir_url in dirs_to_probe:
        label = dir_url.split("/CURRENT/")[-1].rstrip("/")
        hrefs = await loop.run_in_executor(None, _list_hrefs, dir_url)
        # Group by prefix
        by_prefix = {}
        for h in hrefs:
            p = re.sub(r'_\d{8,}.*', '', h.split('/')[-1])
            by_prefix.setdefault(p, []).append(h)

        prefix_tables = {}
        for prefix, files in by_prefix.items():
            try:
                text = await asyncio.wait_for(
                    loop.run_in_executor(None, _read_zip, files[-1]), timeout=15)
                keys = []
                for row in csv.reader(io.StringIO(text or "")):
                    if row and row[0].strip().upper() == "I" and len(row) >= 3:
                        key = f"{row[1].strip()}_{row[2].strip()}".upper()
                        if key not in keys:
                            keys.append(key)
                prefix_tables[prefix] = {"tables": keys, "files": len(files)}
            except Exception as e:
                prefix_tables[prefix] = {"error": str(e), "files": len(files)}

        results[label] = prefix_tables

    return results


@app.get("/api/station-debug")
async def station_debug():
    from scraper import _duid_history
    total_duids = len(_duid_history)
    sample = {}
    for duid in list(_duid_history.keys())[:10]:
        pts = len(_duid_history[duid])
        last = sorted(_duid_history[duid].keys())[-1] if _duid_history[duid] else None
        sample[duid] = {"pts": pts, "last": last}
    return JSONResponse(content={
        "total_duids": total_duids,
        "sample": sample,
    })


@app.api_route("/api/views", methods=["GET", "POST"])
async def record_view(request: Request):
    import json, hashlib, os, base64
    from datetime import datetime, timezone, timedelta
    import httpx

    AEST      = timezone(timedelta(hours=10))
    now_aest  = datetime.now(AEST)
    today     = now_aest.strftime("%Y-%m-%d")
    month     = now_aest.strftime("%Y-%m")

    forwarded = request.headers.get("x-forwarded-for")
    raw_ip    = forwarded.split(",")[0].strip() if forwarded else (request.client.host if request.client else "unknown")
    ip_hash   = hashlib.sha256(raw_ip.encode()).hexdigest()[:16]

    GH_TOKEN = os.environ.get("GITHUB_TOKEN", "")
    GH_REPO  = os.environ.get("GITHUB_REPO", "")   # e.g. "danielteng/nem-dashboard"
    GH_PATH  = "data/views.json"
    GH_HEADERS = {
        "Authorization": f"token {GH_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }

    # ── Load current data from repo file ─────────────────────────────────────
    data: dict = {}
    file_sha: str = ""
    if not GH_TOKEN:
        logger.warning("views: GITHUB_TOKEN not set")
    if not GH_REPO:
        logger.warning("views: GITHUB_REPO not set")
    if GH_TOKEN and GH_REPO:
        try:
            async with httpx.AsyncClient(timeout=8) as client:
                r = await client.get(
                    f"https://api.github.com/repos/{GH_REPO}/contents/{GH_PATH}",
                    headers=GH_HEADERS,
                )
                logger.info(f"views: GET {GH_PATH} status={r.status_code}")
                if r.status_code == 200:
                    resp_json = r.json()
                    file_sha  = resp_json.get("sha", "")
                    raw       = base64.b64decode(resp_json["content"]).decode()
                    data      = json.loads(raw)
                else:
                    logger.warning(f"views: GET failed body={r.text[:200]}")
        except Exception as e:
            logger.error(f"views: load error: {e}")

    # ── Update counts ────────────────────────────────────────────────────────
    data.setdefault("total", 0)
    data.setdefault("by_day", {})
    data.setdefault("by_month", {})
    data.setdefault("unique_ips", [])
    data.setdefault("unique_by_day", {})
    data.setdefault("unique_by_month", {})

    data["total"] += 1
    data["by_day"][today]   = data["by_day"].get(today, 0) + 1
    data["by_month"][month] = data["by_month"].get(month, 0) + 1

    if ip_hash not in data["unique_ips"]:
        data["unique_ips"].append(ip_hash)
    if ip_hash not in data["unique_by_day"].setdefault(today, []):
        data["unique_by_day"][today].append(ip_hash)
    if ip_hash not in data["unique_by_month"].setdefault(month, []):
        data["unique_by_month"][month].append(ip_hash)

    # Prune daily buckets older than 60 days
    if len(data["by_day"]) > 60:
        for old in sorted(data["by_day"].keys())[:-60]:
            data["by_day"].pop(old, None)
            data["unique_by_day"].pop(old, None)

    # ── Write back to repo file ───────────────────────────────────────────────
    if GH_TOKEN and GH_REPO:
        try:
            encoded = base64.b64encode(json.dumps(data, indent=2).encode()).decode()
            payload = {
                "message": f"views: {today} total={data['total']}",
                "content": encoded,
            }
            if file_sha:
                payload["sha"] = file_sha
            async with httpx.AsyncClient(timeout=8) as client:
                r = await client.put(
                    f"https://api.github.com/repos/{GH_REPO}/contents/{GH_PATH}",
                    headers=GH_HEADERS,
                    json=payload,
                )
                logger.info(f"views: PUT {GH_PATH} status={r.status_code}")
                if r.status_code not in (200, 201):
                    logger.warning(f"views: PUT failed body={r.text[:200]}")
        except Exception as e:
            logger.error(f"views: write error: {e}")

    # Build unique_by_month counts for frontend
    unique_by_month_counts = {m: len(ips) for m, ips in data.get("unique_by_month", {}).items()}

    return {
        "total":            data["total"],
        "today":            data["by_day"].get(today, 0),
        "this_month":       data["by_month"].get(month, 0),
        "unique_total":     len(data["unique_ips"]),
        "unique_today":     len(data["unique_by_day"].get(today, [])),
        "unique_month":     len(data["unique_by_month"].get(month, [])),
        "by_month":         data.get("by_month", {}),
        "unique_by_month":  unique_by_month_counts,
    }


@app.get("/", response_class=HTMLResponse)
async def dashboard():
    html_path = Path(__file__).parent / "static" / "index.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text())
    return HTMLResponse(content="<h1>Loading...</h1>")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
